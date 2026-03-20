"""
Publish batch service.

Runs the publish gate check and imports validated Excel data into planning tables.
Called from POST /api/batches/{id}/publish.

Gate conditions (raises ValueError if not met):
  1. Batch exists and is not already PUBLISHED
  2. All 6 required batch files are present with no BLOCKED issues
  3. All 4 masterdata types have at least one successful upload

On success:
  - Any existing PUBLISHED batch is archived (status → ARCHIVED)
  - Planning tables are cleared for this batch and re-imported from Excel
  - Batch status set to PUBLISHED with published_at / published_by

Data import notes:
  - master_stock: snapshot_date = batch.plan_cycle_date
  - demand_plan: one DB row per (item × month column); period_end = last day of month
  - line_capacity_calendar: standard_hours = planned_hours (Phase 1 simplification)
  - free_stock_ea is clamped to ≥ 0 (DB CHECK constraint; negative = back-order, treated as 0)
"""

import calendar as _cal
from datetime import date

import openpyxl
import pyodbc

from app.services.excel_utils import get_headers, get_data_rows, to_date, to_decimal, to_bit
from app.services.validation_service import (
    FILE_SCHEMAS, REQUIRED_FILE_TYPES,
    month_col_to_date, _detect_month_columns,
)
from app.services.masterdata_service import VALID_MASTERDATA_TYPES


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def publish_batch(conn: pyodbc.Connection, batch_id: int, published_by: str | None) -> dict:
    """
    Publish a batch. Raises ValueError with a user-facing message on gate failure.
    Returns the updated batch dict on success.
    """
    # --- Gate checks ---
    failures = _check_publish_gate(conn, batch_id)
    if failures:
        raise ValueError("; ".join(failures))

    # --- Archive any existing PUBLISHED batch ---
    cursor = conn.cursor()
    cursor.execute(
        """
        UPDATE dbo.import_batches
        SET status = 'ARCHIVED'
        WHERE status = 'PUBLISHED' AND batch_id <> ?
        """,
        batch_id,
    )

    # --- Get batch plan_cycle_date for snapshot_date ---
    cursor.execute(
        "SELECT plan_cycle_date FROM dbo.import_batches WHERE batch_id = ?",
        batch_id,
    )
    plan_cycle_date: date = cursor.fetchone()[0]

    # --- Clear previous planning data for this batch (idempotent re-publish) ---
    _clear_planning_data(conn, batch_id)

    # --- Import each file ---
    files = _get_current_files(conn, batch_id)
    for file_type, stored_path in files.items():
        _import_file(conn, batch_id, file_type, stored_path, plan_cycle_date)

    # --- Set status to PUBLISHED ---
    cursor.execute(
        """
        UPDATE dbo.import_batches
        SET status = 'PUBLISHED',
            published_at = GETUTCDATE(),
            published_by = ?
        WHERE batch_id = ?
        """,
        published_by,
        batch_id,
    )
    conn.commit()

    # Return updated batch
    cursor.execute(
        """
        SELECT batch_id, batch_name, plan_cycle_date, status, notes,
               created_by, created_at, published_at, published_by
        FROM dbo.import_batches WHERE batch_id = ?
        """,
        batch_id,
    )
    row = cursor.fetchone()
    return {
        "batch_id": row[0],
        "batch_name": row[1],
        "plan_cycle_date": str(row[2]),
        "status": row[3],
        "notes": row[4],
        "created_by": row[5],
        "created_at": str(row[6]),
        "published_at": str(row[7]) if row[7] else None,
        "published_by": row[8],
    }


# ---------------------------------------------------------------------------
# Gate check
# ---------------------------------------------------------------------------

def _check_publish_gate(conn: pyodbc.Connection, batch_id: int) -> list[str]:
    """Returns list of failure messages. Empty list = gate passed."""
    failures: list[str] = []
    cursor = conn.cursor()

    # Check batch exists and isn't already published
    cursor.execute(
        "SELECT status FROM dbo.import_batches WHERE batch_id = ?", batch_id
    )
    row = cursor.fetchone()
    if not row:
        failures.append("Batch not found")
        return failures
    if row[0] == "PUBLISHED":
        failures.append("Batch is already published")
        return failures

    # Check all 6 required files are present with no BLOCKED issues
    cursor.execute(
        """
        SELECT file_type,
               MAX(CASE WHEN ivr.severity = 'BLOCKED' THEN 1 ELSE 0 END) AS has_blocked
        FROM dbo.import_batch_files ibf
        LEFT JOIN dbo.import_validation_results ivr ON ivr.batch_file_id = ibf.batch_file_id
        WHERE ibf.batch_id = ? AND ibf.is_current_version = 1
          AND ibf.file_type IN (
              'master_stock', 'demand_plan', 'line_capacity_calendar',
              'headcount_plan', 'portfolio_changes', 'production_orders'
          )
        GROUP BY ibf.file_type
        """,
        batch_id,
    )
    present = {r[0]: r[1] for r in cursor.fetchall()}

    missing = REQUIRED_FILE_TYPES - set(present.keys())
    if missing:
        failures.append(f"Missing required files: {', '.join(sorted(missing))}")

    blocked_files = [ft for ft, has_b in present.items() if has_b]
    if blocked_files:
        failures.append(
            f"BLOCKED issues on: {', '.join(sorted(blocked_files))} — resolve before publishing"
        )

    # Check all 4 masterdata types have at least one successful upload
    cursor.execute(
        "SELECT DISTINCT masterdata_type FROM dbo.masterdata_uploads",
    )
    uploaded_types = {r[0] for r in cursor.fetchall()}
    missing_md = VALID_MASTERDATA_TYPES - uploaded_types
    if missing_md:
        failures.append(
            f"Masterdata not uploaded: {', '.join(sorted(missing_md))}"
        )

    return failures


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _clear_planning_data(conn: pyodbc.Connection, batch_id: int) -> None:
    cursor = conn.cursor()
    for table in ("master_stock", "demand_plan", "line_capacity_calendar",
                  "headcount_plan", "plant_headcount_plan",
                  "portfolio_changes", "production_orders"):
        cursor.execute(f"DELETE FROM dbo.{table} WHERE batch_id = ?", batch_id)  # noqa: S608


def _get_current_files(conn: pyodbc.Connection, batch_id: int) -> dict[str, str]:
    """Return {file_type: stored_file_path} for current-version files."""
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT file_type, stored_file_path
        FROM dbo.import_batch_files
        WHERE batch_id = ? AND is_current_version = 1
          AND file_type IN (
              'master_stock', 'demand_plan', 'line_capacity_calendar',
              'headcount_plan', 'portfolio_changes', 'production_orders'
          )
        """,
        batch_id,
    )
    return {r[0]: r[1] for r in cursor.fetchall()}


def _load_workbook(stored_path: str, header_row: int):
    """Open workbook and return (ws, headers, data_rows)."""
    schema = FILE_SCHEMAS  # just for header_row / data_start_row lookup
    wb = openpyxl.load_workbook(stored_path, read_only=True, data_only=True)
    ws = wb.active
    headers = get_headers(ws, header_row)
    return ws, headers


def _import_file(
    conn: pyodbc.Connection,
    batch_id: int,
    file_type: str,
    stored_path: str,
    plan_cycle_date: date,
) -> None:
    schema = FILE_SCHEMAS[file_type]
    header_row = schema.get("header_row", 1)
    data_start_row = schema.get("data_start_row", header_row + 1)

    wb = openpyxl.load_workbook(stored_path, read_only=True, data_only=True)
    ws = wb.active
    headers = get_headers(ws, header_row)
    data_rows = get_data_rows(ws, headers, data_start_row)

    handlers = {
        "master_stock":           _import_master_stock,
        "demand_plan":            _import_demand_plan,
        "line_capacity_calendar": _import_line_capacity_calendar,
        "headcount_plan":         _import_headcount_plan,
        "portfolio_changes":      _import_portfolio_changes,
        "production_orders":      _import_production_orders,
    }
    handlers[file_type](conn, batch_id, headers, data_rows, plan_cycle_date)

    # Sheet 2 — plant support headcount (headcount_plan only)
    if file_type == "headcount_plan":
        sheet2 = next(
            (wb[s] for s in wb.sheetnames
             if 'plant' in s.lower() or 'support' in s.lower()),
            None,
        )
        if sheet2 is not None:
            headers2 = get_headers(sheet2, header_row=2)
            data_rows2 = get_data_rows(sheet2, headers2, start_row=3)
            _import_plant_headcount(conn, batch_id, data_rows2)


# ---------------------------------------------------------------------------
# Per-file import handlers
# ---------------------------------------------------------------------------

def _import_master_stock(conn, batch_id, headers, data_rows, plan_cycle_date):
    header_set = set(headers)
    cursor = conn.cursor()
    cursor.execute("SELECT item_code FROM dbo.items")
    valid_items = {str(r[0]).strip() for r in cursor.fetchall()}
    for row_num, row in data_rows:
        item_code = _str(row.get("material"))
        warehouse_code = _str(row.get("plant"))
        if not item_code or not warehouse_code:
            continue
        if item_code not in valid_items:
            continue

        total_stock = _dec(row.get("unrestrictedstock")) or 0.0
        # free_stock_ea has a DB CHECK >= 0; negative = back-order → clamp to 0
        free_stock_raw = _dec(row.get("unrestricted_-_sales")) or 0.0
        free_stock = max(0.0, free_stock_raw)
        safety_stock = _dec(row.get("safety_stock")) if "safety_stock" in header_set else None
        mrp_type = _str(row.get("mrp_type")) if "mrp_type" in header_set else None

        cursor.execute(
            """
            INSERT INTO dbo.master_stock
                (batch_id, warehouse_code, item_code, snapshot_date,
                 mrp_type, total_stock_ea, free_stock_ea, safety_stock_ea,
                 source_row_number)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            batch_id, warehouse_code, item_code, plan_cycle_date,
            mrp_type, total_stock, free_stock, safety_stock, row_num,
        )


def _import_demand_plan(conn, batch_id, headers, data_rows, plan_cycle_date):
    month_cols = _detect_month_columns(headers)
    cursor = conn.cursor()
    cursor.execute("SELECT item_code FROM dbo.items")
    valid_items = {str(r[0]).strip() for r in cursor.fetchall()}
    for row_num, row in data_rows:
        item_code = _str(row.get("material_id"))
        warehouse_code = _str(row.get("plant"))
        if not item_code or not warehouse_code:
            continue
        if item_code not in valid_items:
            continue

        for col in month_cols:
            period_start = month_col_to_date(col)
            if period_start is None:
                continue
            last_day = _cal.monthrange(period_start.year, period_start.month)[1]
            period_end = date(period_start.year, period_start.month, last_day)

            qty_raw = row.get(col)
            qty = _dec(qty_raw) if qty_raw is not None else 0.0
            if qty is None:
                qty = 0.0

            cursor.execute(
                """
                INSERT INTO dbo.demand_plan
                    (batch_id, warehouse_code, item_code, period_type,
                     period_start_date, period_end_date, demand_quantity,
                     source_row_number)
                VALUES (?, ?, ?, 'MONTHLY', ?, ?, ?, ?)
                """,
                batch_id, warehouse_code, item_code,
                period_start, period_end, qty, row_num,
            )


def _import_line_capacity_calendar(conn, batch_id, headers, data_rows, plan_cycle_date):
    header_set = set(headers)
    cursor = conn.cursor()
    for row_num, row in data_rows:
        line_code = _str(row.get("line_code"))
        cal_date = to_date(row.get("calendar_date"))
        if not line_code or cal_date is None:
            continue

        is_working = to_bit(row.get("is_working_day"), default=1)
        planned_hrs = _dec(row.get("planned_hours")) or 0.0
        # standard_hours = planned_hours for Phase 1 (net = planned - losses)
        maint = _dec(row.get("maintenance_hours")) if "maintenance_hours" in header_set else 0.0
        ph = _dec(row.get("public_holiday_hours")) if "public_holiday_hours" in header_set else 0.0
        downtime = _dec(row.get("planned_downtime_hours")) if "planned_downtime_hours" in header_set else 0.0
        other = _dec(row.get("other_loss_hours")) if "other_loss_hours" in header_set else 0.0
        notes = _str(row.get("notes")) if "notes" in header_set else None

        cursor.execute(
            """
            INSERT INTO dbo.line_capacity_calendar
                (batch_id, line_code, calendar_date, is_working_day,
                 standard_hours, planned_hours,
                 maintenance_hours, public_holiday_hours,
                 planned_downtime_hours, other_loss_hours,
                 notes, source_row_number)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            batch_id, line_code, cal_date, is_working,
            planned_hrs, planned_hrs,
            maint or 0.0, ph or 0.0, downtime or 0.0, other or 0.0,
            notes, row_num,
        )


def _import_plant_headcount(conn, batch_id, data_rows):
    cursor = conn.cursor()
    for row_num, row in data_rows:
        plant_code = _str(row.get("plant_code"))
        role_code  = _str(row.get("resource_type_code"))
        plan_dt    = to_date(row.get("plan_date"))
        if not plant_code or not role_code or plan_dt is None:
            continue
        headcount = _dec(row.get("planned_headcount")) or 0.0
        cursor.execute(
            """
            INSERT INTO dbo.plant_headcount_plan
                (batch_id, plant_code, resource_type_code, plan_date,
                 planned_headcount, source_row_number)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            batch_id, plant_code, role_code, plan_dt, headcount, row_num,
        )


def _import_headcount_plan(conn, batch_id, headers, data_rows, plan_cycle_date):
    header_set = set(headers)
    cursor = conn.cursor()
    for row_num, row in data_rows:
        line_code = _str(row.get("line_code"))
        plan_dt = to_date(row.get("plan_date"))
        if not line_code or plan_dt is None:
            continue

        headcount = _dec(row.get("planned_headcount")) or 0.0
        shift_code = _str(row.get("shift_code")) if "shift_code" in header_set else None
        avail_hrs = _dec(row.get("available_hours")) if "available_hours" in header_set else None
        notes = _str(row.get("notes")) if "notes" in header_set else None

        cursor.execute(
            """
            INSERT INTO dbo.headcount_plan
                (batch_id, line_code, plan_date, shift_code,
                 planned_headcount, available_hours, notes, source_row_number)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            batch_id, line_code, plan_dt, shift_code,
            headcount, avail_hrs, notes, row_num,
        )


def _import_portfolio_changes(conn, batch_id, headers, data_rows, plan_cycle_date):
    """0 data rows is valid — nothing to import."""
    header_set = set(headers)
    cursor = conn.cursor()
    for row_num, row in data_rows:
        change_type = _str(row.get("change_type"))
        effective_dt = to_date(row.get("effective_date")) if "effective_date" in header_set else None
        item_code = _str(row.get("item_code")) if "item_code" in header_set else None
        description = _str(row.get("description")) if "description" in header_set else None
        impact_notes = _str(row.get("impact_notes")) if "impact_notes" in header_set else None
        # initial_demand: only meaningful for NEW_LAUNCH rows; stored as-is for all rows
        initial_demand = _dec(row.get("initial_demand")) if "initial_demand" in header_set else None

        cursor.execute(
            """
            INSERT INTO dbo.portfolio_changes
                (batch_id, item_code, change_type, effective_date,
                 description, impact_notes, initial_demand, source_row_number)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            batch_id, item_code, change_type, effective_dt,
            description, impact_notes, initial_demand, row_num,
        )


def _import_production_orders(conn, batch_id, headers, data_rows, plan_cycle_date):
    header_set = set(headers)
    cursor = conn.cursor()
    cursor.execute("SELECT item_code FROM dbo.items")
    valid_items = {str(r[0]).strip() for r in cursor.fetchall()}
    for row_num, row in data_rows:
        sap_order = _str(row.get("order"))
        item_code = _str(row.get("material"))
        plant_code = _str(row.get("plant"))
        if not sap_order or not item_code or not plant_code:
            continue
        if item_code not in valid_items:
            continue

        order_type = _str(row.get("order_type"))
        mrp_controller = _str(row.get("mrp_controller")) if "mrp_controller" in header_set else None
        order_qty = _dec(row.get("order_quantity_(gmein)")) or 0.0
        delivered_qty = _dec(row.get("delivered_quantity_(gmein)")) or 0.0
        net_qty = max(0.0, order_qty - delivered_qty)
        uom = _str(row.get("unit_of_measure_(=gmein)")) if "unit_of_measure_(=gmein)" in header_set else None
        start_dt = to_date(row.get("basic_start_date"))
        system_status = _str(row.get("system_status")) if "system_status" in header_set else None
        production_line = _str(row.get("production_line")) if "production_line" in header_set else None

        cursor.execute(
            """
            INSERT INTO dbo.production_orders
                (batch_id, sap_order_number, item_code, order_type,
                 mrp_controller, plant_code, order_quantity, delivered_quantity,
                 net_quantity, uom, basic_start_date, system_status,
                 production_line, source_row_number)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            batch_id, sap_order, item_code, order_type,
            mrp_controller, plant_code, order_qty, delivered_qty,
            net_qty, uom, start_dt, system_status,
            production_line, row_num,
        )


# ---------------------------------------------------------------------------
# Mini coercion helpers (publish-side, avoids re-importing full masterdata helpers)
# ---------------------------------------------------------------------------

def _str(val) -> str | None:
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return None
    return str(val).strip()


def _dec(val) -> float | None:
    return to_decimal(val)
