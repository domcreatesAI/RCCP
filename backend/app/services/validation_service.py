"""
7-stage validation pipeline for RCCP planning data files, plus stage 8 cross-file checks.

Stages:
  1. REQUIRED_FILE_CHECK        — batch-level: all 6 required files present?
  2. TEMPLATE_STRUCTURE_CHECK   — file opens as valid Excel with a header row
  3. FIELD_MAPPING_CHECK        — required columns present with correct names
  4. DATA_TYPE_CHECK            — correct types per column
  5. REFERENCE_CHECK            — FK values exist in masterdata tables
  6. BUSINESS_RULE_CHECK        — domain-specific business rules
  7. BATCH_READINESS            — overall batch can-publish status
  8. CROSS_FILE_CHECK           — cross-file consistency (WARNING only, never blocks publish)

master_stock: the authoritative per-batch SKU list (fresh SAP export). No material FK check —
  it IS the reference. plant → dbo.warehouses is BLOCKED.

production_orders / demand_plan: material codes are validated against master_stock (stage 8,
  WARNING). master_stock must be uploaded first for this check to run. plant → dbo.warehouses
  remains BLOCKED (stage 5).

demand_plan: wide-format SAP PIR export (material_id x month). Stages 3-6 fully implemented.
  Row 1 = descriptions (ignored), Row 2 = headers, Row 3+ = data.

production_orders: SAP COOIS export. mrp_controller is now required.

line_capacity_calendar: maintenance_hours is now required.

headcount_plan: available_hours is now required (>= 0).

portfolio_changes (phase-in): item_code + line_code required (the SKU and the line it is phased in on).

Stage 8 cross-file checks (WARNING only):
  master_stock:      SKUs not present in demand_plan or production_orders
  headcount_plan:    (line, date) pairs in line_capacity_calendar missing from headcount_plan
  portfolio_changes: phase-in SKUs absent from the production plan (production_orders)
"""

import re
from datetime import date, datetime
from decimal import Decimal

import openpyxl
import pyodbc

from app.services.excel_utils import (
    find_header_row as _find_header_row,
    get_headers as _get_headers,
    get_data_rows as _get_data_rows,
    is_valid_date as _is_valid_date,
    is_valid_decimal as _is_valid_decimal,
    is_valid_bit as _is_valid_bit,
    is_valid_int as _is_valid_int,
    to_date as _to_date,
)
from app.services.reasons import CAPACITY_DOWNTIME_REASONS, is_known_reason

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

REQUIRED_FILE_TYPES = frozenset({
    "master_stock",
    "demand_plan",
    "line_capacity_calendar",
    "headcount_plan",
    "portfolio_changes",
    "production_orders",
})

STAGE_NAMES = {
    1: "REQUIRED_FILE_CHECK",
    2: "TEMPLATE_STRUCTURE_CHECK",
    3: "FIELD_MAPPING_CHECK",
    4: "DATA_TYPE_CHECK",
    5: "REFERENCE_CHECK",
    6: "BUSINESS_RULE_CHECK",
    7: "BATCH_READINESS",
    8: "CROSS_FILE_CHECK",
}

# ---------------------------------------------------------------------------
# File schemas
# ---------------------------------------------------------------------------

_MONTH_HEADER_RE = re.compile(
    r'^('
    r'\d{2}[./]\d{4}'
    r'|\d{4}[./]\d{2}'
    r'|[a-z]{3}[_\-]\d{4}'
    r'|\d{4}-\d{2}'
    r'|m\d{2}\.\d{4}'
    r')$'
)


def _detect_month_columns(headers: list) -> list:
    return [h for h in headers if _MONTH_HEADER_RE.match(str(h).strip())]


def month_col_to_date(col: str) -> date | None:
    """Convert a normalised month column header to the first day of that month."""
    import calendar as _cal
    h = col.strip()
    if re.match(r'^m?\d{2}[./]\d{4}$', h):
        parts = re.split(r'[./]', h.lstrip('m'))
        mm, yyyy = int(parts[0]), int(parts[1])
    elif re.match(r'^\d{4}[./-]\d{2}$', h):
        parts = re.split(r'[./-]', h)
        yyyy, mm = int(parts[0]), int(parts[1])
    elif re.match(r'^[a-z]{3}[_\-]\d{4}$', h):
        mon_abbr, yr = re.split(r'[_\-]', h)
        month_map = {m.lower(): i for i, m in enumerate(_cal.month_abbr) if m}
        mm = month_map.get(mon_abbr)
        yyyy = int(yr)
        if mm is None:
            return None
    else:
        return None
    try:
        return date(yyyy, mm, 1)
    except ValueError:
        return None


FILE_SCHEMAS: dict = {
    "master_stock": {
        # SAP master_stock report usually has headers at row 1, but some SAP
        # configurations add a title/filter row at row 1 pushing headers to row 2.
        # header_row_candidates triggers auto-detection: we scan rows 1-2 and pick
        # whichever has more required column matches.
        # Column A and B are both labelled "Material" in the SAP export —
        # get_headers deduplicates the second to "material_2"; column_remap
        # renames it to "material_description".
        "header_row_candidates": [1, 2],
        "header_row": 1,   # fallback used by _stage2 for the emptiness check
        "column_remap": {
            "material_2": "material_description",
        },
        "required": [
            "material", "plant", "unrestrictedstock", "unrestricted_-_sales",
            "safety_stock",
        ],
        "optional": [
            "material_description", "abc_indicator",
            "mrp_type", "rounding_value", "volume",
        ],
        "types": {
            "material":                 "str",
            "plant":                    "str",
            "unrestrictedstock":        "decimal",
            "unrestricted_-_sales":     "decimal",
            "safety_stock":             "decimal",
            "material_description":     "str",
            "abc_indicator":            "str",
            "mrp_type":                 "str",
            "rounding_value":           "decimal",
            "volume":                   "decimal",
        },
        "fk_checks": {
            "plant": ("dbo.warehouses", "warehouse_code"),
        },
        # Soft FK — material should exist in dbo.items (sku_masterdata), but WARNING not BLOCKED
        # because master_stock may contain materials not yet loaded into sku_masterdata.
        "warning_fk_checks": {
            "material": ("dbo.items", "item_code"),
        },
        "min_rows": 1,
    },
    "demand_plan": {
        "is_sap": False,
        "header_row": 2,
        "data_start_row": 3,
        "required": ["material_id", "plant"],
        "optional": [],
        "types": {
            "material_id": "str",
            "plant":       "str",
        },
        "fk_checks": {
            "plant": ("dbo.warehouses", "warehouse_code"),
        },
        # No material_id FK check — validated against master_stock in stage 8 (cross-file).
        "min_rows": 1,
        "wide_format": True,
    },
    "line_capacity_calendar": {
        "is_sap": False,
        "header_row": 2,
        "data_start_row": 3,
        # planned_hours = scheduled shift; downtime_hours (with a reason) subtracts.
        "required": ["line_code", "calendar_date", "is_working_day", "planned_hours"],
        "optional": ["downtime_hours", "downtime_reason"],
        "types": {
            "line_code":        "str",
            "calendar_date":    "date",
            "is_working_day":   "bit",
            "planned_hours":    "decimal",
            "downtime_hours":   "decimal",
            "downtime_reason":  "str",
        },
        "fk_checks": {"line_code": ("dbo.lines", "line_code")},
        "min_rows": 1,
    },
    "headcount_plan": {
        "is_sap": False,
        "header_row": 2,
        "data_start_row": 3,
        # Phase 2 v2: the primary sheet is POOL headcount — people available per
        # labour POOL (which may span plants), by role, per month. Covers ALL roles
        # (line + shared). Absences live on the 'Exceptions' sheet.
        "required": ["pool_code", "resource_type_code", "planned_headcount"],
        "optional": ["plan_month"],
        "types": {
            "pool_code":          "str",
            "resource_type_code": "str",
            "plan_month":         "date",
            "plan_date":          "date",   # tolerated if a file still uses daily dates
            "planned_headcount":  "decimal",
        },
        "fk_checks": {
            "pool_code":          ("dbo.labour_pools",   "pool_code"),
            "resource_type_code": ("dbo.resource_types", "resource_type_code"),
        },
        "min_rows": 1,
    },
    "portfolio_changes": {
        "is_sap": False,
        "header_row": 2,
        "data_start_row": 3,
        # Phase-in: lists the SKUs being phased in. Volume & hours are derived from
        # the production plan (production_orders) for information.
        # Required columns must be present in the header even if there are 0 data rows.
        "required": ["effective_date", "item_code", "line_code"],
        "optional": ["comments"],
        "types": {
            "effective_date": "date",
            "item_code":      "str",
            "line_code":      "str",
            "comments":       "str",
        },
        # item_code must be a real SKU (drives the derived volume); line_code is the
        # line the SKU is phased in on — both must exist in masterdata.
        "fk_checks": {
            "item_code": ("dbo.items", "item_code"),
            "line_code": ("dbo.lines", "line_code"),
        },
        "min_rows": 0,  # Empty file valid — no phase-ins this cycle
    },
    "actual_production": {
        # SAP MB51 goods receipts export (list download format — single header row)
        "header_row": 1,
        "data_start_row": 2,
        "required": ["material", "quantity", "posting_date", "plant"],
        "optional": [
            "material_description", "movement_type", "storage_location",
            "material_document", "document_date", "order", "batch",
            "supplier", "reference", "customer", "sales_order",
            "purchase_order", "user_name",
        ],
        "types": {
            "material":       "str",
            "quantity":       "decimal",
            "posting_date":   "date",
            "plant":          "str",
            "movement_type":  "str",
        },
        "fk_checks": {
            "plant": ("dbo.warehouses", "warehouse_code"),
        },
        "blocked_fk_checks": {
            "material": ("dbo.items", "item_code"),
        },
        "min_rows": 1,
    },
    "production_orders": {
        "header_row": 2,
        "data_start_row": 3,
        # mrp_controller promoted to required
        "required": [
            "order",
            "material",
            "plant",
            "order_type",
            "mrp_controller",
            "order_quantity_(gmein)",
            "delivered_quantity_(gmein)",
            "basic_start_date",
        ],
        "optional": [
            "material_description",
            "unit_of_measure_(=gmein)",
            "basic_finish_date",
            "system_status",
            "production_line",
        ],
        "types": {
            "order":                       "str",
            "material":                    "str",
            "order_type":                  ("enum", ["LA", "YPAC"]),
            "mrp_controller":              "str",
            "plant":                       "str",
            "order_quantity_(gmein)":      "decimal",
            "delivered_quantity_(gmein)":  "decimal",
            "unit_of_measure_(=gmein)":    "str",
            "basic_start_date":            "date",
            "basic_finish_date":           "date",
            "system_status":               "str",
            "production_line":             "str",
        },
        "fk_checks": {
            "plant": ("dbo.warehouses", "warehouse_code"),
        },
        # Hard FK — material MUST exist in dbo.items (sku_masterdata). BLOCKED if any missing.
        # Summary format (one BLOCKED per unique missing value) avoids per-row overflow.
        "blocked_fk_checks": {
            "material": ("dbo.items", "item_code"),
        },
        "min_rows": 1,
    },
}


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def run_validation(
    conn: pyodbc.Connection,
    batch_id: int,
    batch_file_id: int,
    file_type: str,
    stored_file_path: str,
) -> None:
    """Run all 8 stages for one file. Writes results to import_validation_results
    and updates validation_status on import_batch_files. Commits when done."""

    cursor = conn.cursor()

    # Clear previous results for this file
    cursor.execute(
        "DELETE FROM dbo.import_validation_results WHERE batch_file_id = ?",
        batch_file_id,
    )

    schema = FILE_SCHEMAS.get(file_type, {"is_sap": True})
    is_sap = schema.get("is_sap", False)

    # --- Stage 1: batch-level required file check ---
    _stage1(cursor, conn, batch_id, batch_file_id)

    # --- Stage 2: structure check ---
    wb = _stage2(cursor, batch_file_id, stored_file_path, schema.get("header_row", 1))

    if wb is None:
        for n in range(3, 7):
            _write(cursor, batch_file_id, n, STAGE_NAMES[n], "INFO",
                   "Stage skipped — file could not be opened (Stage 2 blocked)")
    elif is_sap:
        for n in range(3, 7):
            _write(cursor, batch_file_id, n, STAGE_NAMES[n], "INFO",
                   "SAP export — column mapping not yet confirmed. "
                   "Update file_schemas when SAP export headers are available.")
    else:
        ws = wb.worksheets[0]
        # Resolve header row — auto-detect when candidates list is provided
        candidates = schema.get("header_row_candidates")
        if candidates:
            header_row = _find_header_row(ws, schema.get("required", []), tuple(candidates))
            data_start_row = header_row + 1
        else:
            header_row = schema.get("header_row", 1)
            data_start_row = schema.get("data_start_row", header_row + 1)

        headers = _get_headers(ws, header_row)
        # Apply column renames (e.g. duplicate SAP "Material" → "material_description")
        if schema.get("column_remap"):
            headers = [schema["column_remap"].get(h, h) for h in headers]
        data_rows = _get_data_rows(ws, headers, data_start_row)

        sap_uom_cols = set(schema.get("sap_uom_cols", []))
        if sap_uom_cols:
            data_rows = [
                (rn, {k: (_strip_sap_uom(v) if k in sap_uom_cols else v) for k, v in rd.items()})
                for rn, rd in data_rows
            ]

        blocked_cols = _stage3(cursor, batch_file_id, schema, headers)

        if blocked_cols:
            for n in range(4, 7):
                _write(cursor, batch_file_id, n, STAGE_NAMES[n], "INFO",
                       "Stage skipped — required columns missing (Stage 3 blocked)")
        elif not data_rows and schema.get("min_rows", 1) == 0:
            for n in range(4, 7):
                _write(cursor, batch_file_id, n, STAGE_NAMES[n], "PASS",
                       "No data rows — empty file is valid for this file type")
        elif not data_rows:
            for n in range(4, 7):
                _write(cursor, batch_file_id, n, STAGE_NAMES[n], "BLOCKED",
                       f"File has a header row but no data rows. "
                       f"At least {schema.get('min_rows', 1)} data row(s) required.")
        else:
            _stage4(cursor, batch_file_id, schema, headers, data_rows)
            _stage5(cursor, conn, batch_file_id, schema, headers, data_rows)
            _stage6(cursor, conn, batch_file_id, file_type, schema, headers, data_rows)


    # --- Plant Support Sheet 2 (headcount_plan only, advisory) ---
    if file_type == "headcount_plan" and wb is not None:
        try:
            _validate_plant_support_sheet(cursor, conn, batch_file_id, wb)
        except Exception as exc:
            try:
                c_err = conn.cursor()
                c_err.execute(
                    """INSERT INTO dbo.import_validation_results
                           (batch_file_id, validation_stage, stage_name, severity, message)
                       VALUES (?, 6, 'BUSINESS_RULE_CHECK', 'WARNING', ?)""",
                    batch_file_id,
                    f"Plant Support sheet check failed: {str(exc)[:300]}",
                )
            except Exception:
                pass

    # --- Stage 7: batch readiness (always runs) ---
    _stage7(cursor, conn, batch_id, batch_file_id)

    # --- Stage 8: cross-file checks (only for participating file types, only when file opened) ---
    _STAGE8_TYPES = ("master_stock", "portfolio_changes",
                     "production_orders", "demand_plan")
    if file_type in _STAGE8_TYPES and wb is not None:
        try:
            _stage8(cursor, conn, batch_id, batch_file_id, file_type, stored_file_path)
        except Exception as exc:
            # Stage 8 is advisory only — log but don't let it block the commit.
            # Use a fresh cursor so a failed stage 8 INSERT doesn't poison the transaction.
            try:
                c_err = conn.cursor()
                c_err.execute(
                    """INSERT INTO dbo.import_validation_results
                           (batch_file_id, validation_stage, stage_name, severity, message)
                       VALUES (?, 8, 'CROSS_FILE_CHECK', 'WARNING', ?)""",
                    batch_file_id,
                    f"Cross-file check could not complete: {str(exc)[:300]}",
                )
            except Exception:
                pass  # If even the fallback write fails, silently skip — don't block commit

    # Update validation_status on the file record (stages 2-6 only; stage 8 is WARNING-only)
    _update_file_status(cursor, batch_file_id)

    conn.commit()


# ---------------------------------------------------------------------------
# Stage implementations
# ---------------------------------------------------------------------------

def _stage1(cursor, conn: pyodbc.Connection, batch_id: int, batch_file_id: int) -> None:
    stage_num, stage_name = 1, STAGE_NAMES[1]
    c2 = conn.cursor()
    c2.execute(
        """
        SELECT COUNT(DISTINCT file_type)
        FROM dbo.import_batch_files
        WHERE batch_id = ? AND is_current_version = 1
          AND file_type IN (
              'master_stock', 'demand_plan', 'line_capacity_calendar',
              'headcount_plan', 'portfolio_changes', 'production_orders'
          )
        """,
        batch_id,
    )
    count = c2.fetchone()[0]
    if count >= 6:
        _write(cursor, batch_file_id, stage_num, stage_name, "PASS",
               "All 6 required files are present in this batch")
    else:
        missing = 6 - count
        _write(cursor, batch_file_id, stage_num, stage_name, "WARNING",
               f"{count}/6 required files present — {missing} still to upload")


def _stage2(cursor, batch_file_id: int, stored_file_path: str, header_row: int = 1):
    stage_num, stage_name = 2, STAGE_NAMES[2]
    try:
        wb = openpyxl.load_workbook(stored_file_path, read_only=True, data_only=True)
    except Exception as exc:
        _write(cursor, batch_file_id, stage_num, stage_name, "BLOCKED",
               f"File could not be opened as a valid Excel workbook: {str(exc)[:200]}")
        return None

    ws = wb.worksheets[0] if wb.worksheets else None
    if ws is None:
        _write(cursor, batch_file_id, stage_num, stage_name, "BLOCKED",
               "Workbook has no sheets")
        return None

    header_vals = [cell.value for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row), [])]
    if all(v is None for v in header_vals):
        _write(cursor, batch_file_id, stage_num, stage_name, "BLOCKED",
               f"Header row (row {header_row}) is empty — expected column names")
        return None

    data_row_count = max(0, (ws.max_row or header_row) - header_row)
    _write(cursor, batch_file_id, stage_num, stage_name, "PASS",
           f"File opened successfully. Sheet: '{ws.title}'. "
           f"~{data_row_count} data row(s) detected.")
    return wb


def _stage3(cursor, batch_file_id: int, schema: dict, headers: list) -> list:
    """Returns list of blocked (missing required) column names."""
    stage_num, stage_name = 3, STAGE_NAMES[3]
    header_set = set(headers)
    required = schema.get("required", [])
    optional = schema.get("optional", [])

    blocked_cols = [c for c in required if c not in header_set]
    warning_cols = [c for c in optional if c not in header_set]

    for col in blocked_cols:
        _write(cursor, batch_file_id, stage_num, stage_name, "BLOCKED",
               f"Required column '{col}' not found in header row",
               field_name=col)

    for col in warning_cols:
        _write(cursor, batch_file_id, stage_num, stage_name, "WARNING",
               f"Optional column '{col}' not found — will be treated as NULL",
               field_name=col)

    # headcount_plan accepts either plan_month (preferred, monthly format) or
    # plan_date (legacy daily format). At least one must be present.
    if schema.get("types", {}).get("plan_month") == "date" \
            and "plan_month" not in header_set and "plan_date" not in header_set:
        _write(cursor, batch_file_id, stage_num, stage_name, "BLOCKED",
               "Required column 'plan_month' not found (or legacy 'plan_date' for daily files)",
               field_name="plan_month")
        blocked_cols.append("plan_month")

    if schema.get("wide_format") and not blocked_cols:
        month_cols = _detect_month_columns(headers)
        if not month_cols:
            _write(cursor, batch_file_id, stage_num, stage_name, "BLOCKED",
                   "No month columns detected. Expected columns like 'M03.2026', '01/2026', 'Jan-2026' or '2026-01' "
                   "alongside MaterialID and Plant.")
            blocked_cols.append("__month_columns__")
        else:
            sample = ", ".join(month_cols[:3]) + ("…" if len(month_cols) > 3 else "")
            _write(cursor, batch_file_id, stage_num, stage_name, "PASS",
                   f"Found {len(month_cols)} month column(s): {sample}")
    elif not blocked_cols and not warning_cols:
        _write(cursor, batch_file_id, stage_num, stage_name, "PASS",
               "All expected columns are present")

    return blocked_cols


def _stage4(cursor, batch_file_id: int, schema: dict, headers: list, data_rows: list) -> None:
    stage_num, stage_name = 4, STAGE_NAMES[4]
    types = schema.get("types", {})
    required_cols = set(schema.get("required", []))
    header_set = set(headers)
    errors: list[tuple] = []

    for row_num, row_dict in data_rows:
        for col, type_spec in types.items():
            if col not in header_set:
                continue
            val = row_dict.get(col)
            is_empty = val is None or (isinstance(val, str) and val.strip() == "")

            if is_empty:
                if col in required_cols:
                    errors.append((row_num, col, "BLOCKED", "Required value is empty", None))
                continue

            if type_spec == "date":
                if not _is_valid_date(val):
                    errors.append((row_num, col, "BLOCKED",
                                   f"Expected a date, got: '{val}'", str(val)))
            elif type_spec == "decimal":
                if not _is_valid_decimal(val):
                    errors.append((row_num, col, "BLOCKED",
                                   f"Expected a number, got: '{val}'", str(val)))
            elif type_spec == "bit":
                if not _is_valid_bit(val):
                    errors.append((row_num, col, "BLOCKED",
                                   f"Expected 0/1 or Yes/No, got: '{val}'", str(val)))
            elif isinstance(type_spec, tuple) and type_spec[0] == "enum":
                allowed = type_spec[1]
                if str(val).strip().upper() not in [a.upper() for a in allowed]:
                    errors.append((row_num, col, "BLOCKED",
                                   f"'{val}' not in allowed values: {', '.join(allowed)}",
                                   str(val)))

    if schema.get("wide_format"):
        month_cols = _detect_month_columns(headers)
        for col in month_cols:
            for row_num, row_dict in data_rows:
                val = row_dict.get(col)
                if val is None or (isinstance(val, str) and val.strip() == ""):
                    continue
                if not _is_valid_decimal(val):
                    errors.append((row_num, col, "BLOCKED",
                                   f"Expected a number in month column '{col}', got: '{val}'",
                                   str(val)))

    _emit_errors(cursor, batch_file_id, stage_num, stage_name, errors,
                 pass_msg=f"All data types valid across {len(data_rows)} row(s)")


def _stage5(cursor, conn: pyodbc.Connection, batch_file_id: int,
            schema: dict, headers: list, data_rows: list) -> None:
    stage_num, stage_name = 5, STAGE_NAMES[5]
    header_set = set(headers)
    errors: list[tuple] = []

    # Hard FK checks — BLOCKED if value not found
    for col, (table, pk_col) in schema.get("fk_checks", {}).items():
        if col not in header_set:
            continue
        c2 = conn.cursor()
        c2.execute(f"SELECT {pk_col} FROM {table}")  # noqa: S608
        valid_values = {str(r[0]).strip() for r in c2.fetchall()}
        for row_num, row_dict in data_rows:
            val = row_dict.get(col)
            if val is None or (isinstance(val, str) and val.strip() == ""):
                continue
            if str(val).strip() not in valid_values:
                errors.append((row_num, col, "BLOCKED",
                               f"'{val}' not found in {table} ({pk_col})", str(val)))

    # Soft FK checks — WARNING, one row per unique missing value (deduped).
    # Written directly (not through _emit_errors) so the full list is stored.
    fk_summary_written = False
    for col, (table, pk_col) in schema.get("warning_fk_checks", {}).items():
        if col not in header_set:
            continue
        c2 = conn.cursor()
        c2.execute(f"SELECT {pk_col} FROM {table}")  # noqa: S608
        valid_values = {str(r[0]).strip() for r in c2.fetchall()}
        missing = sorted({
            str(row_dict.get(col)).strip()
            for _, row_dict in data_rows
            if row_dict.get(col) and str(row_dict.get(col)).strip() not in valid_values
        })
        for val in missing:
            _write(cursor, batch_file_id, stage_num, stage_name, "WARNING",
                   f"'{val}' not found in {table} ({pk_col}) — add to sku_masterdata to fix",
                   field_name=col, sample_value=val)
            fk_summary_written = True

    # Hard FK summary checks — BLOCKED, one row per unique missing value (deduped).
    for col, (table, pk_col) in schema.get("blocked_fk_checks", {}).items():
        if col not in header_set:
            continue
        c2 = conn.cursor()
        c2.execute(f"SELECT {pk_col} FROM {table}")  # noqa: S608
        valid_values = {str(r[0]).strip() for r in c2.fetchall()}
        missing = sorted({
            str(row_dict.get(col)).strip()
            for _, row_dict in data_rows
            if row_dict.get(col) and str(row_dict.get(col)).strip() not in valid_values
        })
        for val in missing:
            _write(cursor, batch_file_id, stage_num, stage_name, "BLOCKED",
                   f"'{val}' not found in {table} ({pk_col}) — add to sku_masterdata before validating",
                   field_name=col, sample_value=val)
            fk_summary_written = True

    # Per-row errors (fk_checks) go through _emit_errors as before.
    # PASS only if nothing was written by either path.
    if errors:
        _emit_errors(cursor, batch_file_id, stage_num, stage_name, errors,
                     pass_msg="All reference checks passed")
    elif not fk_summary_written:
        _write(cursor, batch_file_id, stage_num, stage_name, "PASS",
               "All reference checks passed")


def _horizon_context(conn: pyodbc.Connection, batch_file_id: int) -> tuple:
    """Return (plan_cycle_date, active_line_count) for coverage guardrails."""
    c = conn.cursor()
    c.execute(
        """SELECT b.plan_cycle_date
           FROM dbo.import_batch_files f
           JOIN dbo.import_batches b ON f.batch_id = b.batch_id
           WHERE f.batch_file_id = ?""",
        batch_file_id,
    )
    row = c.fetchone()
    plan_cycle = row[0] if row else None
    c.execute("SELECT COUNT(*) FROM dbo.lines WHERE is_active = 1")
    active_lines = c.fetchone()[0] or 0
    return plan_cycle, active_lines


def _stage6(cursor, conn: pyodbc.Connection, batch_file_id: int, file_type: str,
            schema: dict, headers: list, data_rows: list) -> None:
    stage_num, stage_name = 6, STAGE_NAMES[6]
    header_set = set(headers)
    errors: list[tuple] = []

    if file_type == "headcount_plan":
        for col in ("planned_headcount", "available_hours"):
            if col not in header_set:
                continue
            for row_num, row_dict in data_rows:
                val = row_dict.get(col)
                if val is not None and _is_valid_decimal(val) and float(val) < 0:
                    errors.append((row_num, col, "BLOCKED",
                                   f"{col} cannot be negative: {val}", str(val)))

        # Coverage guardrail — catch a template/sample uploaded by mistake.
        # Primary sheet is pool headcount (pool + role + month).
        plan_cycle, _active_lines = _horizon_context(conn, batch_file_id)
        months = []
        for _, rd in data_rows:
            raw = rd.get("plan_month") if "plan_month" in header_set else rd.get("plan_date")
            d = _to_date(raw)
            if d:
                months.append(d)
        max_month = max(months) if months else None
        reasons = []
        if max_month is not None and plan_cycle is not None and max_month < plan_cycle:
            reasons.append(f"latest month in file is {max_month:%Y-%m}, before the plan cycle ({plan_cycle:%Y-%m})")
        if len(data_rows) <= 3:
            reasons.append(f"only {len(data_rows)} row(s)")
        if reasons:
            errors.append((None, "planned_headcount", "WARNING",
                           "headcount_plan (pool headcount) looks incomplete (" + "; ".join(reasons) + "). "
                           "Expected pool headcount per pool per role across the planning horizon — "
                           "did you upload the template/sample instead of the full plan?",
                           None))

    elif file_type == "line_capacity_calendar":
        if "planned_hours" in header_set:
            for row_num, row_dict in data_rows:
                val = row_dict.get("planned_hours")
                if val is not None and _is_valid_decimal(val):
                    f = float(val)
                    if f < 0:
                        errors.append((row_num, "planned_hours", "BLOCKED",
                                       f"planned_hours cannot be negative: {val}", str(val)))
                    elif f > 24:
                        errors.append((row_num, "planned_hours", "BLOCKED",
                                       f"planned_hours cannot exceed 24: {val}", str(val)))
        # Downtime hours + reason. Downtime subtracts from the shift; a reason is
        # expected when there's downtime, and an off-list reason is a soft WARNING.
        if "downtime_hours" in header_set:
            for row_num, row_dict in data_rows:
                val = row_dict.get("downtime_hours")
                dt = float(val) if (val is not None and _is_valid_decimal(val)) else 0.0
                if dt < 0:
                    errors.append((row_num, "downtime_hours", "BLOCKED",
                                   f"downtime_hours cannot be negative: {val}", str(val)))
                    continue
                if dt > 0:
                    planned = row_dict.get("planned_hours")
                    if planned is not None and _is_valid_decimal(planned) and dt > float(planned):
                        errors.append((row_num, "downtime_hours", "WARNING",
                                       f"downtime_hours ({dt}) exceeds planned_hours ({planned}) — capacity floored at 0", str(val)))
                    reason = str(row_dict.get("downtime_reason", "") or "").strip()
                    if not reason:
                        errors.append((row_num, "downtime_reason", "WARNING",
                                       "downtime_hours recorded without a downtime_reason", None))
                    elif not is_known_reason(reason, CAPACITY_DOWNTIME_REASONS):
                        errors.append((row_num, "downtime_reason", "WARNING",
                                       f"'{reason}' is not a standard downtime reason "
                                       f"({', '.join(CAPACITY_DOWNTIME_REASONS)}) — imported as-is", reason))

        # Coverage guardrail — catch a template/sample uploaded by mistake.
        if "calendar_date" in header_set:
            plan_cycle, _active = _horizon_context(conn, batch_file_id)
            dates = [_to_date(rd.get("calendar_date")) for _, rd in data_rows]
            dates = [d for d in dates if d is not None]
            distinct_dates = set(dates)
            max_date = max(dates) if dates else None
            reasons = []
            if max_date is not None and plan_cycle is not None and max_date < plan_cycle:
                reasons.append(f"latest date in file is {max_date:%Y-%m-%d}, before the plan cycle ({plan_cycle:%Y-%m})")
            if len(distinct_dates) < 28:
                reasons.append(f"spans only {len(distinct_dates)} distinct day(s)")
            if reasons:
                errors.append((None, "calendar_date", "WARNING",
                               "line_capacity_calendar looks incomplete (" + "; ".join(reasons) + "). "
                               "Expected a full calendar spanning the planning horizon for every active line — "
                               "did you upload the template/sample instead of the generated capacity calendar?",
                               None))

    elif file_type == "demand_plan":
        month_cols = _detect_month_columns(headers)
        for col in month_cols:
            for row_num, row_dict in data_rows:
                val = row_dict.get(col)
                if val is not None and _is_valid_decimal(val) and float(val) < 0:
                    errors.append((row_num, col, "BLOCKED",
                                   f"Demand quantity cannot be negative: {val}", str(val)))

    elif file_type == "master_stock":
        # unrestrictedstock and safety_stock must be >= 0
        # unrestricted_-_sales may be negative (back-orders) — no check needed
        for col in ("unrestrictedstock", "safety_stock"):
            if col not in header_set:
                continue
            for row_num, row_dict in data_rows:
                val = row_dict.get(col)
                if val is not None and _is_valid_decimal(val) and float(val) < 0:
                    errors.append((row_num, col, "BLOCKED",
                                   f"{col} cannot be negative: {val}", str(val)))

        # SKU routing completeness: warn about materials with no RCCP configuration.
        # pack_size_l and units_per_pallet come from this file (volume/rounding_value) —
        # pack_type_code (warehouse category) and primary_line_code (routing) must come
        # from sku_masterdata. Single summary WARNINGs to avoid hitting the 20-row cap.
        if "material" in header_set:
            materials_in_file = {
                str(row_dict.get("material", "")).strip()
                for _, row_dict in data_rows
                if row_dict.get("material")
            }
            if materials_in_file:
                placeholders = ",".join("?" * len(materials_in_file))
                c_read = conn.cursor()
                c_read.execute(
                    f"""
                    SELECT item_code,
                           CASE WHEN primary_line_code IS NULL THEN 1 ELSE 0 END AS no_line,
                           CASE WHEN pack_type_code    IS NULL THEN 1 ELSE 0 END AS no_pack_type
                    FROM dbo.items
                    WHERE item_code IN ({placeholders})
                      AND (primary_line_code IS NULL OR pack_type_code IS NULL)
                    """,  # noqa: S608
                    *materials_in_file,
                )
                incomplete_rows = c_read.fetchall()
                no_line_codes    = [r[0] for r in incomplete_rows if r[1]]
                no_pack_codes    = [r[0] for r in incomplete_rows if r[2]]
                if no_line_codes:
                    sample = ", ".join(f"'{c}'" for c in no_line_codes[:5])
                    errors.append((None, "material", "WARNING",
                                   f"{len(no_line_codes)} SKU(s) have no RCCP routing config (no line assignment) — "
                                   f"add primary_line_code to sku_masterdata: {sample}"
                                   + ("..." if len(no_line_codes) > 5 else ""),
                                   None))
                if no_pack_codes:
                    sample = ", ".join(f"'{c}'" for c in no_pack_codes[:5])
                    errors.append((None, "material", "WARNING",
                                   f"{len(no_pack_codes)} SKU(s) have no pack type — "
                                   f"add pack_type_code to sku_masterdata: {sample}"
                                   + ("..." if len(no_pack_codes) > 5 else ""),
                                   None))

    elif file_type == "portfolio_changes":
        # Every phase-in row needs the SKU and the affected line. item_code / line_code
        # existence is enforced via the hard FKs (stage 5); here we flag blank values.
        # (Whether the SKU is in the production plan is a WARNING — see stage 8.)
        for col, label in (("item_code", "the SKU being phased in"),
                           ("line_code", "the line the SKU is phased in on")):
            for row_num, row_dict in data_rows:
                val = row_dict.get(col)
                if val is None or (isinstance(val, str) and val.strip() == ""):
                    errors.append((row_num, col, "BLOCKED",
                                   f"Phase-in row requires {col} — {label}", None))

    elif file_type == "production_orders":
        if "order_quantity_(gmein)" in header_set:
            for row_num, row_dict in data_rows:
                val = row_dict.get("order_quantity_(gmein)")
                if val is not None and _is_valid_decimal(val) and float(val) <= 0:
                    errors.append((row_num, "order_quantity_(gmein)", "BLOCKED",
                                   f"Order quantity must be > 0: {val}", str(val)))
        if "delivered_quantity_(gmein)" in header_set:
            for row_num, row_dict in data_rows:
                val = row_dict.get("delivered_quantity_(gmein)")
                if val is not None and _is_valid_decimal(val) and float(val) < 0:
                    errors.append((row_num, "delivered_quantity_(gmein)", "BLOCKED",
                                   f"Delivered quantity cannot be negative: {val}", str(val)))

    elif file_type == "actual_production":
        if "quantity" in header_set:
            for row_num, row_dict in data_rows:
                val = row_dict.get("quantity")
                if val is not None and _is_valid_decimal(val) and float(val) <= 0:
                    errors.append((row_num, "quantity", "BLOCKED",
                                   f"Quantity must be > 0: {val}", str(val)))

    _emit_errors(cursor, batch_file_id, stage_num, stage_name, errors,
                 pass_msg="All business rules passed")


def _validate_plant_support_sheet(cursor, conn, batch_file_id: int, wb) -> None:
    """Validate Sheet 2 (Plant Support) of headcount_plan.xlsx. Optional — INFO if absent."""
    stage_num, stage_name = 6, STAGE_NAMES[6]

    sheet2 = next(
        (wb[s] for s in wb.sheetnames if 'plant' in s.lower() or 'support' in s.lower()),
        None,
    )
    if sheet2 is None:
        _write(cursor, batch_file_id, stage_num, stage_name, "INFO",
               "Plant Support sheet not found in headcount_plan.xlsx — "
               "add a 'Plant Support' sheet to track plant-level headcount (optional).")
        return

    # Check required columns in header row 2. Accept either plan_month
    # (preferred — monthly format) or legacy plan_date (daily format).
    base_required = {"plant_code", "resource_type_code", "planned_headcount"}
    headers2 = _get_headers(sheet2, header_row=2)
    header_set = set(headers2)
    missing = base_required - header_set
    if "plan_month" not in header_set and "plan_date" not in header_set:
        missing = missing | {"plan_month"}
    if missing:
        _write(cursor, batch_file_id, stage_num, stage_name, "BLOCKED",
               f"Plant Support sheet: missing required columns: {', '.join(sorted(missing))}. "
               f"Expected header row 2 with: plant_code, resource_type_code, plan_month "
               f"(or legacy plan_date), planned_headcount.")
        return

    data_rows2 = _get_data_rows(sheet2, headers2, start_row=3)
    if not data_rows2:
        _write(cursor, batch_file_id, stage_num, stage_name, "INFO",
               "Plant Support sheet has no data rows — sheet is present but empty.")
        return

    # FK lookup sets
    c2 = conn.cursor()
    c2.execute("SELECT plant_code FROM dbo.plants")
    valid_plants = {r[0] for r in c2.fetchall()}
    c2.execute("SELECT resource_type_code FROM dbo.resource_types WHERE scope = 'PLANT'")
    valid_roles = {r[0] for r in c2.fetchall()}

    errors: list[str] = []
    for row_num, row in data_rows2:
        pc = (row.get("plant_code") or "").strip()
        rc = (row.get("resource_type_code") or "").strip()
        pd_raw = row.get("plan_month") if "plan_month" in header_set else row.get("plan_date")
        hc_raw = row.get("planned_headcount")

        if not pc or not rc or pd_raw is None:
            errors.append(f"Row {row_num}: plant_code, resource_type_code, plan_month are required.")
            continue
        if pc not in valid_plants:
            errors.append(f"Row {row_num}: plant_code '{pc}' not found in dbo.plants.")
        if rc not in valid_roles:
            errors.append(f"Row {row_num}: resource_type_code '{rc}' not found in dbo.resource_types (scope=PLANT).")
        try:
            hc = float(hc_raw) if hc_raw is not None else None
            if hc is None or hc < 0:
                errors.append(f"Row {row_num}: planned_headcount must be a number ≥ 0.")
        except (TypeError, ValueError):
            errors.append(f"Row {row_num}: planned_headcount '{hc_raw}' is not a valid number.")

    if errors:
        sample = errors[:5]
        suffix = f" (+{len(errors) - 5} more)" if len(errors) > 5 else ""
        _write(cursor, batch_file_id, stage_num, stage_name, "BLOCKED",
               f"Plant Support sheet has {len(errors)} error(s): "
               + "; ".join(sample) + suffix)
    else:
        _write(cursor, batch_file_id, stage_num, stage_name, "PASS",
               f"Plant Support sheet: {len(data_rows2)} row(s) validated OK.")


def _stage7(cursor, conn: pyodbc.Connection, batch_id: int, batch_file_id: int) -> None:
    stage_num, stage_name = 7, STAGE_NAMES[7]
    c2 = conn.cursor()

    c2.execute(
        """
        SELECT COUNT(DISTINCT file_type)
        FROM dbo.import_batch_files
        WHERE batch_id = ? AND is_current_version = 1
          AND file_type IN (
              'master_stock', 'demand_plan', 'line_capacity_calendar',
              'headcount_plan', 'portfolio_changes', 'production_orders'
          )
        """,
        batch_id,
    )
    required_count = c2.fetchone()[0]

    # Count required files that have been uploaded but NOT yet validated (validation_status IS NULL)
    c2.execute(
        """
        SELECT COUNT(DISTINCT file_type)
        FROM dbo.import_batch_files
        WHERE batch_id = ? AND is_current_version = 1
          AND file_type IN (
              'master_stock', 'demand_plan', 'line_capacity_calendar',
              'headcount_plan', 'portfolio_changes', 'production_orders'
          )
          AND validation_status IS NULL
        """,
        batch_id,
    )
    unvalidated_count = c2.fetchone()[0]

    c2.execute(
        """
        SELECT COUNT(*) FROM dbo.import_validation_results ivr
        JOIN dbo.import_batch_files ibf ON ivr.batch_file_id = ibf.batch_file_id
        WHERE ibf.batch_id = ? AND ibf.is_current_version = 1
          AND ivr.severity = 'BLOCKED'
          AND ivr.validation_stage BETWEEN 2 AND 6
        """,
        batch_id,
    )
    blocked_count = c2.fetchone()[0]

    c2.execute(
        """
        SELECT COUNT(*) FROM dbo.import_validation_results ivr
        JOIN dbo.import_batch_files ibf ON ivr.batch_file_id = ibf.batch_file_id
        WHERE ibf.batch_id = ? AND ibf.is_current_version = 1
          AND ivr.severity = 'WARNING'
          AND ivr.validation_stage BETWEEN 2 AND 6
        """,
        batch_id,
    )
    warning_count = c2.fetchone()[0]

    if required_count < 6:
        _write(cursor, batch_file_id, stage_num, stage_name, "WARNING",
               f"Batch not ready to publish: {required_count}/6 required files uploaded")
    elif unvalidated_count > 0:
        _write(cursor, batch_file_id, stage_num, stage_name, "BLOCKED",
               f"{unvalidated_count} file(s) uploaded but not yet validated — "
               f"click Re-validate to run the full validation pipeline")
    elif blocked_count > 0:
        _write(cursor, batch_file_id, stage_num, stage_name, "BLOCKED",
               f"Batch cannot be published: {blocked_count} BLOCKED issue(s) across all files")
    elif warning_count > 0:
        _write(cursor, batch_file_id, stage_num, stage_name, "WARNING",
               f"Batch can be published but has {warning_count} warning(s). Review before publishing.")
    else:
        _write(cursor, batch_file_id, stage_num, stage_name, "PASS",
               "All files validated. Batch is ready to publish.")


# ---------------------------------------------------------------------------
# Stage 8: Cross-file checks
# ---------------------------------------------------------------------------

def _stage8(cursor, conn: pyodbc.Connection, batch_id: int, batch_file_id: int,
            file_type: str, stored_file_path: str) -> None:
    """Run cross-file consistency checks. WARNING severity only — never blocks publish."""
    if file_type == "master_stock":
        _stage8_sku_coverage(cursor, conn, batch_id, batch_file_id, stored_file_path)
    # headcount_plan: no cross-file per-line coverage check — headcount is now
    # pooled per plant (Phase 2), so line-level coverage no longer applies.
    elif file_type == "portfolio_changes":
        _stage8_portfolio_in_plan(cursor, conn, batch_id, batch_file_id, stored_file_path)
    elif file_type == "production_orders":
        _stage8_material_vs_master_stock(cursor, conn, batch_id, batch_file_id,
                                         stored_file_path, "production_orders", "material")
    elif file_type == "demand_plan":
        _stage8_material_vs_master_stock(cursor, conn, batch_id, batch_file_id,
                                         stored_file_path, "demand_plan", "material_id")


def _get_sibling_path(conn: pyodbc.Connection, batch_id: int, file_type: str) -> str | None:
    """Return stored_file_path for the current-version file of the given type in this batch."""
    c = conn.cursor()
    c.execute(
        """SELECT stored_file_path FROM dbo.import_batch_files
           WHERE batch_id = ? AND file_type = ? AND is_current_version = 1""",
        batch_id, file_type,
    )
    row = c.fetchone()
    return row[0] if row else None


def _read_file_column(stored_path: str, file_type: str, column_name: str) -> set:
    """Open a sibling file and return all non-blank values from a single column as a set."""
    schema = FILE_SCHEMAS[file_type]
    try:
        wb = openpyxl.load_workbook(stored_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        headers = _get_headers(ws, schema["header_row"])
        rows = _get_data_rows(ws, headers, schema["data_start_row"])
        return {
            str(row.get(column_name, "")).strip()
            for _, row in rows
            if row.get(column_name) is not None and str(row.get(column_name, "")).strip()
        }
    except Exception:
        return set()


def _stage8_sku_coverage(cursor, conn, batch_id, batch_file_id, stored_file_path):
    """Warn about SKUs in master_stock that have no demand in demand_plan or production_orders."""
    stage_num, stage_name = 8, STAGE_NAMES[8]

    ms_items = _read_file_column(stored_file_path, "master_stock", "material")
    if not ms_items:
        return

    dp_path = _get_sibling_path(conn, batch_id, "demand_plan")
    po_path = _get_sibling_path(conn, batch_id, "production_orders")

    if not dp_path and not po_path:
        _write(cursor, batch_file_id, stage_num, stage_name, "INFO",
               "demand_plan and production_orders not yet uploaded — "
               "SKU coverage check pending. Run Re-validate when all files are uploaded.")
        return

    demand_items = _read_file_column(dp_path, "demand_plan", "material_id") if dp_path else set()
    po_items = _read_file_column(po_path, "production_orders", "material") if po_path else set()

    covered = demand_items | po_items
    uncovered = sorted(ms_items - covered)

    if not uncovered:
        _write(cursor, batch_file_id, stage_num, stage_name, "PASS",
               f"All {len(ms_items)} SKUs in master_stock are covered by demand_plan or production_orders")
        return

    for item_code in uncovered[:50]:
        _write(cursor, batch_file_id, stage_num, stage_name, "WARNING",
               f"SKU '{item_code}' is in master_stock but has no demand in demand_plan "
               f"or open orders in production_orders",
               field_name="sku_coverage", sample_value=item_code)
    if len(uncovered) > 50:
        _write(cursor, batch_file_id, stage_num, stage_name, "WARNING",
               f"{len(uncovered) - 50} more uncovered SKUs not shown — "
               f"review the coverage report for the full list",
               field_name="sku_coverage")


def _stage8_headcount_coverage(cursor, conn, batch_id, batch_file_id, stored_file_path):
    """Warn about (line, month) combinations in line_capacity_calendar missing from headcount_plan.

    Coverage check is month-level (not date-level) since the headcount file is
    now monthly. Legacy daily files are still tolerated — each daily row maps
    to a (line, month) pair, so coverage works either way.
    """
    stage_num, stage_name = 8, STAGE_NAMES[8]

    lcc_path = _get_sibling_path(conn, batch_id, "line_capacity_calendar")
    if not lcc_path:
        _write(cursor, batch_file_id, stage_num, stage_name, "INFO",
               "line_capacity_calendar not yet uploaded — "
               "headcount coverage check pending. Run Re-validate when all files are uploaded.")
        return

    # Collect (line_code, year-month) from line_capacity_calendar — operating months only
    lcc_schema = FILE_SCHEMAS["line_capacity_calendar"]
    calendar_pairs: set[tuple] = set()
    try:
        wb = openpyxl.load_workbook(lcc_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        headers = _get_headers(ws, lcc_schema["header_row"])
        rows = _get_data_rows(ws, headers, lcc_schema["data_start_row"])
        for _, row in rows:
            lc = str(row.get("line_code", "")).strip()
            dt = _to_date(row.get("calendar_date"))
            if lc and dt:
                calendar_pairs.add((lc, (dt.year, dt.month)))
    except Exception:
        _write(cursor, batch_file_id, stage_num, stage_name, "WARNING",
               "Could not read line_capacity_calendar — headcount coverage check skipped",
               field_name="headcount_coverage")
        return

    if not calendar_pairs:
        return

    # Collect (line_code, year-month) from headcount_plan (accept plan_month or plan_date)
    hc_schema = FILE_SCHEMAS["headcount_plan"]
    hc_pairs: set[tuple] = set()
    try:
        wb2 = openpyxl.load_workbook(stored_file_path, read_only=True, data_only=True)
        ws2 = wb2.active
        headers2 = _get_headers(ws2, hc_schema["header_row"])
        rows2 = _get_data_rows(ws2, headers2, hc_schema["data_start_row"])
        for _, row in rows2:
            lc = str(row.get("line_code", "")).strip()
            dt = _to_date(row.get("plan_month")) or _to_date(row.get("plan_date"))
            if lc and dt:
                hc_pairs.add((lc, (dt.year, dt.month)))
    except Exception:
        return

    missing_pairs = calendar_pairs - hc_pairs
    if not missing_pairs:
        _write(cursor, batch_file_id, stage_num, stage_name, "PASS",
               f"headcount_plan covers all {len(calendar_pairs)} line-month combinations "
               f"in line_capacity_calendar")
        return

    # Summarise by line — avoids thousands of individual result rows
    from collections import defaultdict
    missing_by_line: dict[str, int] = defaultdict(int)
    for lc, _ in missing_pairs:
        missing_by_line[lc] += 1

    for lc in sorted(missing_by_line.keys())[:50]:
        count = missing_by_line[lc]
        _write(cursor, batch_file_id, stage_num, stage_name, "WARNING",
               f"Line '{lc}': missing planned_headcount for {count} month(s) "
               f"that exist in line_capacity_calendar",
               field_name="headcount_coverage", sample_value=lc)


def _stage8_portfolio_in_plan(cursor, conn, batch_id, batch_file_id, stored_file_path):
    """Warn when a phase-in SKU is absent from the production plan.

    The phase-in panel highlights each SKU's volume/hours as read from
    production_orders. A phase-in SKU with no orders shows zero impact — usually a
    sign the production plan (MRP) needs updating to reflect the launch.
    """
    stage_num, stage_name = 8, STAGE_NAMES[8]

    po_path = _get_sibling_path(conn, batch_id, "production_orders")
    if not po_path:
        _write(cursor, batch_file_id, stage_num, stage_name, "INFO",
               "production_orders not yet uploaded — "
               "portfolio plan-coverage check pending. Run Re-validate when all files are uploaded.")
        return

    plan_items = _read_file_column(po_path, "production_orders", "material")

    # Flagged launches + phase-outs from this portfolio file
    pc_schema = FILE_SCHEMAS["portfolio_changes"]
    try:
        wb = openpyxl.load_workbook(stored_file_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        headers = _get_headers(ws, pc_schema["header_row"])
        rows = _get_data_rows(ws, headers, pc_schema["data_start_row"])
    except Exception:
        return

    flagged: list[str] = []
    for _, row in rows:
        item_code = str(row.get("item_code", "")).strip()
        if item_code:
            flagged.append(item_code)

    missing = sorted({i for i in flagged if i not in plan_items})

    if not missing:
        _write(cursor, batch_file_id, stage_num, stage_name, "PASS",
               "All phase-in SKUs appear in the production plan (production_orders)")
        return

    for item_code in missing[:50]:
        _write(cursor, batch_file_id, stage_num, stage_name, "WARNING",
               f"Phase-in SKU '{item_code}' has no orders in production_orders — it will show "
               f"zero volume/hours impact. Add it to the production plan (MRP).",
               field_name="portfolio_plan_coverage", sample_value=item_code)


def _stage8_material_vs_master_stock(cursor, conn, batch_id, batch_file_id,
                                      stored_file_path, file_type, material_col):
    """Warn about materials in this file that are not present in the batch's master_stock.

    master_stock is the authoritative per-batch SKU list (fresh SAP export). Any material
    code not found there is either a new SKU not yet in master_stock, or a data entry error.
    """
    stage_num, stage_name = 8, STAGE_NAMES[8]

    ms_path = _get_sibling_path(conn, batch_id, "master_stock")
    if not ms_path:
        _write(cursor, batch_file_id, stage_num, stage_name, "INFO",
               "master_stock not yet uploaded — SKU reference check pending. "
               "Run Re-validate once master_stock is uploaded.")
        return

    ms_materials = _read_file_column(ms_path, "master_stock", "material")
    if not ms_materials:
        _write(cursor, batch_file_id, stage_num, stage_name, "INFO",
               "master_stock could not be read — SKU reference check skipped.",
               field_name="sku_reference")
        return

    this_materials = _read_file_column(stored_file_path, file_type, material_col)
    unknown = sorted(this_materials - ms_materials)

    if not unknown:
        _write(cursor, batch_file_id, stage_num, stage_name, "PASS",
               f"All {len(this_materials)} material codes are present in master_stock")
        return

    for item_code in unknown[:50]:
        _write(cursor, batch_file_id, stage_num, stage_name, "WARNING",
               f"Material '{item_code}' not found in master_stock — "
               f"upload an updated master_stock that includes this SKU",
               field_name="sku_reference", sample_value=item_code)
    if len(unknown) > 50:
        _write(cursor, batch_file_id, stage_num, stage_name, "WARNING",
               f"{len(unknown) - 50} more unknown materials not shown",
               field_name="sku_reference")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write(cursor, batch_file_id: int, stage_num: int, stage_name: str,
           severity: str, message: str, *,
           field_name: str | None = None,
           row_number: int | None = None,
           sample_value: str | None = None) -> None:
    cursor.execute(
        """
        INSERT INTO dbo.import_validation_results
            (batch_file_id, validation_stage, stage_name, severity,
             field_name, row_number, message, sample_value)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        batch_file_id, stage_num, stage_name, severity,
        field_name, row_number, message,
        sample_value[:500] if sample_value else None,
    )


def _emit_errors(cursor, batch_file_id: int, stage_num: int, stage_name: str,
                 errors: list, *, pass_msg: str) -> None:
    if not errors:
        _write(cursor, batch_file_id, stage_num, stage_name, "PASS", pass_msg)
        return
    for row_num, col, severity, msg, sample in errors[:20]:
        _write(cursor, batch_file_id, stage_num, stage_name, severity, msg,
               field_name=col, row_number=row_num, sample_value=sample)
    if len(errors) > 20:
        _write(cursor, batch_file_id, stage_num, stage_name, "BLOCKED",
               f"{len(errors) - 20} more error(s) not shown — fix the file and re-upload")


def _update_file_status(cursor, batch_file_id: int) -> None:
    """Set validation_status to worst severity from stages 2–6 only.
    Stage 7 (batch-level) and stage 8 (cross-file WARNINGs) do not affect per-file status."""
    cursor.execute(
        """
        UPDATE dbo.import_batch_files
        SET validation_status = CASE
            WHEN EXISTS (
                SELECT 1 FROM dbo.import_validation_results
                WHERE batch_file_id = ? AND validation_stage BETWEEN 2 AND 6 AND severity = 'BLOCKED'
            ) THEN 'BLOCKED'
            WHEN EXISTS (
                SELECT 1 FROM dbo.import_validation_results
                WHERE batch_file_id = ? AND validation_stage BETWEEN 2 AND 6 AND severity = 'WARNING'
            ) THEN 'WARNING'
            ELSE 'PASS'
        END
        WHERE batch_file_id = ?
        """,
        batch_file_id,
        batch_file_id,
        batch_file_id,
    )


def _strip_sap_uom(val):
    """Strip SAP unit-of-measure suffix from cells like '0 EA' or '5.000 ERR'."""
    if not isinstance(val, str):
        return val
    parts = val.strip().split()
    if len(parts) == 2:
        try:
            return float(parts[0].replace(',', ''))
        except ValueError:
            pass
    return val


