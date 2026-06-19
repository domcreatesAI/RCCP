"""
Generates downloadable Excel templates for the 4 template-based planning files.

demand_plan comes from the SAP PIR export — the entry here shows the expected format for reference.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.services.reasons import CAPACITY_DOWNTIME_REASONS, HEADCOUNT_ABSENCE_REASONS

# ---------------------------------------------------------------------------
# Template definitions
# Each entry: (header_label, column_key, description, sample_value)
# ---------------------------------------------------------------------------

TEMPLATES: dict[str, dict] = {
    "line_capacity_calendar": {
        "title": "Line Capacity Calendar",
        "description": (
            "One row per production line per day, covering the full planning horizon "
            "(typically 12–18 months). planned_hours = the scheduled shift for that line "
            "that day. downtime_hours = lost time that SUBTRACTS from the shift "
            "(available = planned − downtime); record a downtime_reason so the app can "
            "report why capacity is reduced. Leave downtime 0 on normal days."
        ),
        "columns": [
            ("line_code",        "Line Code",        "Production line code (e.g. A101, A202). Must match a line in the masterdata.",   "A101"),
            ("calendar_date",    "Calendar Date",    "Date in DD/MM/YYYY format (e.g. 01/03/2026).",                                   "01/03/2026"),
            ("is_working_day",   "Is Working Day",   "1 = working day, 0 = non-working day (weekend, bank holiday).",                  "1"),
            ("planned_hours",    "Planned Hours",    "Scheduled production hours for this line on this date (0–24). Required.",        "8.0"),
            ("downtime_hours",   "Downtime Hours",   "Hours lost this day (maintenance, breakdown, etc.). Subtracts from planned_hours. Enter 0 if none.", "0"),
            ("downtime_reason",  "Downtime Reason",  "Why the line is down — pick from the list (Breakdown, Maintenance, Stock check, Planned shutdown). Required when downtime_hours > 0.", "Maintenance"),
        ],
        "sample_rows": [
            ["A101", "01/03/2026", 1, 8.0, 0, ""],
            ["A101", "02/03/2026", 0, 0.0, 0, "Weekend"],
            ["A101", "03/03/2026", 1, 8.0, 8, "Maintenance"],
            ["A202", "01/03/2026", 1, 8.0, 4, "Stock check"],
        ],
        "dropdowns": {"downtime_reason": CAPACITY_DOWNTIME_REASONS},
    },
    "headcount_plan": {
        "title": "Pool Headcount",
        "description": (
            "People you ACTUALLY have in each labour POOL, by role, per month. "
            "A pool can span plants (POOL-FLEX = Plants 1/3/4; POOL-P2 = Plant 2). "
            "One row per pool × role × month — use the 1st of the month for plan_month. "
            "Covers ALL roles (operators, line leaders, palletisers AND shared roles like "
            "forklift drivers). The engine compares this against the demand-driven crew need "
            "to flag staffing gaps. Known absences (holiday/sick) go on the 'Exceptions' sheet."
        ),
        "columns": [
            ("pool_code",          "Pool Code",          "Labour pool (e.g. POOL-FLEX, POOL-P2). Must match a pool in labour_pools.",            "POOL-FLEX"),
            ("resource_type_code", "Role Code",          "Any role: LINE_OPERATOR, LINE_LEADER, PALLETISING_OPERATOR, FORKLIFT_DRIVER, MATERIAL_HANDLER, ROBOT_OPERATOR, TECHNICIAN.", "LINE_OPERATOR"),
            ("plan_month",         "Plan Month",         "1st of the month in DD/MM/YYYY format (e.g. 01/05/2026 for May 2026). Required.",      "01/05/2026"),
            ("planned_headcount",  "Planned Headcount",  "People available in the pool for this role that month. Required. Must be ≥ 0.",        "12"),
        ],
        "sample_rows": [
            ["POOL-FLEX", "LINE_OPERATOR",        "01/05/2026", 12],
            ["POOL-FLEX", "LINE_LEADER",          "01/05/2026", 3],
            ["POOL-FLEX", "PALLETISING_OPERATOR", "01/05/2026", 2],
            ["POOL-P2",   "LINE_OPERATOR",        "01/05/2026", 2],
        ],
    },
    "portfolio_changes": {
        "title": "Phase-in",
        "description": (
            "Lists the SKUs being phased in (new launches) this cycle. REQUIRED every cycle, but may "
            "contain zero data rows if there are no phase-ins — upload an empty file (header row only). "
            "This file carries NO volumes: the monthly volume & hours impact is derived from the "
            "production plan (production_orders) for the SKUs you list, and shown for information only "
            "on the Executive Summary. Every SKU listed should also appear in the production plan (MRP) "
            "— if it doesn't, validation raises a WARNING (it does not block publish)."
        ),
        "columns": [
            ("effective_date", "Effective Date", "Month the phase-in takes effect (DD/MM/YYYY).",                                                                     "01/04/2026"),
            ("item_code",      "Item Code",      "SKU / item code being phased in. Required — must exist in the SKU Masterdata. Its production-plan volume drives the impact shown.", "101221"),
            ("line_code",      "Line Code",      "Production line affected. Required — the line the SKU is being phased in on.",                                       "A101"),
            ("comments",       "Comments",       "Brief description of the change. Optional.",                                                                        "New 1L SKU launch"),
        ],
        "sample_rows": [
            ["01/04/2026", "101221", "A101", "New 1L SKU launch"],
            ["01/09/2026", "101245", "A305", "Autumn range addition"],
        ],
    },
    "master_stock": {
        "title": "Master Stock",
        "no_desc_row": True,   # SAP export has headers at row 1 — no amber description row
        "description": (
            "SAP master stock report — export directly and upload without modification."
            "One row per material per plant/warehouse. "
            "material must already exist in the SKU Masterdata — upload sku_masterdata first. "
            "Stock values may include a unit suffix (e.g. '1200 EA') — stripped automatically on import. "
            "abc_indicator is optional but recommended — used by the capacity planning filter in Settings."
        ),
        "columns": [
            ("material",             "Material",             "SAP material number. Must already exist in the SKU Masterdata (dbo.items). Required.",                                              "100000"),
            ("material_description", "Material Description", "SAP material description. Optional — stored in dbo.items on publish.",                                                             "TTA EXALUB AL 46 CAN 5L"),
            ("plant",                "Plant",                "Warehouse/plant code (e.g. UKP1, UKP3). Must match a warehouse in the masterdata. Required.",                                      "UKP1"),
            ("abc_indicator",        "ABC Indicator",        "SAP ABC planning indicator (A/B/C/F/G/L/T/X). Optional — drives the planning filter in Settings. Stored in dbo.items on publish.", "A"),
            ("mrp_type",             "MRP Type",             "SAP MRP planning type (e.g. ZN, PD, ND). Optional — stored in dbo.items on publish.",                                              "ZN"),
            ("unrestrictedstock",    "UnrestrictedStock",    "Total unrestricted stock in eaches. Required. Must be ≥ 0. Unit suffix (e.g. '1200 EA') stripped on import.",                      "1200 EA"),
            ("unrestricted_-_sales", "Unrestricted - Sales", "Stock available for sales (unrestricted minus allocated). Required. May be negative (back-orders). Unit suffix ok.",               "800 EA"),
            ("safety_stock",         "Safety Stock",         "Minimum target stock level in eaches. Required. Must be ≥ 0. Unit suffix ok.",                                                     "200 EA"),
            ("rounding_value",       "Rounding value",       "Units per pallet (EA per pallet). Optional — stored as units_per_pallet in dbo.items on publish. Must be > 0 if provided.",        "120"),
            ("volume",               "Volume",               "Pack volume in litres (e.g. 0.5, 1, 5, 60). Optional — stored as pack_size_l in dbo.items on publish. Must be > 0 if provided.",   "5"),
        ],
        "sample_rows": [
            ["100000", "TTA EXALUB AL 46 CAN 5L", "UKP1", "A", "ZN", "1200 EA", "800 EA", "200 EA", 120, 5.0],
            ["100000", "TTA EXALUB AL 46 CAN 5L", "UKP3", "A", "ZN",  "600 EA", "600 EA",   "0 EA", 120, 5.0],
            ["100001", "TTA STARWAY HT 100 5L",   "UKP1", "B", "ND",    "0 EA",   "0 EA",  "50 EA",  96, 1.0],
        ],
    },
    "production_orders": {
        "title": "Production Orders (SAP COOIS Export)",
        "description": (
            "SAP COOIS production order export — planned orders (LA) and released/firmed orders (YPAC). "
            "Row 1 = field descriptions (amber, ignored by validator). Row 2 = column headers. Row 3+ = data. "
            "Export directly from SAP COOIS transaction and upload without modification. "
            "net_quantity is computed at import as MAX(0, order_quantity - delivered_quantity). "
            "production_line is expected for planned orders (LA); may be blank for released orders (YPAC) — "
            "blank lines generate a WARNING, not a BLOCKED."
        ),
        "columns": [
            ("order",                      "Order",                      "Unique SAP order number. Required.",                                                                  "5104801"),
            ("material",                   "Material",                   "SAP material number. Must match an item in the items masterdata. Required.",                         "100556"),
            ("material_description",       "Material description",       "SKU description from SAP. Optional — stored for reference only.",                                   "MOBIL BRAKE FLUID DOT4 AP C2/3 12x0.5L"),
            ("order_type",                 "Order Type",                 "LA = MRP planned order. YPAC = firmed or released order. Required.",                                 "LA"),
            ("mrp_controller",             "MRP controller",             "SAP MRP controller code (e.g. 005 = lubes, 006 = chemicals). Optional.",                            "006"),
            ("plant",                      "Plant",                      "SAP plant code. Must match a warehouse in the masterdata (e.g. UKP1). Required.",                   "UKP1"),
            ("order_quantity_(gmein)",     "Order quantity (GMEIN)",     "Total order quantity in eaches. Must be > 0. Required.",                                             "36480"),
            ("delivered_quantity_(gmein)", "Delivered quantity (GMEIN)", "Quantity already produced/delivered. Must be ≥ 0. net_quantity = MAX(0, order_qty - delivered_qty). Required.", "0"),
            ("unit_of_measure_(=gmein)",   "Unit of measure (=GMEIN)",   "Unit of measure (typically EA). Optional.",                                                          "EA"),
            ("basic_start_date",           "Basic start date",           "Date the order is scheduled to start production (DD/MM/YYYY or YYYY-MM-DD). Required.",              "27/03/2026"),
            ("basic_finish_date",          "Basic finish date",          "Expected completion date. Optional — not stored.",                                                   "30/03/2026"),
            ("system_status",              "System Status",              "REL = released, CRTD = firmed, blank = MRP proposal. Optional.",                                    ""),
            ("production_version",         "Production Version",         "SAP production version. Optional — not stored.",                                                    "0001"),
            ("entered_by",                 "Entered By",                 "SAP user who created the order. Optional — not stored.",                                            ""),
            ("production_line",            "Production line",            "Production line code (e.g. A304). Populated for LA orders by MRP; may be blank for YPAC. Optional.", "A304"),
        ],
        "sample_rows": [
            ["5104801", "100556", "MOBIL BRAKE FLUID DOT4 AP C2/3 12x0.5L", "LA",   "006", "UKP1", 36480, 0,     "EA", "27/03/2026", "30/03/2026", "",    "0001", "", "A304"],
            ["5104802", "100557", "MOBIL BRAKE FLUID DOT4 AP C2/3 12x1L",   "LA",   "006", "UKP1", 24000, 0,     "EA", "15/04/2026", "16/04/2026", "",    "0001", "", "A305"],
            ["5104750", "100558", "MOBIL HYDRAULIC OIL 68 12x0.5L",         "YPAC", "005", "UKP1", 12000, 8000,  "EA", "03/03/2026", "03/03/2026", "REL", "0001", "", ""],
        ],
    },
    "demand_plan": {
        "title": "Demand Plan (SAP PIR Export)",
        "description": (
            "SAP Planned Independent Requirements (PIR) export — wide format, one row per material per plant. "
            "Row 1 = field descriptions (amber, ignored by validator). Row 2 = column keys. Row 3+ = data. "
            "Columns material_id and plant are required. Columns mrp_area, version, req_type, version_active, "
            "req_plan, req_seg, uom are present in the export but ignored by the validator. "
            "Month columns follow in M{MM}.{YYYY} format (e.g. M03.2026) — always 12 months, columns J to U. "
            "Filter to UK plants (UKP1, UKP3) before uploading — non-UK plants will be rejected."
        ),
        "columns": [
            ("material_id",    "material_id",    "SAP material number. Must match an item in the items masterdata.",                                  "100000"),
            ("plant",          "plant",          "SAP plant code (UK only: UKP1, UKP3). Must match a warehouse in the masterdata.",                   "UKP1"),
            ("mrp_area",       "mrp_area",       "SAP MRP area — same as Plant in most cases. Present in export, ignored by validator.",              "UKP1"),
            ("version",        "version",        "SAP planning version. Present in export, ignored by validator.",                                    "00"),
            ("req_type",       "req_type",       "SAP requirement type (e.g. VSF = forecast). Present in export, ignored by validator.",              "VSF"),
            ("version_active", "version_active", "Whether this planning version is active. Present in export, ignored by validator.",                  "Yes"),
            ("req_plan",       "req_plan",       "SAP requirements plan. Usually blank. Present in export, ignored by validator.",                    ""),
            ("req_seg",        "req_seg",        "SAP requirements segment. Usually blank. Present in export, ignored by validator.",                  ""),
            ("uom",            "uom",            "Unit of measure (always EA). Present in export, ignored by validator.",                             "EA"),
            ("m03.2026",       "M03.2026",       "Demand quantity in eaches for March 2026. 12 monthly columns follow (M{MM}.{YYYY} format).",        500),
            ("m04.2026",       "M04.2026",       "Demand quantity in eaches for April 2026.",                                                         480),
            ("m05.2026",       "M05.2026",       "Demand quantity in eaches for May 2026 — continue for all 12 rolling months.",                      520),
        ],
        "sample_rows": [
            ["100000", "UKP1", "UKP1", "00", "VSF", "Yes", "", "", "EA", 500, 480, 520],
            ["100000", "UKP3", "UKP3", "00", "VSF", "Yes", "", "", "EA", 200, 190, 210],
            ["100001", "UKP1", "UKP1", "00", "VSF", "Yes", "", "", "EA",  60,  55,  70],
        ],
    },

    "actual_production": {
        "title": "Actual Production (SAP MB51 Export)",
        "description": (
            "SAP MB51 goods receipt export for movement type 101 (GR from production order). "
            "One row per goods receipt posting. Row 1 = column headers. Row 2+ = data. "
            "Export directly from SAP MB51 and upload without modification. "
            "Quantities are in EA — converted to litres on publish using pack_size_l from SKU Masterdata. "
            "If the mvt column is present, only rows with mvt = 101 are imported. "
            "material must already exist in the SKU Masterdata. plant must match a warehouse in masterdata."
        ),
        "no_desc_row": True,
        "columns": [
            ("material",             "Material",             "SAP material number. Must already exist in the SKU Masterdata. Required.",                "100556"),
            ("material_description", "Material Description", "SKU description from SAP. Optional — ignored on import.",                               "MOBIL BRAKE FLUID DOT4 AP C2/3 12x0.5L"),
            ("quantity",             "Quantity",             "Goods receipt quantity in EA (eaches). Must be > 0. Required.",                          "960"),
            ("posting_date",         "Posting Date",         "Date the goods receipt was posted in SAP (DD/MM/YYYY or YYYY-MM-DD). Required.",         "24/04/2026"),
            ("movement_type",        "Movement type",        "SAP movement type. Only rows with 101 (GR from production order) are imported.",         "101"),
            ("plant",                "Plant",                "SAP plant/warehouse code (e.g. UKP1). Must match a warehouse in masterdata. Required.",  "UKP1"),
            ("storage_location",     "Storage location",     "SAP storage location code. Optional — not stored.",                                      "UK1D"),
            ("material_document",    "Material Document",    "SAP material document number. Optional.",                                                "5000268077"),
            ("document_date",        "Document Date",        "SAP document date. Optional — not stored.",                                              "24/04/2026"),
            ("customer",             "Customer",             "Customer number. Optional — not stored.",                                                ""),
            ("sales_order",          "Sales order",          "Sales order number. Optional — not stored.",                                             ""),
            ("purchase_order",       "Purchase order",       "Purchase order number. Optional — not stored.",                                          ""),
            ("batch",                "Batch",                "SAP batch number. Optional — not stored.",                                               "0000063643"),
            ("supplier",             "Supplier",             "Supplier number. Optional — not stored.",                                                ""),
            ("reference",            "Reference",            "Reference document. Optional — not stored.",                                             "UKP1100001139832"),
            ("order",                "Order",                "SAP production order number. Optional.",                                                 "1016164"),
            ("user_name",            "User Name",            "SAP user who posted the document. Optional — not stored.",                               "DDIC_BTC"),
        ],
        "sample_rows": [
            ["100556", "MOBIL BRAKE FLUID DOT4 AP C2/3 12x0.5L", 960, "24/04/2026", "101", "UKP1", "UK1D", "5000268077", "24/04/2026", "", "", "", "0000063643", "", "UKP1100001139832", "1016164", "DDIC_BTC"],
            ["100557", "MOBIL BRAKE FLUID DOT4 AP C2/3 12x1L",   600, "24/04/2026", "101", "UKP1", "UK1D", "5000268078", "24/04/2026", "", "", "", "",           "", "",                  "1016165", "DDIC_BTC"],
        ],
    },

    # ------------------------------------------------------------------
    # Reference templates (planning aids — download only, not uploaded/validated)
    # ------------------------------------------------------------------
    "headcount_exceptions": {
        "title": "Headcount Exceptions (Reference)",
        "description": (
            "REFERENCE ONLY — a planning aid for known headcount changes vs the standard "
            "(annual leave, sickness, training, vacancies). This is the SAME data as the "
            "'Exceptions' sheet inside the headcount_plan upload — it is NOT uploaded "
            "separately and NOT validated. Maintain it here, then copy the rows into the "
            "Exceptions sheet of headcount_plan before uploading. "
            "For a line-role exception, fill line_code and leave plant_code blank. "
            "For a plant-shared role exception, fill plant_code + resource_type_code and leave line_code blank. "
            "delta_headcount is negative for absences (e.g. -1 = one person out)."
        ),
        "columns": [
            ("line_code",          "Line Code",          "Production line code (e.g. A101). For line-role exceptions; leave blank for plant-shared exceptions.", "A101"),
            ("plant_code",         "Plant Code",         "Plant code (e.g. 'Plant 1'). For plant-shared role exceptions; leave blank for line exceptions.",      ""),
            ("resource_type_code", "Resource Type Code", "Required for plant rows. Optional for line rows — blank applies the delta across all line roles.",      ""),
            ("start_date",         "Start Date",         "First affected date in DD/MM/YYYY format.",                                                             "15/05/2026"),
            ("end_date",           "End Date",           "Last affected date inclusive (DD/MM/YYYY). Same as start for a one-day event.",                         "19/05/2026"),
            ("delta_headcount",    "Delta Headcount",    "Change vs the standard headcount during the range. Negative for absences (e.g. -1 = one person out).", "-1"),
            ("reason",             "Reason",             "Annual leave, sickness, training, etc. Surfaces on the People Fit panel.",                              "Annual leave"),
        ],
        "sample_rows": [
            ["A101", "",        "",                "15/05/2026", "19/05/2026", -1, "Annual leave"],
            ["",     "Plant 1", "FORKLIFT_DRIVER", "01/06/2026", "12/06/2026", -1, "Sickness"],
        ],
        "dropdowns": {"reason": HEADCOUNT_ABSENCE_REASONS},
    },
    "line_capacity_exceptions": {
        "title": "Line Capacity Exceptions (Reference)",
        "description": (
            "REFERENCE ONLY — a planning aid for known losses of line capacity "
            "(breakdown, maintenance, stock check, planned shutdown). Capacity downtime is "
            "actually entered as downtime_hours + downtime_reason in the line_capacity_calendar "
            "upload — this sheet is NOT uploaded separately and NOT validated. "
            "Use it to plan the events, then reflect the lost hours in the calendar before uploading."
        ),
        "columns": [
            ("line_code",          "Line Code",          "Production line code affected (e.g. A101). Must exist in the masterdata.",          "A101"),
            ("event_type",         "Event Type",         "Breakdown, Maintenance, Stock check, or Planned shutdown.",                         "Maintenance"),
            ("start_date",         "Start Date",         "First affected date in DD/MM/YYYY format.",                                         "15/06/2026"),
            ("end_date",           "End Date",           "Last affected date inclusive (DD/MM/YYYY). Same as start for a one-day event.",     "15/06/2026"),
            ("hours_lost_per_day", "Hours Lost / Day",   "How many production hours are lost per day during the event (→ downtime_hours).",   "4"),
            ("reason",             "Reason",             "Free text — what's happening (e.g. Quarterly PM, Annual cleaning).",                "Quarterly PM"),
        ],
        "sample_rows": [
            ["A101", "Maintenance",      "15/06/2026", "15/06/2026", 4, "Quarterly PM"],
            ["A103", "Planned shutdown", "22/06/2026", "26/06/2026", 8, "Annual cleaning"],
        ],
        "dropdowns": {"event_type": CAPACITY_DOWNTIME_REASONS},
    },

    # ------------------------------------------------------------------
    # Masterdata templates (served via GET /api/masterdata/{type}/template)
    # ------------------------------------------------------------------
    "sku_masterdata": {
        "title": "SKU Masterdata",
        "description": (
            "RCCP routing and classification for each SKU. Upload this BEFORE uploading any batch files. "
            "SAP-sourced attributes (description, ABC indicator, MRP type, pack size, rounding value) "
            "are now populated automatically from the master_stock upload — maintain only RCCP config here. "
            "MERGE by item_code: new rows are inserted, existing rows are updated. "
            "Blank cells keep the existing value in the database — partial uploads are valid. "
            "Items are never deleted by this upload — contact an admin to deactivate a SKU. "
            "primary_line_code is the preferred filling line; secondary through quaternary are "
            "capable alternatives (leave blank if not applicable)."
        ),
        "columns": [
            ("item_code",            "Item Code",             "SAP material number — unique key. Required.",                                                                              "100000"),
            ("item_description",     "Item Description",      "SKU description — for reference only. Not written to DB from this file (sourced from master_stock on publish). Optional.", "MOBIL BRAKE FLUID DOT4 12x0.5L"),
            ("moq",                  "MOQ",                   "Minimum order quantity in eaches. 0 = no minimum. Must be ≥ 0 if provided. Optional.",                                   "240"),
            ("pack_type_code",       "Pack Type Code",        "Warehouse capacity category. Must match pack_types masterdata (SMALL_PACK, 60L, BARREL_200L, IBC). Optional.",            "SMALL_PACK"),
            ("sku_status",           "SKU Status",            "SAP lifecycle status. 1 = In Design | 2 = Phasing Out | 3 = Obsolete. Optional.",                                        "1"),
            ("plant_code",           "Plant Code",            "Primary manufacturing plant code (e.g. P1). Must match plants masterdata. Optional.",                                     "P1"),
            ("primary_line_code",    "Primary Line Code",     "Preferred filling line code (e.g. A101). Must match lines masterdata. Optional.",                                         "A101"),
            ("secondary_line_code",  "Secondary Line Code",   "First alternative filling line. Must match lines masterdata. Optional.",                                                  "A102"),
            ("tertiary_line_code",   "Tertiary Line Code",    "Second alternative filling line. Optional — leave blank if not applicable.",                                              ""),
            ("quaternary_line_code", "Quaternary Line Code",  "Third alternative filling line. Optional — leave blank if not applicable.",                                               ""),
            ("unit_cost",            "Unit Cost (£)",         "Standard cost per EA in GBP. Must be ≥ 0 if provided. Optional — leave blank until cost data is available.",             "0.85"),
        ],
        "sample_rows": [
            ["100000", "MOBIL BRAKE FLUID DOT4 12x0.5L", 240, "SMALL_PACK", 1, "P1", "A101", "A102", "", "", 0.85],
            ["100001", "MOBIL BRAKE FLUID DOT4 12x1L",   120, "SMALL_PACK", 1, "P1", "A101", "",     "", "", 0.90],
            ["100002", "MOBIL HYDRAULIC OIL 68 60L",       1, "60L",        2, "P1", "A202", "A203", "", "", 12.50],
        ],
    },
    "line_pack_capabilities": {
        "title": "Line Pack Capabilities",
        "description": (
            "Fill speeds, pack size capabilities, and OEE targets per production line. "
            "Each row = one line + pack size combination. "
            "bottles_per_minute, is_active, and oee_target are optional — leave blank to use system defaults. "
            "oee_target is the OEE assumption for this line/pack combination (e.g. 0.65 = 65%). "
            "If blank, the line-level OEE target is used instead. "
            "Upload replaces ALL existing line pack capability data."
        ),
        "columns": [
            ("line_code",          "Line Code",          "Production line code (e.g. A101). Must exist in the masterdata.",                                          "A101"),
            ("pack_size_l",        "Pack Size (L)",       "Pack size in litres (e.g. 1, 5, 60, 200). Must be > 0.",                                                  "1"),
            ("bottles_per_minute", "Bottles per Minute",  "Fill rate for this line/pack combination. Required. Must be > 0.",                                     "80"),
            ("is_active",          "Is Active",           "1 = this capability is active, 0 = disabled. Optional — defaults to 1.",                                  "1"),
            ("oee_target",         "OEE Target",          "OEE target for this line/pack combination (0–1, e.g. 0.65 = 65%). Optional — leave blank for line default.", "0.65"),
        ],
        "sample_rows": [
            ["A101", 1,   80,  1, 0.65],
            ["A101", 5,   45,  1, 0.60],
            ["A202", 1,   95,  1, 0.70],
            ["A202", 60,  12,  1, 0.75],
        ],
    },
    "line_resource_requirements": {
        "title": "Line Resource Requirements",
        "description": (
            "Headcount required per resource role to run each production line. "
            "Each row = one line + resource type combination. "
            "LINE-scope roles are LINE_OPERATOR, LINE_LEADER and PALLETISING_OPERATOR "
            "(manual end-of-line palletising). "
            "resource_type_code must match a code in the resource_types masterdata table. "
            "Every active line must have a row for every LINE-scope role — use 0 where a "
            "role is not needed on that line (e.g. a line with no manual palletising). "
            "Upload replaces ALL existing line resource requirement data."
        ),
        "columns": [
            ("line_code",          "Line Code",           "Production line code (e.g. A101). Must exist in the masterdata.",                    "A101"),
            ("resource_type_code", "Resource Type Code",  "Role code (LINE_OPERATOR, LINE_LEADER, PALLETISING_OPERATOR). Must exist in resource_types.", "LINE_OPERATOR"),
            ("headcount_required", "Headcount Required",  "Number of people required for this role on this line. Must be ≥ 0 (0 = role not needed on this line).", "3"),
        ],
        "sample_rows": [
            ["A101", "LINE_OPERATOR",        3],
            ["A101", "LINE_LEADER",          1],
            ["A101", "PALLETISING_OPERATOR", 0],
            ["A303", "LINE_OPERATOR",        3],
            ["A303", "LINE_LEADER",          1],
            ["A303", "PALLETISING_OPERATOR", 2],
        ],
    },
    "plant_resource_requirements": {
        "title": "Plant Resource Requirements",
        "description": (
            "Shared headcount required per resource role per manufacturing plant. "
            "These are plant-level roles shared across lines: FORKLIFT_DRIVER, "
            "MATERIAL_HANDLER, ROBOT_OPERATOR and TECHNICIAN. "
            "resource_type_code must match a code in the resource_types masterdata table. "
            "Every active plant must have a row for every PLANT-scope role — use 0 where a "
            "role is not needed at that plant. "
            "Upload replaces ALL existing plant resource requirement data."
        ),
        "columns": [
            ("plant_code",         "Plant Code",          "Manufacturing plant code (e.g. 'Plant 1', 'Plant 2'). Must exist in the masterdata.", "Plant 1"),
            ("resource_type_code", "Resource Type Code",  "Plant-level role code (FORKLIFT_DRIVER, MATERIAL_HANDLER, ROBOT_OPERATOR, TECHNICIAN). Must exist in resource_types.", "FORKLIFT_DRIVER"),
            ("headcount_required", "Headcount Required",  "Number of people required for this role across the whole plant. Must be ≥ 0 (0 = role not needed at this plant).", "2"),
        ],
        "sample_rows": [
            ["Plant 1", "FORKLIFT_DRIVER",   2],
            ["Plant 1", "MATERIAL_HANDLER",  2],
            ["Plant 1", "ROBOT_OPERATOR",    1],
            ["Plant 1", "TECHNICIAN",        1],
        ],
    },
    "warehouse_capacity": {
        "title": "Warehouse Capacity",
        "description": (
            "Maximum pallet positions per pack type per warehouse. "
            "Each row = one warehouse + pack type combination. "
            "warehouse_code must match a code in the warehouses masterdata table. "
            "pack_type_code must match a code in the pack_types masterdata table. "
            "Upload replaces ALL existing warehouse capacity data."
        ),
        "columns": [
            ("warehouse_code",     "Warehouse Code",      "Warehouse code (e.g. UKP1, UKP3, UKP4, UKP5). Must exist in the masterdata.",    "UKP1"),
            ("pack_type_code",     "Pack Type Code",      "Pack type (e.g. SMALL_PACK, 60L, BARREL_200L, IBC). Must exist in pack_types.", "SMALL_PACK"),
            ("max_pallet_capacity","Max Pallet Capacity", "Maximum number of pallet positions available for this pack type. Must be > 0.",  "500"),
        ],
        "sample_rows": [
            ["UKP1", "SMALL_PACK",   500],
            ["UKP1", "60L",          200],
            ["UKP1", "BARREL_200L",  150],
            ["UKP3", "SMALL_PACK",   300],
        ],
    },
    "pack_types": {
        "title": "Pack Types",
        "description": (
            "Reference list of pack type categories used for warehouse capacity planning. "
            "Each row defines one pack type (e.g. SMALL_PACK, 60L, BARREL_200L, IBC). "
            "Uploaded via MERGE — existing codes are updated, new codes are added. "
            "Codes referenced by other data (items, warehouse_capacity) cannot be removed by upload."
        ),
        "columns": [
            ("pack_type_code", "Pack Type Code", "Unique identifier for this pack type (e.g. SMALL_PACK, 60L, BARREL_200L, IBC). Required.", "SMALL_PACK"),
            ("pack_type_name", "Pack Type Name", "Human-readable name for this pack type. Required.",                                          "Small Pack (1L, 2L, 5L)"),
            ("notes",          "Notes",          "Optional notes about this pack type.",                                                       ""),
            ("is_active",      "Is Active",      "1 = active (used in validation), 0 = inactive. Defaults to 1 if omitted.",                  "1"),
        ],
        "sample_rows": [
            ["SMALL_PACK",  "Small Pack (1L, 2L, 5L)", "",  1],
            ["60L",         "60 Litre Drum",           "",  1],
            ["BARREL_200L", "200 Litre Barrel",        "",  1],
            ["IBC",         "IBC",                     "",  1],
        ],
    },
}

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

_HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")   # Dark navy
_HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
_DESC_FILL    = PatternFill("solid", fgColor="FFE066")   # Amber — delete before uploading
_DESC_FONT    = Font(color="7A5700", italic=True, size=9)
_SAMPLE_FONT  = Font(color="444444", size=10)
_THIN_BORDER  = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _line_resource_sample_rows(conn) -> list:
    """Build a full line × LINE-scope-role skeleton from live masterdata.

    Returns one row per active line per active LINE-scope resource type. Each row
    is pre-filled with the line/role's current headcount_required from
    dbo.line_resource_requirements (0 where no row exists yet). This gives the
    user a complete grid — current data, completed to every line × role — that
    satisfies the stage-6 completeness check (every active line must have every
    LINE-scope role) while preserving values they've already entered.
    """
    cur = conn.cursor()
    cur.execute("SELECT line_code FROM dbo.lines WHERE is_active = 1 ORDER BY line_code")
    lines = [str(r[0]).strip() for r in cur.fetchall()]
    cur.execute("SELECT resource_type_code FROM dbo.resource_types WHERE scope = 'LINE' AND is_active = 1")
    roles = [str(r[0]).strip() for r in cur.fetchall()]
    # Preferred display order; unknown roles fall to the end alphabetically.
    order = {"LINE_OPERATOR": 0, "LINE_LEADER": 1, "PALLETISING_OPERATOR": 2}
    roles.sort(key=lambda rc: (order.get(rc, 99), rc))

    # Current values, so the template reflects existing data rather than blank zeros.
    cur.execute("SELECT line_code, resource_type_code, headcount_required FROM dbo.line_resource_requirements")
    existing = {
        (str(r[0]).strip(), str(r[1]).strip()): int(r[2])
        for r in cur.fetchall()
    }

    rows = []
    for line_code in lines:
        for role in roles:
            rows.append([line_code, role, existing.get((line_code, role), 0)])
    return rows


def _plant_resource_sample_rows(conn) -> list:
    """Build a full plant × PLANT-scope-role skeleton from live masterdata.

    Mirror of _line_resource_sample_rows for plant_resource_requirements. One row
    per active plant per active PLANT-scope resource type, pre-filled with the
    current headcount_required from dbo.plant_resource_requirements (0 where no
    row exists yet). Satisfies the stage-6 completeness check (every active plant
    must have every PLANT-scope role) while preserving entered values.
    """
    cur = conn.cursor()
    cur.execute("SELECT plant_code FROM dbo.plants WHERE is_active = 1 ORDER BY plant_code")
    plants = [str(r[0]).strip() for r in cur.fetchall()]
    cur.execute("SELECT resource_type_code FROM dbo.resource_types WHERE scope = 'PLANT' AND is_active = 1 ORDER BY resource_type_code")
    roles = [str(r[0]).strip() for r in cur.fetchall()]

    cur.execute("SELECT plant_code, resource_type_code, headcount_required FROM dbo.plant_resource_requirements")
    existing = {
        (str(r[0]).strip(), str(r[1]).strip()): int(r[2])
        for r in cur.fetchall()
    }

    rows = []
    for plant_code in plants:
        for role in roles:
            rows.append([plant_code, role, existing.get((plant_code, role), 0)])
    return rows


def _line_capacity_sample_rows(conn) -> list:
    """Full Line Capacity Calendar — one row per active line per day, 2026–2030,
    weekends and UK bank holidays set non-working. Shared logic lives in
    app.services.uk_calendar so the template and the standalone script agree."""
    from app.services.uk_calendar import build_calendar_rows
    cur = conn.cursor()
    cur.execute("SELECT line_code FROM dbo.lines WHERE is_active = 1 ORDER BY line_code")
    lines = [str(r[0]).strip() for r in cur.fetchall()]
    if not lines:
        return []
    return build_calendar_rows(lines)


def _headcount_months() -> list[str]:
    """First-of-month DD/MM/YYYY strings across the planning horizon (2026–2030)."""
    out, y, m = [], 2026, 1
    while (y, m) <= (2030, 12):
        out.append(f"01/{m:02d}/{y}")
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return out


def _hc_num(v):
    """Whole numbers as int (3.0 → 3); keep fractional (3.5)."""
    return int(v) if float(v).is_integer() else float(v)


def _pool_headcount_sample_rows(conn) -> list:
    """Complete Pool Headcount — every labour pool × role × month, fully-staffed
    default. Covers ALL roles: line roles (Σ per-line crew across the pool's lines)
    and shared roles (Σ plant requirement across the pool's plants). Edit DOWN to
    the people actually available."""
    cur = conn.cursor()
    # Line-scope roles: sum per-line crew across each pool's lines.
    cur.execute("""
        SELECT l.labour_pool_code, lrr.resource_type_code, SUM(lrr.headcount_required)
        FROM dbo.line_resource_requirements lrr
        JOIN dbo.lines l ON l.line_code = lrr.line_code
        WHERE l.labour_pool_code IS NOT NULL
        GROUP BY l.labour_pool_code, lrr.resource_type_code
    """)
    combos = {(str(r[0]).strip(), str(r[1]).strip()): _hc_num(r[2]) for r in cur.fetchall()}
    # Shared (plant-scope) roles: sum plant requirement across each pool's plants.
    cur.execute("""
        SELECT pl.labour_pool_code, prr.resource_type_code, SUM(prr.headcount_required)
        FROM dbo.plant_resource_requirements prr
        JOIN (SELECT DISTINCT labour_pool_code, plant_code FROM dbo.lines
              WHERE labour_pool_code IS NOT NULL) pl ON pl.plant_code = prr.plant_code
        WHERE prr.headcount_required > 0
        GROUP BY pl.labour_pool_code, prr.resource_type_code
    """)
    for r in cur.fetchall():
        combos[(str(r[0]).strip(), str(r[1]).strip())] = _hc_num(r[2])

    months = _headcount_months()
    # Month-major ordering: each month is a contiguous block of all pool×role rows.
    rows = []
    for mo in months:
        for (pool_code, role), hc in sorted(combos.items()):
            rows.append([pool_code, role, mo, hc])
    return rows


def generate_template(file_type: str, conn=None) -> bytes:
    """Return an in-memory .xlsx file for the given file_type.

    If conn is provided, line_resource_requirements and plant_resource_requirements
    are pre-populated with a full role-coverage skeleton from live masterdata,
    line_capacity_calendar with the full 2026–2030 working calendar, and
    headcount_plan (Pool Headcount + Plant Support sheets) fully across all
    plants/roles × months.
    """
    if file_type not in TEMPLATES:
        raise ValueError(f"No template defined for file_type '{file_type}'")

    spec = TEMPLATES[file_type]

    # Dynamic skeleton for the resource requirement files when a DB connection is available.
    sample_rows = spec["sample_rows"]
    if conn is not None:
        if file_type == "line_resource_requirements":
            dynamic_rows = _line_resource_sample_rows(conn)
            if dynamic_rows:
                sample_rows = dynamic_rows
        elif file_type == "plant_resource_requirements":
            dynamic_rows = _plant_resource_sample_rows(conn)
            if dynamic_rows:
                sample_rows = dynamic_rows
        elif file_type == "line_capacity_calendar":
            dynamic_rows = _line_capacity_sample_rows(conn)
            if dynamic_rows:
                sample_rows = dynamic_rows
        elif file_type == "headcount_plan":
            dynamic_rows = _pool_headcount_sample_rows(conn)
            if dynamic_rows:
                sample_rows = dynamic_rows

    wb = Workbook()

    # --- Data sheet ---
    ws = wb.active
    ws.title = "Pool Headcount" if file_type == "headcount_plan" else "Data"

    cols = spec["columns"]
    no_desc_row = spec.get("no_desc_row", False)

    if no_desc_row:
        # Row 1: column keys (header only — no amber description row)
        header_row = 1
        data_start = 2
    else:
        # Row 1: field descriptions (amber — validator ignores this row)
        for col_idx, (key, label, desc, sample) in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=col_idx, value=desc)
            cell.font = _DESC_FONT
            cell.fill = _DESC_FILL
            cell.alignment = _LEFT
        header_row = 2
        data_start = 3

    # Header row: column keys — machine-readable header the validator reads
    for col_idx, (key, label, desc, sample) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=key)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER

    # Sample data rows
    for row_offset, sample_row in enumerate(sample_rows, start=data_start):
        for col_idx, val in enumerate(sample_row, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=val)
            cell.font = _SAMPLE_FONT
            cell.alignment = _LEFT

    # Dropdowns (non-strict data validation) on named columns of the Data sheet.
    # spec["dropdowns"] maps a column key to its list of suggested values.
    # showErrorMessage=False = list is suggested but free text is still allowed.
    dropdowns = spec.get("dropdowns")
    if dropdowns:
        key_to_col = {key: i + 1 for i, (key, *_rest) in enumerate(cols)}
        for col_key, values in dropdowns.items():
            if col_key not in key_to_col:
                continue
            letter = get_column_letter(key_to_col[col_key])
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(values) + '"',
                allow_blank=True,
                showErrorMessage=False,
            )
            ws.add_data_validation(dv)
            dv.add(f"{letter}{data_start}:{letter}1000")

    # Column widths
    col_widths = {
        "line_code": 12, "calendar_date": 16, "plan_date": 16,
        "effective_date": 16,
        "is_working_day": 16, "planned_hours": 16,
        "planned_headcount": 18, "shift_code": 12,
        "available_hours": 16, "oee_target": 14,
        "change_type": 18, "item_code": 14,
        "downtime_hours": 16, "downtime_reason": 22,
        "description": 32, "impact_notes": 36, "notes": 28, "comments": 40,
        # exception reference columns
        "resource_type_code": 22, "start_date": 14, "end_date": 14,
        "delta_headcount": 18, "event_type": 18, "hours_lost_per_day": 18,
        "reason": 34,
        # master_stock columns
        "material": 14, "material_description": 40, "plant": 10,
        "abc_indicator": 14, "base_unit_of_measure": 22,
        "unrestrictedstock": 20, "unrestricted_-_sales": 22, "safety_stock": 14,
        # sku_masterdata columns
        "item_code": 14, "item_description": 38,
        "abc_indicator": 14, "mrp_type": 12, "pack_size_l": 14,
        "moq": 10, "pack_type_code": 18, "sku_status": 12,
        "rounding_value": 16, "plant_code": 14,
        "primary_line_code": 20, "secondary_line_code": 22,
        "tertiary_line_code": 20, "quaternary_line_code": 22,
        "unit_cost": 14,
        # demand_plan columns
        "material_id": 14, "mrp_area": 12, "version": 10, "req_type": 10,
        "version_active": 16, "req_plan": 10, "req_seg": 10, "uom": 8,
        "m03.2026": 12, "m04.2026": 12, "m05.2026": 12,
        # production_orders columns
        "order": 12, "material": 14, "material_description": 40,
        "order_type": 12, "mrp_controller": 16,
        "order_quantity_(gmein)": 24, "delivered_quantity_(gmein)": 26,
        "unit_of_measure_(=gmein)": 24, "basic_start_date": 18, "basic_finish_date": 18,
        "system_status": 14, "production_version": 18, "entered_by": 14, "production_line": 16,
        # actual_production columns
        "quantity": 14, "posting_date": 16, "movement_type": 16, "storage_location": 18,
        "material_document": 20, "document_date": 16, "sales_order": 16,
        "purchase_order": 16, "reference": 26, "order": 12, "user_name": 14,
    }
    for col_idx, (key, *_) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(key, 18)

    # Freeze panes below header row
    ws.freeze_panes = f"A{data_start}"

    # (v2: the separate Plant Support sheet is gone — shared roles are part of the
    # single Pool Headcount sheet, since they flex with the pool.)
    if file_type == "headcount_plan":

        # --- Exceptions sheet (planned absences vs the standard headcount) ---
        ex_cols = [
            ("line_code",          "Line Code",           "Production line code (e.g. A101). Required for line-role exceptions; leave blank when entering a plant-shared exception.", "A101"),
            ("plant_code",         "Plant Code",          "Plant code (e.g. 'Plant 1'). Required for plant-shared role exceptions; leave blank for line exceptions.",                  ""),
            ("resource_type_code", "Role Code",           "Required for PLANT rows. Optional for LINE rows — leave blank to apply the delta across all line roles proportionally.",       ""),
            ("start_date",         "Start Date",          "First affected date in DD/MM/YYYY format.",                                                                                  "15/05/2026"),
            ("end_date",           "End Date",            "Last affected date in DD/MM/YYYY format (inclusive). Same as start for a one-day event.",                                    "19/05/2026"),
            ("delta_headcount",    "Delta Headcount",     "Change vs the standard headcount during the date range. Negative for absences (e.g. −1 = one person out).",                "-1"),
            ("reason",             "Reason",              "Free text: annual leave, sickness, training, etc. Surfaces on the People Fit panel.",                                       "Annual leave"),
        ]
        ws3 = wb.create_sheet("Exceptions")

        # Row 1: descriptions
        for col_idx, (_key, _label, desc, _sample) in enumerate(ex_cols, start=1):
            cell = ws3.cell(row=1, column=col_idx, value=desc)
            cell.font = _DESC_FONT
            cell.fill = _DESC_FILL
            cell.alignment = _LEFT

        # Row 2: column keys
        for col_idx, (key, _label, _desc, _sample) in enumerate(ex_cols, start=1):
            cell = ws3.cell(row=2, column=col_idx, value=key)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER
            cell.border = _THIN_BORDER

        # Sample rows
        sample_rows3 = [
            ["A101", "",        "",                "15/05/2026", "19/05/2026", -1,   "Annual leave"],
            ["",     "Plant 1", "FORKLIFT_DRIVER", "01/05/2026", "31/05/2026", -0.5, "Long-term sick"],
        ]
        for row_offset, sample_row in enumerate(sample_rows3, start=3):
            for col_idx, val in enumerate(sample_row, start=1):
                cell = ws3.cell(row=row_offset, column=col_idx, value=val)
                cell.font = _SAMPLE_FONT
                cell.alignment = _LEFT

        ex_widths = {
            "line_code": 14, "plant_code": 14, "resource_type_code": 24,
            "start_date": 14, "end_date": 14, "delta_headcount": 18, "reason": 36,
        }
        for col_idx, (key, *_) in enumerate(ex_cols, start=1):
            ws3.column_dimensions[get_column_letter(col_idx)].width = ex_widths.get(key, 18)
        ws3.freeze_panes = "A3"

        # Dropdown (non-strict) on the reason column. showErrorMessage=False means
        # the list is suggested but users can still type free text (e.g. a specific note).
        reason_col = get_column_letter(len(ex_cols))  # reason is the last column (G)
        reason_dv = DataValidation(
            type="list",
            formula1='"' + ",".join(HEADCOUNT_ABSENCE_REASONS) + '"',
            allow_blank=True,
            showErrorMessage=False,
        )
        reason_dv.prompt = "Pick a standard reason or type your own"
        reason_dv.promptTitle = "Reason"
        ws3.add_data_validation(reason_dv)
        reason_dv.add(f"{reason_col}3:{reason_col}1000")

    # --- Instructions sheet ---
    wi = wb.create_sheet("Instructions")
    wi.column_dimensions["A"].width = 100

    if no_desc_row:
        how_to = [
            (f"1. Fill in your data on the 'Data' sheet starting from row 2.", Font(size=10)),
            ("2. Do NOT modify row 1 — it is the column header the system reads.", Font(size=10)),
            ("3. Save as .xlsx before uploading.", Font(size=10)),
        ]
    else:
        how_to = [
            (f"1. Fill in your data on the 'Data' sheet starting from row 3.", Font(size=10)),
            ("2. Do NOT modify rows 1 or 2 — row 1 is field descriptions, row 2 is the column header the system reads.", Font(size=10)),
            ("3. Row 1 (amber/yellow) is ignored by the validator — leave it in as a reference or delete it if you prefer.", Font(size=10)),
            ("   WARNING: If you delete row 1, what was row 2 shifts to row 1 and the validator will fail to find column names.", Font(size=10)),
            ("4. Save as .xlsx before uploading.", Font(size=10)),
        ]

    instructions = [
        (f"RCCP One — {spec['title']} Template", Font(bold=True, size=13)),
        ("", None),
        (spec["description"], Font(size=10)),
        ("", None),
        ("HOW TO USE THIS TEMPLATE", Font(bold=True, size=11)),
        *how_to,
        ("", None),
        ("COLUMN REFERENCE", Font(bold=True, size=11)),
    ]
    for key, label, desc, sample in cols:
        instructions.append((f"  {key}  |  {label}  —  {desc}  (Example: {sample})", Font(size=10)))

    for row_idx, (text, font) in enumerate(instructions, start=1):
        cell = wi.cell(row=row_idx, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = _LEFT

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
