"""
Build the dashboard data-pack workbook from RCCP engine output.

A defensible audit pack — every figure on the Executive Summary should be
traceable back to a sheet in this file. Used to verify calculations, defend
numbers in the RCCP review meeting, and answer "where did that come from?".

Sheets, in reading order:
  1.  README                      — what each sheet is, units, glossary
  2.  KPIs Summary                — every headline figure (planning month + 12-mo)
  3.  Per Line                    — capacity vs volumes (legacy "verification" view)
  4.  Capacity & Headcount        — line × month: capacity drivers + heads
  5.  Monthly Detail              — line × month: capacity vs volumes with window flag
  6.  Headcount by Line × Month   — needs (heads + FTE) vs planned + gap
  7.  Plant-shared Headcount      — per plant × role × month
  8.  Headcount Exceptions        — absence events with prorated FTE delta
  9.  FTE Breakdown               — the math: role-hours ÷ FTE-month hours
  10. Planned Downtime            — line × month: downtime hours by reason
  11. Phase-in                    — phased-in SKUs: monthly Litres (+ total hrs) from the plan
  12. Action Items                — auto-generated talking points

Used by both the API endpoint (in-memory) and scripts/export_sop_verification.py (to disk).
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Moove brand ────────────────────────────────────────────────────────────────
NAVY = "0C3C5D"
LIME_TINT = "F0F7CC"
WHITE = "FFFFFF"
INK = "0F1A24"

LINE_ORDER = ["A101", "A102", "A103", "A201", "A202", "A302", "A303",
              "A304", "A305", "A307", "A308", "A401", "A501", "A502"]

# Lines hidden from the dashboards — excluded here too so the workbook reconciles
# with what users see. Mirrors HIDDEN_LINE_CODES in frontend/src/components/rccp/brand.ts.
HIDDEN_LINES = {"A501", "A502"}

THIN = Side(style="thin", color="E2E6EA")
BORDER = Border(bottom=THIN)


def _add_months(yyyy_mm: str, n: int) -> str:
    y, m = int(yyyy_mm[:4]), int(yyyy_mm[5:7])
    m += n
    while m > 12:
        m -= 12
        y += 1
    while m < 1:
        m += 12
        y -= 1
    return f"{y:04d}-{m:02d}"


def _line_sort_key(code: str) -> tuple:
    return (LINE_ORDER.index(code) if code in LINE_ORDER else 999, code)


# ─── Shared helpers (mirror frontend brand.ts) ─────────────────────────────────

def _site_fte_month_hours(lines: list, period: str) -> tuple[int, float, float]:
    """site_working_days × shift_hours — calendar-derived 1-FTE envelope.

    Returns (working_days, shift_hours, fte_month_hours). Matches the frontend
    fteSummary helper. Defaults shift_hours to 7 (420 min/day) if no line data.
    """
    site_wd = 0
    shift_mins = 0
    for l in lines:
        m = next((x for x in l["monthly"] if x["period"] == period), None)
        if not m:
            continue
        if (m.get("working_days") or 0) > site_wd:
            site_wd = m["working_days"]
        mins = l.get("available_mins_per_day") or 420
        if mins > shift_mins:
            shift_mins = mins
    shift_hours = shift_mins / 60.0
    return site_wd, shift_hours, site_wd * shift_hours


def _fte_summary(lines: list, plant_support: dict, period: str) -> dict:
    """Per-period FTE needed/planned/gap. Mirrors frontend fteSummary()."""
    site_wd, shift_hours, fte_month_hours = _site_fte_month_hours(lines, period)
    if fte_month_hours <= 0:
        return {"needed": None, "planned": None, "gap": None,
                "month_hours": 0, "working_days": 0, "shift_hours": shift_hours}

    # Plant operating envelope (max line available hours per plant)
    plant_op_hours: dict[str, float] = {}
    for l in lines:
        m = next((x for x in l["monthly"] if x["period"] == period), None)
        if not m:
            continue
        ah = m.get("available_hours") or 0.0
        plant_op_hours[l["plant_code"]] = max(plant_op_hours.get(l["plant_code"], 0.0), ah)

    total_role_hours = 0.0
    total_planned = 0.0

    # Line roles — counted only when the line is actually scheduled
    for l in lines:
        m = next((x for x in l["monthly"] if x["period"] == period), None)
        if not m:
            continue
        prod_h = m.get("production_hours") or 0.0
        if prod_h > 0:
            line_crew = sum(r.get("required", 0) for r in l.get("hc_roles", []))
            total_role_hours += prod_h * line_crew
        if m.get("hc_planned_avg") is not None:
            total_planned += m["hc_planned_avg"]

    # Plant-shared roles — present whenever the plant operates
    for plant_code, roles in (plant_support or {}).items():
        op_h = plant_op_hours.get(plant_code, 0.0)
        for role in roles:
            if op_h > 0:
                total_role_hours += role.get("required", 0) * op_h
            monthly = next((x for x in (role.get("monthly") or []) if x["period"] == period), None)
            if monthly and monthly.get("hc_planned_avg") is not None:
                total_planned += monthly["hc_planned_avg"]

    needed = total_role_hours / fte_month_hours
    return {
        "needed": round(needed, 1),
        "planned": round(total_planned, 1),
        "gap": round(needed - total_planned, 1),
        "month_hours": round(fte_month_hours, 1),
        "working_days": site_wd,
        "shift_hours": shift_hours,
        "total_role_hours": total_role_hours,
        "plant_op_hours": plant_op_hours,
    }


def _plan_feasibility(lines: list, periods: list[str]) -> dict:
    """Plan feasibility + cap/demand totals — matches frontend buildHeadline()."""
    period_set = set(periods)
    prod = avail = dem = 0.0
    firm = planned_v = 0.0
    deliverable = shortfall = 0.0
    any_avail = False
    for l in lines:
        for m in l["monthly"]:
            if m["period"] not in period_set:
                continue
            a = m.get("available_litres")
            p = m.get("production_litres") or 0.0
            d = m.get("demand_litres") or 0.0
            f_ = m.get("firm_litres") or 0.0
            pl = m.get("planned_litres") or 0.0
            if a is not None:
                any_avail = True
                avail += a
                if p > 0:
                    deliverable += min(p, a)
                    shortfall += max(0.0, p - a)
            else:
                deliverable += p
            prod += p
            dem += d
            firm += f_
            planned_v += pl
    return {
        "production_total": prod,
        "firm_total": firm,
        "planned_total": planned_v,
        "deliverable_litres": deliverable,
        "shortfall_litres": shortfall,
        "available_total": avail if any_avail else None,
        "demand_total": dem,
        "plan_feasibility_pct": round((deliverable / prod) * 100, 1) if prod > 0 else None,
        "site_util_theoretical_pct": round((prod / avail) * 100, 1) if (any_avail and avail > 0) else None,
        "demand_cov_pct": round((dem / avail) * 100, 1) if (any_avail and avail > 0) else None,
    }


def _build_action_items(dash: dict, lines: list) -> list[dict]:
    """Auto-generated talking points — same rules as the frontend ActionItemsCard."""
    cycle = dash["plan_cycle_date"][:7]
    horizon = [_add_months(cycle, i) for i in range(12)]
    horizon_set = set(horizon)
    next3 = {_add_months(cycle, i) for i in range(3)}
    cogs = float(dash.get("settings", {}).get("cogs_opex_per_litre") or 0.12)

    items: list[dict] = []

    # 1. Capacity actions — worst overload per line
    for l in lines:
        overs = [m for m in l["monthly"]
                 if m["period"] in horizon_set
                 and m.get("utilisation_pct") is not None
                 and m["utilisation_pct"] > 100]
        if not overs:
            continue
        worst = max(overs, key=lambda m: m["utilisation_pct"])
        prod = worst.get("production_litres") or 0.0
        avail = worst.get("available_litres") or 0.0
        extra_litres = max(0.0, prod - avail)
        ah = worst.get("available_hours") or 0.0
        extra_h = round((extra_litres / avail) * ah) if (avail > 0 and ah > 0) else 0
        cost = extra_litres * cogs
        sev = "critical" if worst["utilisation_pct"] > 115 else "high"
        more = f" (+{len(overs) - 1} more month{'s' if len(overs) > 2 else ''})" if len(overs) > 1 else ""
        items.append({
            "category": "CAPACITY",
            "severity": sev,
            "period": worst["period"],
            "title": f"Approve extra hours on {l['line_code']} — {worst['period']}",
            "detail": f"Order book at {round(worst['utilisation_pct'])}% · need +{extra_h}h to clear{more}",
            "cost_gbp": round(cost),
        })

    # 2. Labour actions — material shortfall in the focus month
    for l in lines:
        if not l.get("material_labour_shortfall"):
            continue
        m = next((x for x in l["monthly"] if x["period"] == cycle), None)
        if not m or (m.get("hc_shortfall") or 0) < 1:
            continue
        items.append({
            "category": "LABOUR",
            "severity": "high",
            "period": cycle,
            "title": f"Resolve {round(m['hc_shortfall'])} FTE labour gap on {l['line_code']} — {cycle}",
            "detail": "Standard crew below the line requirement. Confirm cover with Manufacturing or reschedule the lines.",
            "cost_gbp": None,
        })

    # 3. Portfolio actions — new launches in the next 3 months
    for pc in dash.get("portfolio_changes") or []:
        if pc.get("change_type") != "NEW_LAUNCH":
            continue
        eff = pc.get("effective_period")
        if not eff or eff not in next3:
            continue
        detail = (f"Routing line {pc['line_code']}. Demand will flow through S&OP — sanity-check the line's load that month."
                  if pc.get("line_code")
                  else "No routing line yet — set up the SKU in masterdata before the launch month.")
        items.append({
            "category": "PORTFOLIO",
            "severity": "info",
            "period": eff,
            "title": f"Confirm capacity for {pc.get('item_code') or 'new SKU'} — launching {eff}",
            "detail": detail,
            "cost_gbp": None,
        })

    sev_rank = {"critical": 0, "high": 1, "info": 2}
    items.sort(key=lambda x: (x["period"], sev_rank.get(x["severity"], 9)))
    return items


# ─── Public API ─────────────────────────────────────────────────────────────────
def build_verification_workbook(dash: dict, horizon_months: int = 12) -> Workbook:
    """Build the verification workbook from a compute_dashboard() result."""
    cycle_period = dash["plan_cycle_date"][:7]                       # 'YYYY-MM'
    forward_periods = {_add_months(cycle_period, i) for i in range(horizon_months)}
    # Dashboard charts show actuals for the 3 months before the plan cycle.
    past_periods = {_add_months(cycle_period, -i) for i in (1, 2, 3)}

    lines = sorted(
        (l for l in dash["lines"] if l["line_code"] not in HIDDEN_LINES),
        key=lambda l: _line_sort_key(l["line_code"]),
    )

    summary_rows: list[dict] = []
    detail_rows: list[dict] = []
    for l in lines:
        filled = planned = firmed = capacity = sop = 0.0
        for m in l["monthly"]:
            p = m["period"]
            is_forward = p in forward_periods
            is_past = p in past_periods

            if is_forward:
                window = "forward"
            elif is_past:
                window = "past"            # one of the 3 months shown on the chart
            elif p < cycle_period:
                window = "before"          # older history (not on chart)
            else:
                window = "later"           # beyond the forward horizon
            detail_rows.append({
                "line": l["line_code"],
                "plant": l["plant_code"],
                "period": p,
                "window": window,
                "filled": m.get("actual_litres"),
                "planned": m.get("planned_litres") or 0.0,
                "firmed": m.get("firm_litres") or 0.0,
                "capacity": m.get("available_litres"),
                "sop": m.get("demand_litres") or 0.0,
            })

            if is_past and m.get("actual_litres") is not None:
                filled += m["actual_litres"]
            if is_forward:
                planned += m.get("planned_litres") or 0.0
                firmed += m.get("firm_litres") or 0.0
                capacity += (m.get("available_litres") or 0.0)
                sop += m.get("demand_litres") or 0.0

        summary_rows.append({
            "line": l["line_code"], "plant": l["plant_code"],
            "filled": filled, "planned": planned, "firmed": firmed,
            "capacity": capacity, "sop": sop,
        })

    forward_list = [_add_months(cycle_period, i) for i in range(horizon_months)]
    cogs = float(dash.get("settings", {}).get("cogs_opex_per_litre") or 0.12)
    plant_support = dash.get("plant_support_requirements") or {}

    wb = Workbook()

    # Sheet 1: README (replaces openpyxl's default "Sheet")
    readme = wb.active
    readme.title = "README"
    _build_readme_sheet(readme, dash)

    # Sheet 2: KPIs Summary
    _build_kpis_sheet(wb.create_sheet("KPIs Summary"), dash, lines, plant_support)

    # Sheets 3–5: existing capacity/volume sheets
    _build_summary_sheet(wb.create_sheet("Per Line"), summary_rows, dash["batch_id"], cycle_period, horizon_months, cogs)
    _build_monthly_summary_sheet(wb.create_sheet("Capacity & Headcount"), lines, forward_list)
    _build_detail_sheet(wb.create_sheet("Monthly Detail"), detail_rows, cogs)

    # Sheets 6–11: audit pack sheets
    _build_pool_labour_sheet(wb.create_sheet("Pool Labour"), dash, forward_list)
    _build_exceptions_sheet(wb.create_sheet("Headcount Exceptions"), lines, plant_support)
    _build_downtime_sheet(wb.create_sheet("Planned Downtime"), lines, forward_list)
    _build_portfolio_sheet(wb.create_sheet("Phase-in"), dash)
    _build_actions_sheet(wb.create_sheet("Action Items"), dash, lines)

    return wb


def workbook_bytes(dash: dict, horizon_months: int = 12) -> bytes:
    """Build the workbook and return it as .xlsx bytes (for streaming)."""
    buf = BytesIO()
    build_verification_workbook(dash, horizon_months).save(buf)
    return buf.getvalue()


# ─── Sheet builders ─────────────────────────────────────────────────────────────
def _num(cell, value, *, bold=False, color=INK):
    if value is None:
        cell.value = "—"
    else:
        cell.value = round(value)
        cell.number_format = "#,##0"
    cell.alignment = Alignment(horizontal="right")
    cell.font = Font(name="Calibri", size=11, bold=bold, color=color)


def _pct(cell, value, *, color=INK):
    if value is None:
        cell.value = "—"
    else:
        cell.value = round(value)
        cell.number_format = '0"%"'
    cell.alignment = Alignment(horizontal="right")
    cell.font = Font(name="Calibri", size=11, color=color)


def _dec(cell, value, *, color=INK):
    """One-decimal number (for averaged headcount)."""
    if value is None:
        cell.value = "—"
    else:
        cell.value = round(value, 1)
        cell.number_format = "#,##0.0"
    cell.alignment = Alignment(horizontal="right")
    cell.font = Font(name="Calibri", size=11, color=color)


def _money(cell, value, *, bold=False, color=INK):
    if value is None:
        cell.value = "—"
    else:
        cell.value = round(value)
        cell.number_format = '"£"#,##0'
    cell.alignment = Alignment(horizontal="right")
    cell.font = Font(name="Calibri", size=11, bold=bold, color=color)


def _build_summary_sheet(ws, rows, batch_id, cycle_period, horizon_months, cogs):
    ws.title = "Per Line"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "S&OP Verification — Capacity vs Volumes (litres)"
    ws["A1"].font = Font(name="Calibri", size=15, bold=True, color=NAVY)
    ws["A2"] = (f"Batch {batch_id}  ·  plan cycle {cycle_period}  ·  "
                f"filled = past actuals (MB51)  ·  planned/firmed/capacity/S&OP = forward {horizon_months}M  ·  "
                f"production cost = (firmed + planned) × £{cogs:.2f}/L  ·  "
                f"generated {datetime.now():%d %b %Y %H:%M}")
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="6B7A8A")

    headers = ["Line", "Plant", "Filled volume", "Volume planned",
               "Volume firmed", "Capacity", "S&OP forecast", "Production cost (£)"]
    ncol = len(headers)
    header_row = 4
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="left" if c <= 2 else "right", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color=NAVY))

    r = header_row + 1
    totals = {"filled": 0.0, "planned": 0.0, "firmed": 0.0, "capacity": 0.0, "sop": 0.0}
    for row in rows:
        ws.cell(row=r, column=1, value=row["line"]).font = Font(bold=True, color=NAVY, size=11)
        ws.cell(row=r, column=2, value=row["plant"]).font = Font(color=INK, size=11)
        _num(ws.cell(row=r, column=3), row["filled"])
        _num(ws.cell(row=r, column=4), row["planned"])
        _num(ws.cell(row=r, column=5), row["firmed"])
        _num(ws.cell(row=r, column=6), row["capacity"])
        _num(ws.cell(row=r, column=7), row["sop"])
        _money(ws.cell(row=r, column=8), (row["firmed"] + row["planned"]) * cogs)
        for c in range(1, ncol + 1):
            ws.cell(row=r, column=c).border = BORDER
        for k in totals:
            totals[k] += row[k]
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True, color=NAVY, size=11)
    ws.cell(row=r, column=2, value="")
    _num(ws.cell(row=r, column=3), totals["filled"], bold=True, color=NAVY)
    _num(ws.cell(row=r, column=4), totals["planned"], bold=True, color=NAVY)
    _num(ws.cell(row=r, column=5), totals["firmed"], bold=True, color=NAVY)
    _num(ws.cell(row=r, column=6), totals["capacity"], bold=True, color=NAVY)
    _num(ws.cell(row=r, column=7), totals["sop"], bold=True, color=NAVY)
    _money(ws.cell(row=r, column=8), (totals["firmed"] + totals["planned"]) * cogs, bold=True, color=NAVY)
    for c in range(1, ncol + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = PatternFill("solid", fgColor=LIME_TINT)
        cell.border = Border(top=Side(style="thin", color=NAVY))

    for i, w in enumerate([10, 10, 16, 16, 16, 16, 16, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"


def _build_detail_sheet(ws, rows, cogs):
    ws.sheet_view.showGridLines = False
    headers = ["Line", "Plant", "Period", "Window", "Filled volume",
               "Volume planned", "Volume firmed", "Capacity", "S&OP forecast", "Production cost (£)"]
    ncol = len(headers)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="left" if c <= 4 else "right", vertical="center")

    r = 2
    for row in rows:
        ws.cell(row=r, column=1, value=row["line"]).font = Font(bold=True, color=NAVY, size=10)
        ws.cell(row=r, column=2, value=row["plant"])
        ws.cell(row=r, column=3, value=row["period"])
        wcell = ws.cell(row=r, column=4, value=row["window"])
        if row["window"] == "forward":
            wcell.font = Font(color=NAVY, size=10, bold=True)
        elif row["window"] == "past":
            wcell.font = Font(color="7B9400", size=10, bold=True)
        else:
            wcell.font = Font(color="9CABB9", size=10)
        _num(ws.cell(row=r, column=5), row["filled"])
        _num(ws.cell(row=r, column=6), row["planned"])
        _num(ws.cell(row=r, column=7), row["firmed"])
        _num(ws.cell(row=r, column=8), row["capacity"])
        _num(ws.cell(row=r, column=9), row["sop"])
        _money(ws.cell(row=r, column=10), (row["firmed"] + row["planned"]) * cogs)
        for c in range(1, ncol + 1):
            ws.cell(row=r, column=c).border = BORDER
        r += 1

    for i, w in enumerate([10, 10, 12, 11, 16, 16, 16, 14, 16, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{r - 1}"


def _build_monthly_summary_sheet(ws, lines, periods):
    """Line × month summary of capacity drivers and headcount (forward horizon)."""
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Capacity & Headcount — by line, by month"
    ws["A1"].font = Font(name="Calibri", size=15, bold=True, color=NAVY)
    ws["A2"] = ("Available capacity @ OEE and headcount (required vs planned average) over the forward 12 months. "
                "HC is the total per line (line operators + team leaders); shortfall ≥1 FTE shown in red.")
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="6B7A8A")

    headers = ["Line", "Plant", "Month", "Working days", "Capacity (L)", "Capacity (h)",
               "Utilisation", "HC required", "HC planned (avg)", "HC shortfall"]
    header_row = 4
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="left" if c <= 3 else "right", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color=NAVY))

    r = header_row + 1
    for l in lines:
        bym = {m["period"]: m for m in l["monthly"]}
        for p in periods:
            m = bym.get(p)
            ws.cell(row=r, column=1, value=l["line_code"]).font = Font(bold=True, color=NAVY, size=10)
            ws.cell(row=r, column=2, value=l["plant_code"]).font = Font(color=INK, size=10)
            ws.cell(row=r, column=3, value=p).font = Font(color=INK, size=10)
            _num(ws.cell(row=r, column=4), m.get("working_days") if m else None)
            _num(ws.cell(row=r, column=5), m.get("available_litres") if m else None)
            _num(ws.cell(row=r, column=6), m.get("available_hours") if m else None)
            _pct(ws.cell(row=r, column=7), m.get("utilisation_pct") if m else None)
            _num(ws.cell(row=r, column=8), m.get("hc_required") if m else None)
            _dec(ws.cell(row=r, column=9), m.get("hc_planned_avg") if m else None)
            short = (m.get("hc_shortfall") if m else None)
            _dec(ws.cell(row=r, column=10), short, color="C2410C" if (short or 0) >= 1 else INK)
            for c in range(1, 11):
                ws.cell(row=r, column=c).border = BORDER
            r += 1

    for i, w in enumerate([10, 9, 11, 13, 14, 13, 12, 13, 17, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "D5"
    ws.auto_filter.ref = f"A{header_row}:J{r - 1}"


# ─── New sheet builders for the audit pack ─────────────────────────────────────

def _header_row(ws, row: int, headers: list[str], left_cols: int = 0) -> int:
    """Write a navy-on-white header row; return the next row number."""
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="left" if c <= left_cols else "right", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color=NAVY))
    return row + 1


def _title(ws, title: str, subtitle: str | None = None) -> int:
    """Write title + (optional) subtitle in rows 1–2; return next free row (4)."""
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=15, bold=True, color=NAVY)
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="6B7A8A")
    return 4


def _build_readme_sheet(ws, dash: dict):
    _title(ws,
           "RCCP One — Dashboard data pack",
           f"Batch {dash['batch_id']}  ·  plan cycle {dash['plan_cycle_date'][:7]}  ·  "
           f"generated {datetime.now():%d %b %Y %H:%M}")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 92

    sections: list[tuple[str, str]] = [
        ("", ""),
        ("Purpose", "Every figure on the Executive Summary dashboard should be traceable back to a sheet in this file. "
                    "Open the sheet for the metric in question and you will see the inputs and the math."),
        ("", ""),
        ("Sheet list", ""),
        ("1. README",                    "This sheet."),
        ("2. KPIs Summary",              "Every headline number — planning month and 12-month aggregates."),
        ("3. Per Line",                  "Capacity vs volumes summarised per line over the forward horizon (legacy verification view)."),
        ("4. Capacity & Headcount",      "Line × month grid of capacity drivers + heads-based headcount."),
        ("5. Monthly Detail",            "Line × month: volumes + capacity with a window flag (past / forward / before / later)."),
        ("6. Pool Labour",               "Labour balance per pool × role × month: need vs have vs gap (the v2 headcount model). Pools span plants (POOL-FLEX = Plants 1/3/4, POOL-P2 = Plant 2)."),
        ("7. Headcount Exceptions",      "Known absence events. Prorated delta is the FTE adjustment applied to each affected month."),
        ("8. Planned Downtime",          "Downtime hours per line per month, by reason. Subtracts from available capacity."),
        ("9. Phase-in",                  "Phased-in SKUs with monthly volume (Litres) + total hours from the production plan + the line affected."),
        ("10. Action Items",             "Auto-generated talking points — same logic as the dashboard Action Items card."),
        ("", ""),
        ("Units", ""),
        ("Litres (L)",                   "All volume figures. Capacity, demand, production = litres. £ × COGS_per_litre = OPEX cost."),
        ("Hours (h)",                    "Line operating hours and headcount role-hours. Production hours = runtime needed for the plan."),
        ("FTE",                          "Full-time equivalent. 1 FTE = one person working a standard month "
                                          "(working_days × shift_hours, calendar-derived). Captures part-time, OT, and partial-month operation."),
        ("£",                            "Cost figures use the settings COGS OPEX per litre (default £0.12/L)."),
        ("",                             ""),
        ("Glossary",                     ""),
        ("Plan feasibility",             "% of production deliverable at current capacity = Σ min(prod, avail) ÷ Σ prod."),
        ("Volume to clear",              "Σ max(0, prod − avail) — the unbookable volume that needs OT or extra shift."),
        ("Critical / High",              "Line risk classification. Critical = util > 100% in any month. High = util > 90%."),
        ("Pool labour gap",              "Per pool × role: need − have. Headcount is pooled (people flex across the pool's lines), so gaps are reported per pool, not per line. See the Pool Labour sheet."),
        ("Theoretical capacity",         "Σ available across all lines — the raw ceiling if mix could rebalance. The dashboard's 'optimisation lever' figure."),
    ]
    r = 5
    for label, body in sections:
        c1 = ws.cell(row=r, column=1, value=label)
        c2 = ws.cell(row=r, column=2, value=body)
        if label and not body:
            c1.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        else:
            c1.font = Font(name="Calibri", size=10, bold=True, color=NAVY)
            c2.font = Font(name="Calibri", size=10, color=INK)
        c1.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
        c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        r += 1


def _pool_fte_for_period(pool_labour: dict, period: str) -> dict:
    """Sum pool need/have/gap across all pools & roles for one month (v2 model)."""
    need = have = gap = 0.0
    has_have = False
    for _pool, roles in (pool_labour or {}).items():
        for role in roles:
            m = (role.get("monthly") or {}).get(period)
            if not m:
                continue
            need += m.get("need") or 0.0
            if m.get("have") is not None:
                has_have = True
                have += m["have"]
            if m.get("gap") is not None:
                gap += m["gap"]
    return {
        "need": round(need, 1),
        "have": round(have, 1) if has_have else None,
        "gap": round(gap, 1) if has_have else None,
        "has_have": has_have,
    }


def _build_kpis_sheet(ws, dash: dict, lines: list, plant_support: dict):
    cycle = dash["plan_cycle_date"][:7]
    horizon = [_add_months(cycle, i) for i in range(12)]
    next_row = _title(ws,
                       "KPIs Summary — every headline figure on one page",
                       f"Planning month {cycle}  ·  12-month horizon {horizon[0]} → {horizon[-1]}")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 32

    feas_focus = _plan_feasibility(lines, [cycle])
    feas_12m   = _plan_feasibility(lines, horizon)
    fte_focus  = _fte_summary(lines, plant_support, cycle)
    cogs = float(dash.get("settings", {}).get("cogs_opex_per_litre") or 0.12)

    kpis = dash.get("kpis") or {}

    crit_lines = sum(1 for l in lines if l.get("risk_status") == "Critical")
    high_lines = sum(1 for l in lines if l.get("risk_status") == "High")

    # Pool labour (v2) — the headcount source of truth.
    pool_labour = dash.get("pool_labour") or {}
    pool_focus = _pool_fte_for_period(pool_labour, cycle)
    # Pools short ≥1 FTE on any role in the planning month.
    pools_short = 0
    for _pool, roles in pool_labour.items():
        if any(((r.get("monthly") or {}).get(cycle) or {}).get("gap") is not None
               and ((r.get("monthly") or {}).get(cycle) or {}).get("gap") >= 1
               for r in roles):
            pools_short += 1

    def _section_header(label: str, row: int) -> int:
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for c in (2, 3):
            cc = ws.cell(row=row, column=c)
            cc.fill = PatternFill("solid", fgColor=NAVY)
        return row + 1

    def _safe(text: str | None) -> str:
        # Excel parses leading =, +, -, @ as formula-starters; prefix with a zero-width
        # space when our note text accidentally begins with one. Belt-and-braces.
        if text and text[:1] in ("=", "+", "-", "@"):
            return "​" + text
        return text or ""

    def _row(label: str, value, note: str, row: int, fmt: str = "#,##0") -> int:
        ws.cell(row=row, column=1, value=_safe(label)).font = Font(name="Calibri", size=11, color=INK)
        c2 = ws.cell(row=row, column=2, value=value)
        c2.alignment = Alignment(horizontal="right")
        c2.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        if isinstance(value, (int, float)):
            c2.number_format = fmt
        ws.cell(row=row, column=3, value=_safe(note)).font = Font(name="Calibri", size=10, italic=True, color="6B7A8A")
        for c in (1, 2, 3):
            ws.cell(row=row, column=c).border = BORDER
        return row + 1

    r = next_row
    r = _section_header(f"PLANNING MONTH · {cycle}", r)
    r = _row("Production plan (litres)", feas_focus["production_total"], "Σ firm + MRP this month", r)
    r = _row("  · firm orders (YPAC)",   feas_focus["firm_total"],       "released SAP orders", r)
    r = _row("  · MRP proposals (LA)",   feas_focus["planned_total"],    "planned SAP orders",  r)
    r = _row("Site capacity (litres)",   feas_focus["available_total"],  "Σ available capacity at line mix × OEE", r)
    r = _row("Plan feasibility (%)",     feas_focus["plan_feasibility_pct"], "Σ min(prod, avail) ÷ Σ prod", r, fmt='0"%"')
    r = _row("Deliverable volume (L)",   feas_focus["deliverable_litres"], "what fits in current capacity", r)
    r = _row("Volume to clear (L)",      feas_focus["shortfall_litres"], "shortfall — needs OT or extra shift", r)
    r = _row("Cost to clear (£)",        feas_focus["shortfall_litres"] * cogs, f"shortfall litres × £{cogs:.2f}/L COGS", r, fmt='"£"#,##0')
    r = _row("S&OP demand (litres)",     feas_focus["demand_total"],     "forward demand from PIR upload", r)
    r = _row("Demand coverage (%)",      feas_focus["demand_cov_pct"],   "Σ demand ÷ Σ available", r, fmt='0"%"')
    r += 1

    r = _section_header(f"PLANNING MONTH · headcount (pool FTE)", r)
    r = _row("FTE needed",      pool_focus["need"], "Σ pool need: line roles = crew × util; shared = flat plant req", r, fmt="#,##0.0")
    r = _row("FTE available",   pool_focus["have"] if pool_focus["has_have"] else "—",
             "Σ pool headcount − absences" if pool_focus["has_have"] else "no pool headcount entered yet", r, fmt="#,##0.0")
    r = _row("FTE gap",         pool_focus["gap"] if pool_focus["has_have"] else "—",
             "positive = short (need − available)" if pool_focus["has_have"] else "enter pool headcount to compute", r, fmt="#,##0.0")
    r = _row("Working days", fte_focus["working_days"], "site working days this month (max of lines)", r)
    r = _row("Shift hours",  fte_focus["shift_hours"],  "max line shift hours (available_mins_per_day / 60)", r, fmt="#,##0.0")
    r = _row("1 FTE = X hours", fte_focus["month_hours"], "working_days × shift_hours", r, fmt="#,##0")
    r += 1

    r = _section_header(f"12-MONTH OUTLOOK · {horizon[0]} → {horizon[-1]}", r)
    r = _row("Production plan total (L)",  feas_12m["production_total"],  "Σ firm + MRP over 12 months", r)
    r = _row("Available capacity total (L)", feas_12m["available_total"], "Σ available over 12 months", r)
    r = _row("Plan feasibility (%)",       feas_12m["plan_feasibility_pct"], "across the horizon", r, fmt='0"%"')
    r = _row("Shortfall total (L)",        feas_12m["shortfall_litres"], "Σ max(0, prod − avail) across 12 months", r)
    r = _row("Theoretical site util (%)",  feas_12m["site_util_theoretical_pct"], "Σ prod ÷ Σ avail (optimisation lever)", r, fmt='0"%"')
    r = _row("S&OP demand total (L)",      feas_12m["demand_total"], "Σ S&OP demand over 12 months", r)
    r = _row("Demand vs capacity (%)",     feas_12m["demand_cov_pct"], "12-month S&OP ÷ 12-month capacity", r, fmt='0"%"')
    r += 1

    r = _section_header("RISK & STAFFING (12-month)", r)
    r = _row("Critical lines",                 crit_lines,   "lines with utilisation > 100% in any month", r)
    r = _row("High-risk lines",                high_lines,   "lines > 90% util or with ≥1 FTE labour gap", r)
    r = _row("Pools short of crew",            pools_short, "pools with a ≥1 FTE gap on any role this month", r)
    r = _row("Lines with no data",             kpis.get("lines_with_no_data"), "no capacity figure for any month", r)
    r = _row("Peak utilisation (%)",           kpis.get("peak_util_pct"), f"worst month: {kpis.get('peak_util_period') or '—'}", r, fmt='0.0"%"')
    r = _row("Total capacity gap (L)",         kpis.get("total_gap_litres"), "Σ deficit months × deficit litres (negative gaps only)", r)
    r = _row("Total capacity gap (hours)",     kpis.get("total_gap_hours"),  "as hours, same definition", r)
    r += 1

    r = _section_header("SETTINGS", r)
    r = _row("COGS OPEX per litre (£)", cogs, "from app_settings.cogs_opex_per_litre", r, fmt='"£"0.00')


def _build_pool_labour_sheet(ws, dash: dict, periods: list[str]):
    """Pool labour balance (v2) — per pool × role × month: need vs have vs gap.
    Mirrors the on-screen Staffing Feasibility panel."""
    pool_labour = dash.get("pool_labour") or {}
    pool_info = dash.get("pool_info") or {}

    next_row = _title(ws,
                       "Pool Labour — need vs have vs gap (per pool × role × month)",
                       "Line roles: Σ crew × utilisation. Shared roles: flat plant requirement when the plant runs. "
                       "Have = pool headcount − absences. Gap = need − have (positive = short).")
    headers = ["Pool", "Role", "Scope", "Month", "Need (FTE)", "Have (FTE)", "Gap (FTE)"]
    r = _header_row(ws, next_row, headers, left_cols=4)

    for pool_code in sorted(pool_labour):
        info = pool_info.get(pool_code, {})
        pname = info.get("pool_name", pool_code)
        for role in pool_labour[pool_code]:
            monthly = role.get("monthly") or {}
            for p in periods:
                m = monthly.get(p)
                if not m:
                    continue
                need = m.get("need")
                have = m.get("have")
                gap = m.get("gap")
                ws.cell(row=r, column=1, value=pname).font = Font(bold=True, color=NAVY, size=10)
                ws.cell(row=r, column=2, value=role.get("role_code")).font = Font(color=INK, size=10)
                ws.cell(row=r, column=3, value=role.get("scope") or "").font = Font(color="6B7A8A", size=10)
                ws.cell(row=r, column=4, value=p).font = Font(color=INK, size=10)
                _dec(ws.cell(row=r, column=5), need)
                _dec(ws.cell(row=r, column=6), have)
                _dec(ws.cell(row=r, column=7), gap,
                     color="C2410C" if (gap or 0) >= 1 else ("B45309" if (gap or 0) > 0 else INK))
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).border = BORDER
                r += 1

    # Concurrency note per pool
    r += 1
    ws.cell(row=r, column=1, value="Concurrency (lines you can run, by labour):").font = Font(bold=True, color=NAVY, size=10)
    r += 1
    for pool_code in sorted(pool_info):
        info = pool_info[pool_code]
        ml = info.get("max_concurrent_lines_by_labour")
        txt = f"{info.get('pool_name', pool_code)}: " + (
            f"≈ {ml} lines" + (f" (limited by {info.get('binding_role')})" if info.get("binding_role") else "")
            if ml is not None else "enter pool headcount to compute")
        ws.cell(row=r, column=1, value=txt).font = Font(color=INK, size=10)
        r += 1

    for i, w in enumerate([26, 22, 9, 11, 12, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "E5"


def _build_headcount_by_line_sheet(ws, lines: list, periods: list[str], plant_support: dict):
    next_row = _title(ws,
                       "Headcount by Line × Month — heads + FTE",
                       "Per-line crew demand vs the headcount plan. Crew = LINE_OPERATOR + LINE_LEADER + PALLETISING_OPERATOR per line. "
                       "FTE = (production_hours × crew) ÷ FTE-month hours.")
    headers = ["Line", "Plant", "Month", "Working days", "Production (h)",
               "Crew/line", "Role-hours (h)", "FTE needed", "Heads needed", "Heads planned (avg)", "Heads shortfall"]
    r = _header_row(ws, next_row, headers, left_cols=3)

    fte_hours_cache: dict[str, float] = {}
    for p in periods:
        fte_hours_cache[p] = _site_fte_month_hours(lines, p)[2]

    for l in lines:
        crew = sum(rr.get("required", 0) for rr in l.get("hc_roles", []))
        bym = {m["period"]: m for m in l["monthly"]}
        for p in periods:
            m = bym.get(p)
            if not m:
                continue
            prod_h = m.get("production_hours") or 0.0
            role_hours = prod_h * crew
            fte_hours = fte_hours_cache.get(p, 0.0)
            fte_needed = (role_hours / fte_hours) if fte_hours > 0 else None

            ws.cell(row=r, column=1, value=l["line_code"]).font = Font(bold=True, color=NAVY, size=10)
            ws.cell(row=r, column=2, value=l["plant_code"]).font = Font(color=INK, size=10)
            ws.cell(row=r, column=3, value=p).font = Font(color=INK, size=10)
            _num(ws.cell(row=r, column=4), m.get("working_days"))
            _dec(ws.cell(row=r, column=5), prod_h)
            _num(ws.cell(row=r, column=6), crew)
            _dec(ws.cell(row=r, column=7), role_hours)
            _dec(ws.cell(row=r, column=8), fte_needed)
            _num(ws.cell(row=r, column=9), m.get("hc_required"))
            _dec(ws.cell(row=r, column=10), m.get("hc_planned_avg"))
            short = m.get("hc_shortfall")
            _dec(ws.cell(row=r, column=11), short, color="C2410C" if (short or 0) >= 1 else INK)
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).border = BORDER
            r += 1

    for i, w in enumerate([10, 9, 11, 13, 14, 11, 14, 13, 13, 17, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "D5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{r - 1}"


def _build_plant_shared_sheet(ws, plant_support: dict, periods: list[str], lines: list):
    next_row = _title(ws,
                       "Plant-shared Headcount — per plant × role × month",
                       "Shared crew (forklift, materials handler, robot op, technician). Required and planned, with gap. "
                       "FTE_needed = required × plant_operating_hours ÷ FTE-month hours.")
    headers = ["Plant", "Role", "Month", "Required (heads)", "Plant operating (h)",
               "Role-hours (h)", "FTE needed", "Planned (FTE)", "Heads shortfall"]
    r = _header_row(ws, next_row, headers, left_cols=3)

    fte_hours_cache: dict[str, float] = {p: _site_fte_month_hours(lines, p)[2] for p in periods}

    plant_op_hours_cache: dict[tuple[str, str], float] = {}
    for p in periods:
        for l in lines:
            m = next((x for x in l["monthly"] if x["period"] == p), None)
            if not m:
                continue
            ah = m.get("available_hours") or 0.0
            key = (l["plant_code"], p)
            plant_op_hours_cache[key] = max(plant_op_hours_cache.get(key, 0.0), ah)

    for plant_code, roles in sorted((plant_support or {}).items()):
        for role in roles:
            monthly_by_period = {x["period"]: x for x in (role.get("monthly") or [])}
            for p in periods:
                op_h = plant_op_hours_cache.get((plant_code, p), 0.0)
                role_hours = role.get("required", 0) * op_h
                fte_h = fte_hours_cache.get(p, 0.0)
                fte_needed = (role_hours / fte_h) if fte_h > 0 else None
                m = monthly_by_period.get(p)
                short = m.get("hc_shortfall") if m else None

                ws.cell(row=r, column=1, value=plant_code).font = Font(bold=True, color=NAVY, size=10)
                ws.cell(row=r, column=2, value=role.get("role_code")).font = Font(color=INK, size=10)
                ws.cell(row=r, column=3, value=p).font = Font(color=INK, size=10)
                _num(ws.cell(row=r, column=4), role.get("required"))
                _dec(ws.cell(row=r, column=5), op_h)
                _dec(ws.cell(row=r, column=6), role_hours)
                _dec(ws.cell(row=r, column=7), fte_needed)
                _dec(ws.cell(row=r, column=8), (m.get("hc_planned_avg") if m else None))
                _dec(ws.cell(row=r, column=9), short, color="C2410C" if (short or 0) >= 1 else INK)
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).border = BORDER
                r += 1

    for i, w in enumerate([10, 22, 11, 17, 18, 14, 13, 15, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "D5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{r - 1}"


def _build_exceptions_sheet(ws, lines: list, plant_support: dict):
    next_row = _title(ws,
                       "Headcount Exceptions — known absences vs the standard",
                       "Each row = one event (annual leave, sickness, training, …). Engine prorates the delta by the working-day overlap "
                       "with the affected month and applies the result to the per-role planned figure.")
    headers = ["Scope", "Code", "Role", "Start", "End", "Delta (FTE)",
               "Affected month", "Prorated delta (FTE)", "Reason"]
    r = _header_row(ws, next_row, headers, left_cols=3)

    # Collect from lines first
    for l in sorted(lines, key=lambda x: _line_sort_key(x["line_code"])):
        for m in l["monthly"]:
            for e in m.get("hc_exceptions") or []:
                ws.cell(row=r, column=1, value="LINE").font = Font(bold=True, color=NAVY, size=10)
                ws.cell(row=r, column=2, value=e.get("code") or l["line_code"]).font = Font(color=INK, size=10)
                ws.cell(row=r, column=3, value=e.get("role") or "(all line roles)").font = Font(color=INK, size=10)
                ws.cell(row=r, column=4, value=e.get("start"))
                ws.cell(row=r, column=5, value=e.get("end"))
                _dec(ws.cell(row=r, column=6), e.get("delta"))
                ws.cell(row=r, column=7, value=m["period"]).font = Font(color=INK, size=10)
                _dec(ws.cell(row=r, column=8), e.get("delta_prorated"))
                ws.cell(row=r, column=9, value=e.get("reason") or "").font = Font(color=INK, size=10)
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).border = BORDER
                r += 1

    # And from plant-shared roles
    for plant_code, roles in sorted((plant_support or {}).items()):
        for role in roles:
            for monthly in role.get("monthly") or []:
                for e in monthly.get("hc_exceptions") or []:
                    ws.cell(row=r, column=1, value="PLANT").font = Font(bold=True, color=NAVY, size=10)
                    ws.cell(row=r, column=2, value=e.get("code") or plant_code).font = Font(color=INK, size=10)
                    ws.cell(row=r, column=3, value=e.get("role") or role["role_code"]).font = Font(color=INK, size=10)
                    ws.cell(row=r, column=4, value=e.get("start"))
                    ws.cell(row=r, column=5, value=e.get("end"))
                    _dec(ws.cell(row=r, column=6), e.get("delta"))
                    ws.cell(row=r, column=7, value=monthly["period"]).font = Font(color=INK, size=10)
                    _dec(ws.cell(row=r, column=8), e.get("delta_prorated"))
                    ws.cell(row=r, column=9, value=e.get("reason") or "").font = Font(color=INK, size=10)
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=r, column=c).border = BORDER
                    r += 1

    if r == next_row + 1:
        ws.cell(row=r, column=1, value="(no exceptions recorded in this batch)").font = Font(italic=True, color="9CABB9", size=10)
        r += 1

    for i, w in enumerate([8, 12, 22, 13, 13, 13, 16, 20, 38], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(r - 1, next_row)}"


def _build_fte_breakdown_sheet(ws, lines: list, plant_support: dict, periods: list[str]):
    next_row = _title(ws,
                       "FTE Breakdown — site-level math per month",
                       "Per month: Σ(production_hours × crew) for line roles + Σ(plant_req × plant_operating_hours) for shared roles, "
                       "all divided by FTE-month hours (calendar-derived).")
    headers = ["Month", "Working days", "Shift hours", "FTE-month hours",
               "Line role-hours", "Plant-shared role-hours", "Total role-hours",
               "FTE needed", "FTE planned", "FTE gap"]
    r = _header_row(ws, next_row, headers, left_cols=1)

    for p in periods:
        fte = _fte_summary(lines, plant_support, p)
        # Line role hours = Σ prod_h × line crew across lines
        line_h = 0.0
        for l in lines:
            m = next((x for x in l["monthly"] if x["period"] == p), None)
            if not m:
                continue
            prod_h = m.get("production_hours") or 0.0
            if prod_h > 0:
                line_h += prod_h * sum(rr.get("required", 0) for rr in l.get("hc_roles", []))
        plant_h = (fte.get("total_role_hours") or 0.0) - line_h

        ws.cell(row=r, column=1, value=p).font = Font(color=INK, size=11)
        _num(ws.cell(row=r, column=2), fte["working_days"])
        _dec(ws.cell(row=r, column=3), fte["shift_hours"])
        _dec(ws.cell(row=r, column=4), fte["month_hours"])
        _dec(ws.cell(row=r, column=5), line_h)
        _dec(ws.cell(row=r, column=6), plant_h)
        _dec(ws.cell(row=r, column=7), (fte.get("total_role_hours") or 0.0))
        _dec(ws.cell(row=r, column=8), fte["needed"])
        _dec(ws.cell(row=r, column=9), fte["planned"])
        gap = fte["gap"]
        gap_cell = ws.cell(row=r, column=10)
        _dec(gap_cell, gap, color="C2410C" if (gap or 0) >= 1 else ("7B9400" if (gap or 0) <= -1 else INK))
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
        r += 1

    for i, w in enumerate([11, 13, 12, 16, 17, 22, 17, 13, 13, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{r - 1}"


def _build_downtime_sheet(ws, lines: list, periods: list[str]):
    next_row = _title(ws,
                       "Planned Downtime — line × month by reason",
                       "Downtime hours recorded in the line capacity calendar, by reason. "
                       "Downtime SUBTRACTS from each line's available capacity.")
    headers = ["Line", "Plant", "Month", "Downtime (h)", "By reason"]
    r = _header_row(ws, next_row, headers, left_cols=3)

    for l in lines:
        bym = {m["period"]: m for m in l["monthly"]}
        for p in periods:
            m = bym.get(p)
            if not m:
                continue
            loss = m.get("loss_hours") or 0.0
            if loss <= 0:
                continue
            bd = m.get("loss_breakdown") or {}
            reasons = " · ".join(
                f"{reason} {round(h, 1)}h"
                for reason, h in sorted(bd.items(), key=lambda kv: kv[1], reverse=True)
                if h > 0
            )
            ws.cell(row=r, column=1, value=l["line_code"]).font = Font(bold=True, color=NAVY, size=10)
            ws.cell(row=r, column=2, value=l["plant_code"]).font = Font(color=INK, size=10)
            ws.cell(row=r, column=3, value=p).font = Font(color=INK, size=10)
            _dec(ws.cell(row=r, column=4), loss, color="B45309" if loss > 0 else INK)
            ws.cell(row=r, column=5, value=reasons).font = Font(color=INK, size=10)
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).border = BORDER
            r += 1

    for i, w in enumerate([10, 9, 11, 12, 46], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "D5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{r - 1}"


def _build_portfolio_sheet(ws, dash: dict):
    next_row = _title(ws,
                       "Phase-in — added volume from the production plan",
                       "Information only — volume & hours are derived from the production plan "
                       "(production_orders) for the phase-in SKUs. Monthly figures are Litres; "
                       "Total (hrs) is the fill-time equivalent.")
    months = list(dash.get("horizon_months") or [])
    text_cols = 4   # item, line, plant, effective month
    headers = (["Item code", "Line", "Plant", "Effective month"]
               + months + ["Total (L)", "Total (hrs)", "Comments"])
    r = _header_row(ws, next_row, headers, left_cols=text_cols)

    # Phase-ins only (skip any legacy DISCONTINUE rows).
    pcs = [pc for pc in (dash.get("portfolio_changes") or []) if pc.get("change_type") != "DISCONTINUE"]
    for pc in pcs:
        monthly = pc.get("monthly") or {}
        ws.cell(row=r, column=1, value=pc.get("item_code") or "—").font = Font(bold=True, color=NAVY, size=10)
        ws.cell(row=r, column=2, value=pc.get("line_code") or "—").font = Font(color=INK, size=10)
        ws.cell(row=r, column=3, value=pc.get("plant_code") or "—").font = Font(color=INK, size=10)
        ws.cell(row=r, column=4, value=pc.get("effective_period") or "—").font = Font(color=INK, size=10)

        col = text_cols + 1
        total_l = 0.0
        total_h = 0.0
        for m in months:
            cell_data = monthly.get(m) or {}
            litres = cell_data.get("litres") or 0.0
            total_l += litres
            total_h += cell_data.get("hours") or 0.0
            cell = ws.cell(row=r, column=col, value=round(litres) if litres else None)
            cell.font = Font(color=INK, size=10)
            cell.alignment = Alignment(horizontal="right")
            col += 1
        tl = ws.cell(row=r, column=col, value=round(total_l) if total_l else None)
        tl.font = Font(bold=True, color=NAVY, size=10)
        tl.alignment = Alignment(horizontal="right")
        col += 1
        th = ws.cell(row=r, column=col, value=round(total_h, 1) if total_h else None)
        th.font = Font(bold=True, color=NAVY, size=10)
        th.alignment = Alignment(horizontal="right")
        col += 1
        ws.cell(row=r, column=col, value=pc.get("description") or "").font = Font(color=INK, size=10)
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
        r += 1

    if not pcs:
        ws.cell(row=r, column=1, value="(no phase-ins this batch)").font = Font(italic=True, color="9CABB9", size=10)
        r += 1

    widths = [14, 13, 10, 16] + [11] * len(months) + [12, 12, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(r - 1, next_row)}"


def _build_actions_sheet(ws, dash: dict, lines: list):
    next_row = _title(ws,
                       "Action Items — talking points for the RCCP review",
                       "Auto-generated from the published batch. Same logic as the dashboard's Action Items card.")
    headers = ["Severity", "Category", "Month", "Title", "Detail", "Estimated cost (£)"]
    r = _header_row(ws, next_row, headers, left_cols=5)

    items = _build_action_items(dash, lines)
    for it in items:
        sev_color = {"critical": "C2410C", "high": "B45309", "info": "6B7A8A"}.get(it["severity"], INK)
        ws.cell(row=r, column=1, value=it["severity"].upper()).font = Font(bold=True, color=sev_color, size=10)
        ws.cell(row=r, column=2, value=it["category"]).font = Font(color=INK, size=10)
        ws.cell(row=r, column=3, value=it["period"]).font = Font(color=INK, size=10)
        ws.cell(row=r, column=4, value=it["title"]).font = Font(bold=True, color=NAVY, size=10)
        ws.cell(row=r, column=5, value=it["detail"]).font = Font(color=INK, size=10)
        if it.get("cost_gbp") is not None:
            _money(ws.cell(row=r, column=6), it["cost_gbp"])
        else:
            ws.cell(row=r, column=6, value="—").alignment = Alignment(horizontal="right")
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
        r += 1

    if not items:
        ws.cell(row=r, column=1, value="(no actions outstanding — plan fits in current capacity)").font = Font(italic=True, color="9CABB9", size=10)
        r += 1

    for i, w in enumerate([12, 13, 11, 52, 60, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(r - 1, next_row)}"
