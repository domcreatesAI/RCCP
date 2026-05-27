import io
from datetime import date

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pydantic import BaseModel

from app.database import get_connection
from app.services.auth_service import get_current_user
from app.services import batch_service, validation_service, publish_service

router = APIRouter(prefix="/batches", tags=["batches"])


class CreateBatchRequest(BaseModel):
    batch_name: str
    plan_cycle_date: date


@router.post("")
def create_batch(
    request: CreateBatchRequest,
    current_user: dict = Depends(get_current_user),
):
    if request.plan_cycle_date.day != 1:
        raise HTTPException(
            status_code=422,
            detail="plan_cycle_date must be the 1st of the month",
        )
    conn = get_connection()
    try:
        return batch_service.create_batch(
            conn,
            request.batch_name,
            request.plan_cycle_date,
            current_user["username"],
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("")
def list_batches(current_user: dict = Depends(get_current_user)):
    conn = get_connection()
    try:
        return batch_service.list_batches(conn)
    finally:
        conn.close()


@router.get("/{batch_id}")
def get_batch(batch_id: int, current_user: dict = Depends(get_current_user)):
    conn = get_connection()
    try:
        batch = batch_service.get_batch(conn, batch_id)
        if not batch:
            raise HTTPException(status_code=404, detail="Batch not found")
        batch["files"] = batch_service.get_batch_file_status(conn, batch_id)
        batch["validation_stages"] = batch_service.get_validation_stage_summary(conn, batch_id)
        return batch
    finally:
        conn.close()


@router.post("/{batch_id}/validate")
def validate_batch(batch_id: int, current_user: dict = Depends(get_current_user)):
    """Re-run validation on all current-version files in the batch."""
    conn = get_connection()
    try:
        batch = batch_service.get_batch(conn, batch_id)
        if not batch:
            raise HTTPException(status_code=404, detail="Batch not found")

        files = batch_service.get_current_files_for_validation(conn, batch_id)
        for f in files:
            validation_service.run_validation(
                conn,
                batch_id,
                f["batch_file_id"],
                f["file_type"],
                f["stored_file_path"],
            )

        batch["files"] = batch_service.get_batch_file_status(conn, batch_id)
        batch["validation_stages"] = batch_service.get_validation_stage_summary(conn, batch_id)
        return batch
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.post("/{batch_id}/publish")
def publish_batch(batch_id: int, current_user: dict = Depends(get_current_user)):
    """Publish a batch: gate check, import planning data, set status to PUBLISHED."""
    conn = get_connection()
    try:
        result = publish_service.publish_batch(conn, batch_id, current_user["username"])
        result["files"] = batch_service.get_batch_file_status(conn, batch_id)
        result["validation_stages"] = batch_service.get_validation_stage_summary(conn, batch_id)
        return result
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/{batch_id}/files/{file_type}/download")
def download_batch_file(
    batch_id: int,
    file_type: str,
    current_user: dict = Depends(get_current_user),
):
    """Download the currently uploaded file for a given file type (from DB; filesystem fallback for pre-migration rows)."""
    conn = get_connection()
    try:
        info = batch_service.get_current_file_for_download(conn, batch_id, file_type)
        if not info:
            raise HTTPException(status_code=404, detail="No uploaded file found for this type")
        filename = info["original_filename"] or f"{file_type}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if info.get("file_content"):
            return StreamingResponse(
                io.BytesIO(info["file_content"]),
                media_type=media_type,
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )
        # Fall back to filesystem for rows uploaded before migration 17
        if info.get("stored_file_path"):
            return FileResponse(
                path=info["stored_file_path"],
                filename=filename,
                media_type=media_type,
            )
        raise HTTPException(status_code=404, detail="File content not available")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/{batch_id}/coverage-report")
def get_coverage_report(batch_id: int, current_user: dict = Depends(get_current_user)):
    """
    Return stage 8 cross-file check findings for a batch, grouped by category.
    Results are WARNING-only and do not affect publish eligibility.
    """
    conn = get_connection()
    try:
        batch = batch_service.get_batch(conn, batch_id)
        if not batch:
            raise HTTPException(status_code=404, detail="Batch not found")
        return batch_service.get_coverage_report(conn, batch_id)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/{batch_id}/validation-report")
def download_validation_report(
    batch_id: int,
    current_user: dict = Depends(get_current_user),
):
    """Download a full validation report for the batch as an Excel file."""
    from collections import defaultdict
    from openpyxl.utils import get_column_letter

    conn = get_connection()
    try:
        batch = batch_service.get_batch(conn, batch_id)
        if not batch:
            raise HTTPException(status_code=404, detail="Batch not found")

        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                f.file_type,
                v.validation_stage,
                v.stage_name,
                v.severity,
                v.field_name,
                v.row_number,
                v.message,
                v.sample_value,
                v.created_at
            FROM dbo.import_validation_results v
            JOIN dbo.import_batch_files f ON f.batch_file_id = v.batch_file_id
            WHERE f.batch_id = ?
            ORDER BY
                CASE v.severity WHEN 'BLOCKED' THEN 1 WHEN 'WARNING' THEN 2 WHEN 'INFO' THEN 3 ELSE 4 END,
                f.file_type,
                v.validation_stage,
                v.validation_id
        """, batch_id)
        rows = cursor.fetchall()

        # Shared styles
        hdr_fill = PatternFill("solid", fgColor="1E3A5F")
        hdr_font = Font(color="FFFFFF", bold=True, size=10)
        center    = Alignment(horizontal="center")
        left_top  = Alignment(horizontal="left", vertical="top", wrap_text=True)
        sev_fill  = {
            "BLOCKED": PatternFill("solid", fgColor="FEE2E2"),
            "WARNING": PatternFill("solid", fgColor="FEF9C3"),
        }

        wb = Workbook()

        # ── Sheet 1: Summary ──────────────────────────────────────────────────
        ws_sum = wb.active
        ws_sum.title = "Summary"

        # Batch info header
        ws_sum.cell(row=1, column=1, value="Batch").font = Font(bold=True, size=11)
        ws_sum.cell(row=1, column=2, value=batch.get("batch_name", ""))
        ws_sum.cell(row=2, column=1, value="Cycle date").font = Font(bold=True)
        ws_sum.cell(row=2, column=2, value=str(batch.get("plan_cycle_date", "")))
        ws_sum.cell(row=3, column=1, value="Report generated").font = Font(bold=True)
        ws_sum.cell(row=3, column=2, value=str(date.today()))

        sum_headers = ["File", "Status", "BLOCKED", "WARNING", "INFO", "PASS"]
        for col, h in enumerate(sum_headers, 1):
            cell = ws_sum.cell(row=5, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center

        counts: dict = defaultdict(lambda: defaultdict(int))
        for r in rows:
            counts[r.file_type][r.severity] += 1

        for row_i, (file_type, sev_counts) in enumerate(sorted(counts.items()), 6):
            worst = "PASS"
            if sev_counts.get("BLOCKED", 0):
                worst = "BLOCKED"
            elif sev_counts.get("WARNING", 0):
                worst = "WARNING"
            ws_sum.cell(row=row_i, column=1, value=file_type)
            ws_sum.cell(row=row_i, column=2, value=worst).alignment = center
            for col, sev in enumerate(["BLOCKED", "WARNING", "INFO", "PASS"], 3):
                cnt = sev_counts.get(sev, 0)
                cell = ws_sum.cell(row=row_i, column=col, value=cnt if cnt else "")
                cell.alignment = center
            if worst in sev_fill:
                for col in range(1, 7):
                    ws_sum.cell(row=row_i, column=col).fill = sev_fill[worst]

        ws_sum.column_dimensions["A"].width = 28
        ws_sum.column_dimensions["B"].width = 12
        for col_letter in ["C", "D", "E", "F"]:
            ws_sum.column_dimensions[col_letter].width = 11

        # ── Sheet 2: Missing SKUs / reference data ────────────────────────────
        # Rows from stage 5 (REFERENCE_CHECK) that have a sample_value = the missing code.
        # After re-validation with the updated service, each missing SKU is a separate row.
        missing_rows = [
            r for r in rows
            if r.validation_stage == 5
            and r.severity in ("BLOCKED", "WARNING")
            and r.sample_value
        ]

        ws_miss = wb.create_sheet("Missing Reference Data")
        miss_headers = ["File", "Severity", "Field (column)", "Missing Value", "Action Required"]
        for col, h in enumerate(miss_headers, 1):
            cell = ws_miss.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center

        if missing_rows:
            for row_i, r in enumerate(missing_rows, 2):
                action = (
                    "Add this SKU to sku_masterdata and re-upload"
                    if "sku_masterdata" in (r.message or "").lower() or "item_code" in (r.field_name or "").lower()
                    else "Fix reference data and re-validate"
                )
                values = [r.file_type, r.severity, r.field_name or "", r.sample_value or "", action]
                fill = sev_fill.get(r.severity)
                for col, val in enumerate(values, 1):
                    cell = ws_miss.cell(row=row_i, column=col, value=val)
                    cell.alignment = left_top
                    if fill:
                        cell.fill = fill
        else:
            # Old summary format — no sample_value stored; surface the summary messages instead
            old_ref_rows = [r for r in rows if r.validation_stage == 5 and r.severity in ("BLOCKED", "WARNING")]
            if old_ref_rows:
                ws_miss.cell(row=2, column=1,
                    value="Re-validate the batch to get the full per-SKU list. Summary below:"
                ).font = Font(italic=True, color="7A5700")
                for row_i, r in enumerate(old_ref_rows, 3):
                    values = [r.file_type, r.severity, r.field_name or "", r.message or "",
                              "Re-validate to expand into per-SKU rows"]
                    fill = sev_fill.get(r.severity)
                    for col, val in enumerate(values, 1):
                        cell = ws_miss.cell(row=row_i, column=col, value=val)
                        cell.alignment = left_top
                        if fill:
                            cell.fill = fill
            else:
                ws_miss.cell(row=2, column=1, value="No reference check failures found.")

        miss_col_widths = [24, 12, 22, 20, 52]
        for col, w in enumerate(miss_col_widths, 1):
            ws_miss.column_dimensions[get_column_letter(col)].width = w
        ws_miss.freeze_panes = "A2"

        # ── Sheet 3: All results ──────────────────────────────────────────────
        ws_all = wb.create_sheet("All Results")
        detail_headers = ["File", "Stage", "Stage Name", "Severity", "Field", "Row #", "Message", "Value"]
        for col, h in enumerate(detail_headers, 1):
            cell = ws_all.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center

        for row_i, r in enumerate(rows, 2):
            values = [
                r.file_type, r.validation_stage, r.stage_name, r.severity,
                r.field_name or "", r.row_number or "", r.message or "", r.sample_value or "",
            ]
            fill = sev_fill.get(r.severity)
            for col, val in enumerate(values, 1):
                cell = ws_all.cell(row=row_i, column=col, value=val)
                cell.alignment = left_top
                if fill:
                    cell.fill = fill

        for col, w in enumerate([22, 7, 26, 11, 20, 7, 80, 22], 1):
            ws_all.column_dimensions[get_column_letter(col)].width = w
        ws_all.freeze_panes = "A2"

        # Stream back
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        batch_name = (batch.get("batch_name") or f"batch_{batch_id}").replace(" ", "_")
        filename = f"validation_report_{batch_name}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.post("/{batch_id}/unpublish")
def unpublish_batch(
    batch_id: int,
    current_user: dict = Depends(get_current_user),
):
    """Move a PUBLISHED batch back to VALIDATED. Files and planning data are preserved.
    Use this to upload actual_production then re-publish."""
    conn = get_connection()
    try:
        batch = batch_service.get_batch(conn, batch_id)
        if not batch:
            raise HTTPException(status_code=404, detail="Batch not found")
        if batch["status"] != "PUBLISHED":
            raise HTTPException(status_code=422, detail="Batch is not PUBLISHED")
        batch_service.unpublish_batch(conn, batch_id)
        return {"status": "VALIDATED"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.delete("/{batch_id}/files")
def reset_batch_files(
    batch_id: int,
    current_user: dict = Depends(get_current_user),
):
    """Delete all uploaded files for a batch and reset its status to DRAFT."""
    conn = get_connection()
    try:
        batch = batch_service.get_batch(conn, batch_id)
        if not batch:
            raise HTTPException(status_code=404, detail="Batch not found")
        if batch["status"] == "PUBLISHED":
            raise HTTPException(status_code=422, detail="Cannot reset a published batch")
        count = batch_service.reset_batch_files(conn, batch_id)
        return {"deleted": count}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()
