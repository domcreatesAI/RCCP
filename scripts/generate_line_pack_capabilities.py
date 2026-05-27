"""
Generate a pre-filled line_pack_capabilities upload template.

One row per (line_code, pack_size_l) combination found in dbo.items.primary_line_code.
Known speeds for A101/A102/A103 are pre-filled.
All other speeds are left blank for the user to fill in.

Run from repo root:
    python scripts/generate_line_pack_capabilities.py
Output: uploads/line_pack_capabilities_template.xlsx
"""

import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Known speeds from existing DB data (A101, A102, A103)
# key: (line_code, pack_size_l) → bottles_per_minute
# ---------------------------------------------------------------------------
KNOWN_SPEEDS = {
    ("A101", 2.0):   44,
    ("A101", 4.0):   34,
    ("A101", 4.5):   None,   # not in DB yet — leave blank
    ("A101", 5.0):   34,
    ("A102", 4.0):   32,
    ("A102", 5.0):   32,
    ("A103", 0.4):   None,   # not in DB yet — leave blank
    ("A103", 0.5):   60,
    ("A103", 0.6):   60,
    ("A103", 1.0):   60,
    # A201–A401: converted from units/day ÷ (420 min/day × 0.55 OEE) = units/day ÷ 231
    ("A201",  60.0):  1.56,
    ("A201", 199.0):  1.04,
    ("A201", 205.0):  1.04,
    ("A201", 208.0):  1.04,
    ("A202",  20.0):  6.93,
    ("A202", 1000.0): 0.14,
    ("A302",  20.0):  1.11,
    ("A302",  25.0):  1.11,
    ("A302", 200.0):  0.35,
    ("A302", 205.0):  0.35,
    ("A302", 208.0):  0.35,
    ("A302", 1000.0): 0.14,
    ("A303",   4.0): 34.63,
    ("A303",   5.0): 34.63,
    ("A304",   0.4): 32.21,
    ("A304",   0.5): 41.56,
    ("A304",   1.0): 41.56,
    ("A305",   0.5): 83.12,
    ("A305",   1.0): 54.03,
    ("A305",   5.0): None,   # in masterdata but speed not provided — leave blank
    ("A307",   3.0):  3.32,
    ("A307",  10.0):  2.08,
    ("A307",  20.0):  1.30,
    ("A308",  20.0):  0.52,
    ("A401",   5.0):  8.66,
}

# ---------------------------------------------------------------------------
# Combinations derived from dbo.items (primary_line_code + pack_size_l)
# Only lines/pack sizes that actually have SKUs assigned.
# ---------------------------------------------------------------------------
LINE_PACK_COMBINATIONS = [
    # (line_code, pack_size_l, sku_count)  — ordered by line, then pack size
    ("A101",  2.0,   16),
    ("A101",  4.0,    3),
    ("A101",  4.5,    1),
    ("A101",  5.0,   99),
    ("A102",  4.0,   21),
    ("A102",  5.0,   48),
    ("A103",  0.4,    1),
    ("A103",  0.5,    3),
    ("A103",  0.6,    1),
    ("A103",  1.0,  123),
    ("A201", 60.0,   26),
    ("A201",199.0,   35),
    ("A201",205.0,   16),
    ("A201",208.0,    4),
    ("A202", 20.0,  108),
    ("A202",1000.0,  15),
    ("A302", 20.0,   25),
    ("A302", 25.0,    1),
    ("A302",200.0,    1),
    ("A302",205.0,    5),
    ("A302",208.0,   17),
    ("A302",1000.0,  12),
    ("A303",  4.0,    2),
    ("A303",  5.0,   25),
    ("A304",  0.4,    9),
    ("A304",  0.5,    4),
    ("A304",  1.0,   21),
    ("A305",  0.5,    7),
    ("A305",  1.0,    7),
    ("A305",  5.0,    1),
    ("A307",  3.0,    2),
    ("A307", 10.0,    1),
    ("A307", 20.0,    1),
    ("A308", 20.0,   13),
    ("A401",  5.0,    2),
    # A501 and A502 — no primary_line_code assignments yet; add rows when known
]

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
HDR_FILL   = PatternFill("solid", fgColor="1E3A5F")
HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
DESC_FILL  = PatternFill("solid", fgColor="FFE066")
DESC_FONT  = Font(color="7A5700", italic=True, size=9)
BLANK_FILL = PatternFill("solid", fgColor="FFF9C4")   # yellow — needs filling
KNOWN_FILL = PatternFill("solid", fgColor="F0FDF4")   # green tint — pre-filled
CENTER     = Alignment(horizontal="center", vertical="center")
LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN = Side(style="thin", color="E2E8F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = ["line_code", "pack_size_l", "bottles_per_minute", "is_active", "oee_target"]
DESCRIPTIONS = [
    "Production line code (e.g. A101). Must exist in the masterdata.",
    "Pack size in litres (e.g. 0.5, 5, 60, 200). Must be > 0.",
    "Fill rate for this line/pack combination (bottles per minute). FILL IN ALL YELLOW CELLS.",
    "1 = active, 0 = disabled. Leave blank to default to 1.",
    "OEE target override for this line/pack combination (e.g. 0.65). Leave blank to use line default.",
]
COL_WIDTHS = [14, 14, 22, 12, 14]


def generate():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Row 1: descriptions
    for col, desc in enumerate(DESCRIPTIONS, 1):
        cell = ws.cell(row=1, column=col, value=desc)
        cell.font      = DESC_FONT
        cell.fill      = DESC_FILL
        cell.alignment = LEFT

    # Row 2: column headers
    for col, name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=2, column=col, value=name)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS[col - 1]

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 40

    # Data rows
    prev_line = None
    for row_i, (line_code, pack_size_l, sku_count) in enumerate(LINE_PACK_COMBINATIONS, 3):
        speed = KNOWN_SPEEDS.get((line_code, pack_size_l))
        has_speed = speed is not None

        ws.cell(row=row_i, column=1, value=line_code).alignment = CENTER
        ws.cell(row=row_i, column=2, value=pack_size_l).alignment = CENTER
        ws.cell(row=row_i, column=3, value=speed).alignment = CENTER
        ws.cell(row=row_i, column=4, value=1).alignment = CENTER     # is_active = 1
        ws.cell(row=row_i, column=5, value=None).alignment = CENTER  # oee_target = blank (use line default)

        # Colour: yellow if speed missing, light green if pre-filled
        speed_fill = KNOWN_FILL if has_speed else BLANK_FILL
        for col in range(1, 6):
            ws.cell(row=row_i, column=col).border = BORDER
        ws.cell(row=row_i, column=3).fill = speed_fill

        # Faint separator between line groups
        if prev_line and prev_line != line_code:
            for col in range(1, 6):
                ws.cell(row=row_i, column=col).border = Border(
                    left=THIN, right=THIN, bottom=THIN,
                    top=Side(style="medium", color="94A3B8"),
                )
        prev_line = line_code

    # Info column — SKU count for reference (column F, not uploaded)
    ws.cell(row=2, column=6, value="SKUs on this line/pack").font = Font(italic=True, color="94A3B8", size=9)
    ws.column_dimensions["F"].width = 22
    for row_i, (_, _, sku_count) in enumerate(LINE_PACK_COMBINATIONS, 3):
        cell = ws.cell(row=row_i, column=6, value=sku_count)
        cell.font      = Font(color="94A3B8", size=9)
        cell.alignment = CENTER

    # ── Instructions sheet ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions["A"].width = 80
    lines = [
        "RCCP One — Line Pack Capabilities Template",
        "",
        "HOW TO USE THIS FILE",
        "1. Fill in bottles_per_minute for every YELLOW cell (speed not yet known).",
        "2. Green cells are pre-filled from existing data — verify they are correct.",
        "3. Column F (SKU count) is for reference only — it is ignored on upload.",
        "4. oee_target column: leave blank to use the line-level OEE default (55%).",
        "   Only fill in if this specific line/pack combination has a different OEE.",
        "5. Upload this file via: Masterdata → Line Pack Capabilities → Upload.",
        "   The upload replaces ALL existing line pack capability data.",
        "",
        "NOTES",
        "- A501 and A502 have no SKUs assigned yet in sku_masterdata.",
        "  Add rows for them once primary_line_code is set on their SKUs.",
        "- litres_per_minute is computed automatically: pack_size_l × bottles_per_minute.",
        "- Lines with no capabilities will show as 'No data' on the RCCP dashboard.",
        "",
        f"Combinations: {len(LINE_PACK_COMBINATIONS)} rows across 12 lines",
        "Lines with no data (A501, A502): add rows when SKUs are assigned.",
    ]
    for i, line in enumerate(lines, 1):
        cell = ws2.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=13)
        elif line.startswith("HOW") or line.startswith("NOTES"):
            cell.font = Font(bold=True, size=10)
        else:
            cell.font = Font(size=10)

    out_dir = os.path.join(os.path.dirname(__file__), "..", "uploads")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "line_pack_capabilities_template.xlsx")
    wb.save(out_path)

    print(f"Saved: {os.path.abspath(out_path)}")
    print(f"Rows:  {len(LINE_PACK_COMBINATIONS)}")
    print()
    print("Pre-filled (green):")
    filled = [(l, p, s) for (l, p, _) in LINE_PACK_COMBINATIONS if (s := KNOWN_SPEEDS.get((l, p))) is not None]
    for l, p, s in filled:
        print(f"  {l:6} {p:6}L  →  {s} bottles/min")
    print()
    print("Needs filling (yellow):")
    blank = [(l, p) for (l, p, _) in LINE_PACK_COMBINATIONS if KNOWN_SPEEDS.get((l, p)) is None]
    for l, p in blank:
        print(f"  {l:6} {p:6}L")


if __name__ == "__main__":
    generate()
