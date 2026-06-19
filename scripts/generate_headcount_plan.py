"""
Generate a Pool Headcount workbook for upload to RCCP One (labour model v2).

Two sheets matching the in-app template:
  Sheet 1 — "Pool Headcount"  pool_code, resource_type_code, plan_month, planned_headcount
  Sheet 2 — "Exceptions"      absences (header only — no rows = no absences)

Covers 01/2026 – 12/2030 (60 months).

Pool Headcount defaults to FULLY STAFFED, by pool × role:
  - line roles  (LINE_OPERATOR, LINE_LEADER, PALLETISING_OPERATOR):
                Σ per-line crew across the pool's lines
  - shared roles (FORKLIFT_DRIVER, MATERIAL_HANDLER, ROBOT_OPERATOR, TECHNICIAN):
                Σ plant requirement across the pool's plants
Edit the numbers DOWN to the people you actually have so real gaps surface.

Run from the repo root:
    python scripts/generate_headcount_plan.py
Output: uploads/headcount_plan_2026_2030.xlsx
"""

import os
import sys
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))
from app.database import get_connection  # noqa: E402

START_YEAR, START_MONTH = 2026, 1
END_YEAR,   END_MONTH   = 2030, 12


def months_in_range() -> list[date]:
    out, y, m = [], START_YEAR, START_MONTH
    while (y, m) <= (END_YEAR, END_MONTH):
        out.append(date(y, m, 1))
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return out


def _num(v):
    return int(v) if float(v).is_integer() else float(v)


def load_pool_defaults():
    """Return {(pool_code, role): fully_staffed_default}."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT l.labour_pool_code, lrr.resource_type_code, SUM(lrr.headcount_required)
        FROM dbo.line_resource_requirements lrr
        JOIN dbo.lines l ON l.line_code = lrr.line_code
        WHERE l.labour_pool_code IS NOT NULL
        GROUP BY l.labour_pool_code, lrr.resource_type_code
    """)
    combos = {(str(r[0]).strip(), str(r[1]).strip()): _num(r[2]) for r in cur.fetchall()}
    cur.execute("""
        SELECT pl.labour_pool_code, prr.resource_type_code, SUM(prr.headcount_required)
        FROM dbo.plant_resource_requirements prr
        JOIN (SELECT DISTINCT labour_pool_code, plant_code FROM dbo.lines
              WHERE labour_pool_code IS NOT NULL) pl ON pl.plant_code = prr.plant_code
        WHERE prr.headcount_required > 0
        GROUP BY pl.labour_pool_code, prr.resource_type_code
    """)
    for r in cur.fetchall():
        combos[(str(r[0]).strip(), str(r[1]).strip())] = _num(r[2])
    conn.close()
    return combos


_HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_DESC_FILL   = PatternFill("solid", fgColor="FFE066")
_DESC_FONT   = Font(color="7A5700", italic=True, size=9)
_CENTER      = Alignment(horizontal="center", vertical="center")
_LEFT        = Alignment(horizontal="left",   vertical="center")


def _write_sheet(wb, sheet_name, is_first, columns, rows) -> None:
    ws = wb.active if is_first else wb.create_sheet(sheet_name)
    ws.title = sheet_name
    for col_idx, (_key, desc, _width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=desc)
        cell.font, cell.fill, cell.alignment = _DESC_FONT, _DESC_FILL, _LEFT
    for col_idx, (key, _desc, width) in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=key)
        cell.font, cell.fill, cell.alignment = _HEADER_FONT, _HEADER_FILL, _CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row_offset, row in enumerate(rows, start=3):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_offset, column=col_idx, value=val)
    ws.freeze_panes = "A3"


def generate() -> None:
    months = months_in_range()
    pool_defaults = load_pool_defaults()
    wb = Workbook()

    # ── Sheet 1 — Pool Headcount (all roles per pool) ────────────────────
    pool_cols = [
        ("pool_code",          "Labour pool (e.g. POOL-FLEX, POOL-P2). Must match labour_pools.",                    14),
        ("resource_type_code", "Any role (LINE_OPERATOR, LINE_LEADER, PALLETISING_OPERATOR, FORKLIFT_DRIVER, …).",    26),
        ("plan_month",         "1st of the month in DD/MM/YYYY (e.g. 01/05/2026).",                                   16),
        ("planned_headcount",  "People you ACTUALLY have in the pool for this role/month. Default = fully staffed; edit DOWN.", 22),
    ]
    pool_rows = [
        [pool_code, role, m.strftime("%d/%m/%Y"), hc]
        for (pool_code, role), hc in sorted(pool_defaults.items())
        for m in months
    ]
    _write_sheet(wb, "Pool Headcount", is_first=True, columns=pool_cols, rows=pool_rows)

    # ── Sheet 2 — Exceptions (header only; pool absences) ────────────────
    exc_cols = [
        ("line_code",          "Production line code (e.g. A101). For line-role exceptions; leave blank for plant-shared.", 12),
        ("plant_code",         "Plant code (e.g. 'Plant 1'). For plant-shared role exceptions; leave blank for line.",      12),
        ("resource_type_code", "Required for plant rows. Optional for line rows.",                                          24),
        ("start_date",         "First affected date in DD/MM/YYYY.",                                                        14),
        ("end_date",           "Last affected date in DD/MM/YYYY (inclusive).",                                            14),
        ("delta_headcount",    "Change vs the standard during the range. Negative for absences (e.g. -1 = one out).",       18),
        ("reason",             "Free text: annual leave, sickness, training, etc.",                                        32),
    ]
    _write_sheet(wb, "Exceptions", is_first=False, columns=exc_cols, rows=[])

    # ── Info ─────────────────────────────────────────────────────────────
    ws_info = wb.create_sheet("Info")
    ws_info.column_dimensions["A"].width = 84
    info = [
        ("RCCP One — Pool Headcount 2026–2030 (labour model v2)", Font(bold=True, size=13)),
        ("", None),
        (f"Generated: {date.today().strftime('%d/%m/%Y')}", Font(size=10)),
        (f"Date range: 01/{START_MONTH:02d}/{START_YEAR} – 12/{END_MONTH:02d}/{END_YEAR}  ({len(months)} months)", Font(size=10)),
        (f"Pool Headcount rows: {len(pool_rows):,}", Font(size=10)),
        ("", None),
        ("Pools span plants: POOL-FLEX = Plants 1/3/4, POOL-P2 = Plant 2 (Plant 5 excluded).", Font(size=10)),
        ("Headcount is per POOL per role — defaults to fully staffed; EDIT DOWN to reality.", Font(size=10)),
        ("Absences go on the Exceptions sheet (by line or plant; mapped to the pool).", Font(size=10)),
        ("", None),
        ("Fully-staffed defaults (pool · role):", Font(bold=True, size=10)),
        *[(f"  {pc} · {rc}: {hc}", Font(size=10)) for (pc, rc), hc in sorted(pool_defaults.items())],
    ]
    for i, (text, font) in enumerate(info, start=1):
        cell = ws_info.cell(row=i, column=1, value=text)
        if font is not None:
            cell.font = font
        cell.alignment = _LEFT

    out_dir = os.path.join(os.path.dirname(__file__), "..", "uploads")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "headcount_plan_2026_2030.xlsx")
    wb.save(out_path)
    print(f"Saved: {os.path.abspath(out_path)}")
    print(f"  Sheet 1 (Pool Headcount): {len(pool_rows):,} rows  ({len(pool_defaults)} pool×role combos × {len(months)} months)")
    print(f"  Sheet 2 (Exceptions):     0 rows")


if __name__ == "__main__":
    generate()
