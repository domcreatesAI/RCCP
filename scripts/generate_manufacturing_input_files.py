"""
Generate two lightweight input templates for Manufacturing to fill in
before each RCCP cycle:

  1. uploads/headcount_exceptions_input.xlsx
     Known absences for the planning month — annual leave, sickness,
     training. Rows go directly into Sheet 3 (Exceptions) of the main
     headcount_plan.xlsx that the planner uploads to RCCP.

  2. uploads/line_capacity_exceptions_input.xlsx
     Planned downtime, maintenance, full shutdowns. Each event describes
     a date range with hours-per-day lost. The planner translates these
     into per-day rows on line_capacity_calendar.xlsx before publishing.

Both files have:
  - an Instructions sheet  (how to fill in)
  - a Data Entry sheet     (column headers + 2 sample rows + dropdowns
                            on line_code / plant_code / role_code / event_type)
  - a Reference sheet      (valid codes so Manufacturing can look them up)

Run from the repo root:
    python scripts/generate_manufacturing_input_files.py
"""

import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ─── Masterdata references (mirror the live DB) ──────────────────────────────
LINES = [
    "A101", "A102", "A103",
    "A201", "A202",
    "A302", "A303", "A304", "A305", "A307", "A308",
    "A401",
    "A501", "A502",
]
PLANTS = ["Plant 1", "Plant 2", "Plant 3", "Plant 4", "Plant 5"]
LINE_ROLES = ["LINE_OPERATOR", "TEAM_LEADER"]
PLANT_ROLES = ["FORKLIFT_DRIVER", "MATERIAL_HANDLER", "ROBOT_OPERATOR", "TECHNICIAN"]
ALL_ROLES = LINE_ROLES + PLANT_ROLES

EVENT_TYPES = ["Maintenance", "Planned downtime", "Public holiday", "Full shutdown", "Other"]


# ─── Styling ─────────────────────────────────────────────────────────────────
NAVY = "0C3C5D"
NAVY_TINT = "E8EEF3"
LIME_TINT = "F0F7CC"
AMBER_TINT = "FEF3C7"
INK = "0F1A24"
INK3 = "6B7A8A"
WHITE = "FFFFFF"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
DESC_FILL = PatternFill("solid", fgColor=AMBER_TINT)
DESC_FONT = Font(name="Calibri", size=9, italic=True, color="7A5700")
SAMPLE_FILL = PatternFill("solid", fgColor=NAVY_TINT)
SAMPLE_FONT = Font(name="Calibri", size=10, color=INK)
THIN = Side(style="thin", color="E2E6EA")
BORDER_BOTTOM = Border(bottom=THIN)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def _write_header_row(ws, row: int, columns: list[tuple[str, str, int]]):
    """columns = [(key, description, col_width)]. Row 1 = description (amber).
    Row 2 = column key (navy bold)."""
    for col_idx, (_key, desc, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=desc)
        cell.font = DESC_FONT
        cell.fill = DESC_FILL
        cell.alignment = LEFT
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for col_idx, (key, _desc, _width) in enumerate(columns, start=1):
        cell = ws.cell(row=row + 1, column=col_idx, value=key)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER


def _write_instruction_sheet(ws, title: str, lines: list[tuple[str, str]]):
    """Each line is (label, body). Empty label/body = blank line."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90

    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=15, bold=True, color=NAVY)

    r = 3
    for label, body in lines:
        if label:
            ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", size=10, bold=True, color=NAVY)
        if body:
            ws.cell(row=r, column=2, value=body).font = Font(name="Calibri", size=10, color=INK)
            ws.cell(row=r, column=2).alignment = LEFT
        r += 1


def _write_reference_sheet(ws):
    """Lookups Manufacturing can copy line/plant/role codes from."""
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Reference — valid codes"
    ws["A1"].font = Font(name="Calibri", size=15, bold=True, color=NAVY)
    ws["A2"] = "Copy values from here into the Data Entry sheet so they match the masterdata exactly."
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color=INK3)

    blocks = [
        ("Lines (line_code)", LINES),
        ("Plants (plant_code)", PLANTS),
        ("Line roles (role_code, optional)", LINE_ROLES),
        ("Plant-shared roles (role_code, required for plant rows)", PLANT_ROLES),
    ]
    col = 1
    for title, items in blocks:
        ws.cell(row=4, column=col, value=title).font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        ws.column_dimensions[get_column_letter(col)].width = max(24, len(title) + 4)
        for i, item in enumerate(items, start=5):
            cell = ws.cell(row=i, column=col, value=item)
            cell.font = Font(name="Calibri", size=10, color=INK)
            cell.fill = SAMPLE_FILL
        col += 1


def _add_named_range(wb, name: str, sheet_name: str, col_letter: str, first_row: int, last_row: int):
    """Define a workbook-level named range for use in data validation."""
    from openpyxl.workbook.defined_name import DefinedName
    ref = f"'{sheet_name}'!${col_letter}${first_row}:${col_letter}${last_row}"
    wb.defined_names[name] = DefinedName(name=name, attr_text=ref)


def _set_dropdown(ws, col_letter: str, last_row: int, named_range: str | None = None,
                  explicit_list: list[str] | None = None):
    """Apply a data validation dropdown to col_letter rows 3..last_row.
    Use a named range when the list is long; explicit list for short ones."""
    if explicit_list is not None:
        formula = '"' + ",".join(explicit_list) + '"'
    else:
        formula = "=" + named_range
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Pick a value from the dropdown (matches the Reference sheet)."
    dv.errorTitle = "Invalid value"
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}3:{col_letter}{last_row}")


# ─── Workbook 1: Headcount Exceptions Input ────────────────────────────────
def build_headcount_exceptions_input(out_path: str):
    wb = Workbook()

    # Instructions
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    _write_instruction_sheet(ws_instr,
        "Headcount Exceptions — input from Manufacturing",
        [
            ("Purpose",
             "Tell the planner about known absences for the planning month — annual leave, sickness, training, "
             "long-term sick. Each event reduces the planned headcount; the planner folds these into the main "
             "headcount_plan.xlsx file (Sheet 3 — Exceptions) before publishing the RCCP batch."),
            ("", ""),
            ("How to fill in", ""),
            ("1. Open the 'Data Entry' sheet.",
             "Two sample rows are pre-filled — replace them with real events."),
            ("2. Use ONE of line_code OR plant_code, not both.",
             "Use line_code for events that affect a specific production line (operators / team leaders). "
             "Use plant_code for events that affect a plant-shared role (forklift, materials handler, etc.)."),
            ("3. role_code (optional for LINE rows, required for PLANT rows).",
             "For line events, leave blank to apply the delta across all line roles. "
             "For plant events, pick the specific role from the Reference sheet."),
            ("4. Dates in DD/MM/YYYY format.",
             "start_date and end_date are inclusive. For a one-day event, set them the same."),
            ("5. delta_headcount — negative for absences.",
             "−1 means one person out. Use decimals if part-time (e.g. −0.5)."),
            ("6. reason — free text.",
             "Brief description that will surface on the People Fit panel: 'Annual leave', 'Long-term sick — Joe Smith', etc."),
            ("", ""),
            ("Returning the file",
             "Save and send back to the planner. The Reference sheet lists every valid line, plant and role code."),
            ("", ""),
            ("Quick examples", ""),
            ("Annual leave for one operator at A101",
             "line_code=A101, role_code=(blank), start=15/05/2026, end=19/05/2026, delta=−1, reason='Annual leave'"),
            ("One forklift driver out for the month in Plant 1",
             "plant_code='Plant 1', role_code='FORKLIFT_DRIVER', start=01/06/2026, end=30/06/2026, delta=−1, reason='Long-term sick'"),
        ])

    # Reference
    _write_reference_sheet(wb.create_sheet("Reference"))

    # Set up the named ranges from Reference for dropdowns
    ref_name = "Reference"
    _add_named_range(wb, "LineList", ref_name, "A", 5, 5 + len(LINES) - 1)
    _add_named_range(wb, "PlantList", ref_name, "B", 5, 5 + len(PLANTS) - 1)
    _add_named_range(wb, "AllRoleList", ref_name, "D", 5, 5 + len(PLANT_ROLES) - 1)

    # Data Entry sheet
    ws = wb.create_sheet("Data Entry")
    ws.sheet_view.showGridLines = False
    columns = [
        ("line_code",        "Required for LINE-role events. Leave blank for plant-shared events.",                  14),
        ("plant_code",       "Required for plant-shared role events. Leave blank for line events.",                  14),
        ("resource_type_code", "Role code. Optional for LINE rows (blank = applies to all line roles). Required for PLANT rows.", 24),
        ("start_date",       "First affected date (DD/MM/YYYY).",                                                    14),
        ("end_date",         "Last affected date inclusive (DD/MM/YYYY). Same as start for a one-day event.",        14),
        ("delta_headcount",  "Change vs the standard headcount. Negative for absences (e.g. −1 = one person out).",  18),
        ("reason",           "Free text: annual leave, sickness, training, etc. Surfaces on the People Fit panel.", 36),
    ]
    _write_header_row(ws, row=1, columns=columns)

    # Sample rows (rows 3–4)
    sample_rows = [
        ["A101", "",        "",                  "15/05/2026", "19/05/2026", -1,   "Annual leave (sample — replace or remove)"],
        ["",     "Plant 1", "FORKLIFT_DRIVER",   "01/06/2026", "12/06/2026", -1,   "Long-term sick (sample — replace or remove)"],
    ]
    for r_offset, row in enumerate(sample_rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_offset, column=c_idx, value=val)
            cell.font = SAMPLE_FONT
            cell.alignment = LEFT
            cell.border = BORDER_BOTTOM

    ws.freeze_panes = "A3"

    # Dropdowns (covering rows 3–200 to leave plenty of empty space)
    LAST = 200
    _set_dropdown(ws, "A", LAST, named_range="LineList")
    _set_dropdown(ws, "B", LAST, named_range="PlantList")
    _set_dropdown(ws, "C", LAST, explicit_list=ALL_ROLES)

    wb.save(out_path)


# ─── Workbook 2: Line Capacity Exceptions Input ─────────────────────────────
def build_line_capacity_exceptions_input(out_path: str):
    wb = Workbook()

    ws_instr = wb.active
    ws_instr.title = "Instructions"
    _write_instruction_sheet(ws_instr,
        "Line Capacity Exceptions — input from Manufacturing",
        [
            ("Purpose",
             "Tell the planner about planned events that reduce a production line's available hours for the "
             "planning month — maintenance, planned downtime, full shutdowns. The planner uses these to update "
             "the main line_capacity_calendar.xlsx file (per day per line) before publishing the RCCP batch."),
            ("", ""),
            ("How to fill in", ""),
            ("1. Open the 'Data Entry' sheet.",
             "Two sample rows are pre-filled — replace them with real events."),
            ("2. line_code — required.",
             "Pick the production line from the dropdown. See the Reference sheet for the full list."),
            ("3. event_type — pick from the dropdown.",
             "Maintenance · Planned downtime · Public holiday · Full shutdown · Other."),
            ("4. Dates in DD/MM/YYYY format.",
             "start_date and end_date are inclusive. For a one-day event, set them the same."),
            ("5. hours_lost_per_day.",
             "How many hours are lost per affected day. For a full shutdown use the line's normal shift hours (e.g. 7). "
             "For a 2-hour maintenance window use 2.0."),
            ("6. reason — free text.",
             "Brief description: 'Quarterly maintenance', 'Tool changeover', 'Annual cleaning shutdown', etc."),
            ("", ""),
            ("Returning the file",
             "Save and send back to the planner. The Reference sheet lists every valid line code."),
            ("", ""),
            ("Quick examples", ""),
            ("A 4-hour maintenance on A101 one day",
             "line_code=A101, event_type='Maintenance', start=15/06/2026, end=15/06/2026, hours_lost_per_day=4.0, reason='Quarterly PM'"),
            ("A 5-day full shutdown on A103",
             "line_code=A103, event_type='Full shutdown', start=22/06/2026, end=26/06/2026, hours_lost_per_day=7.0, reason='Annual cleaning'"),
            ("A tool changeover slot on A201",
             "line_code=A201, event_type='Planned downtime', start=10/06/2026, end=10/06/2026, hours_lost_per_day=2.5, reason='Tool changeover'"),
        ])

    _write_reference_sheet(wb.create_sheet("Reference"))

    _add_named_range(wb, "LineList", "Reference", "A", 5, 5 + len(LINES) - 1)

    ws = wb.create_sheet("Data Entry")
    ws.sheet_view.showGridLines = False
    columns = [
        ("line_code",            "Production line affected (pick from the dropdown).",                     14),
        ("event_type",           "Maintenance · Planned downtime · Public holiday · Full shutdown · Other.", 22),
        ("start_date",           "First affected date (DD/MM/YYYY).",                                       14),
        ("end_date",             "Last affected date inclusive (DD/MM/YYYY).",                              14),
        ("hours_lost_per_day",   "How many hours are lost per affected day. e.g. 4.0 for a 4-hour PM, 7.0 for a full day off.", 22),
        ("reason",               "Free text — what's happening.",                                           40),
    ]
    _write_header_row(ws, row=1, columns=columns)

    sample_rows = [
        ["A101", "Maintenance",      "15/06/2026", "15/06/2026", 4.0, "Quarterly PM (sample — replace or remove)"],
        ["A103", "Full shutdown",    "22/06/2026", "26/06/2026", 7.0, "Annual cleaning (sample — replace or remove)"],
    ]
    for r_offset, row in enumerate(sample_rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_offset, column=c_idx, value=val)
            cell.font = SAMPLE_FONT
            cell.alignment = LEFT
            cell.border = BORDER_BOTTOM

    ws.freeze_panes = "A3"

    LAST = 200
    _set_dropdown(ws, "A", LAST, named_range="LineList")
    _set_dropdown(ws, "B", LAST, explicit_list=EVENT_TYPES)

    wb.save(out_path)


def main():
    out_dir = os.path.join(os.path.dirname(__file__), "..", "uploads")
    os.makedirs(out_dir, exist_ok=True)

    hc_path = os.path.abspath(os.path.join(out_dir, "headcount_exceptions_input.xlsx"))
    lc_path = os.path.abspath(os.path.join(out_dir, "line_capacity_exceptions_input.xlsx"))

    build_headcount_exceptions_input(hc_path)
    print(f"Saved: {hc_path}")
    build_line_capacity_exceptions_input(lc_path)
    print(f"Saved: {lc_path}")
    print()
    print("Send these to Manufacturing each cycle. They'll fill in the Data Entry sheet")
    print("(dropdowns + sample rows already in place) and send back.")
    print("Then fold the returned data into:")
    print("  - headcount_plan.xlsx · Sheet 3 (Exceptions)")
    print("  - line_capacity_calendar.xlsx · per-day rows for the affected dates")


if __name__ == "__main__":
    main()
