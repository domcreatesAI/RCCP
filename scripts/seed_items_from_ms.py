"""
Seed dbo.items from master_stock Excel file (uploads/ms.xlsx).

Reads unique material codes and their pack attributes, then MERGE-inserts
into dbo.items. Existing items are left unchanged. New items get:
  - plant_code = 'A1'  (default — update manually if different)
  - pack_size_l        from Volume column (first positive value per material)
  - units_per_pallet   from Rounding value column (first positive int per material)
  - everything else    NULL

Run from repo root:
    backend\\venv\\Scripts\\python.exe scripts\\seed_items_from_ms.py

Optional args:
    --file  PATH    path to .xlsx  (default: uploads/ms.xlsx)
    --plant CODE    default plant  (default: A1)
    --dry-run       print SQL only, do not insert
"""

import argparse
import os
import sys

import openpyxl
import pyodbc

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
DEFAULT_XLSX = os.path.join(os.path.dirname(__file__), "..", "uploads", "ms.xlsx")
DB_CONN = (
    "DRIVER={ODBC Driver 18 for SQL Server};"
    "SERVER=localhost\\SQLEXPRESS;"
    "DATABASE=RCCP_One;"
    "Trusted_Connection=yes;"
    "TrustServerCertificate=yes"
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_pos_float(val):
    """Return positive float or None."""
    if val is None:
        return None
    try:
        f = float(val)
        return f if f > 0 else None
    except (TypeError, ValueError):
        return None


def _safe_pos_int(val):
    """Return positive int or None."""
    f = _safe_pos_float(val)
    if f is None:
        return None
    return int(f) if f > 0 else None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Seed items from master_stock .xlsx")
    parser.add_argument("--file",    default=DEFAULT_XLSX, help="Path to ms.xlsx")
    parser.add_argument("--plant",   default="A1",         help="Default plant_code")
    parser.add_argument("--dry-run", action="store_true",  help="Print counts only, no DB writes")
    args = parser.parse_args()

    # ---- 1. Read Excel ----
    print(f"Reading: {args.file}")
    wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)
    ws = wb.active

    # Find column indices from header row (row 1 in ms.xlsx = plain headers)
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    col = {h: i for i, h in enumerate(headers)}

    required = {"Material", "Volume", "Rounding value"}
    missing = required - set(col.keys())
    if missing:
        print(f"ERROR: Missing columns in file: {missing}")
        sys.exit(1)

    mat_idx  = col["Material"]
    vol_idx  = col["Volume"]
    rv_idx   = col["Rounding value"]

    # Collect per-item attributes (first non-null wins)
    items: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        mat = row[mat_idx]
        if mat is None:
            continue
        item_code = str(mat).strip()
        if not item_code:
            continue
        if item_code not in items:
            items[item_code] = {"pack_size_l": None, "units_per_pallet": None}
        attrs = items[item_code]
        if attrs["pack_size_l"] is None:
            attrs["pack_size_l"] = _safe_pos_float(row[vol_idx])
        if attrs["units_per_pallet"] is None:
            attrs["units_per_pallet"] = _safe_pos_int(row[rv_idx])

    print(f"Unique material codes found: {len(items)}")

    if args.dry_run:
        sample = list(items.items())[:5]
        for code, attrs in sample:
            print(f"  {code}  pack_size_l={attrs['pack_size_l']}  units_per_pallet={attrs['units_per_pallet']}")
        print("Dry run — no DB writes.")
        return

    # ---- 2. Connect and MERGE ----
    print(f"Connecting to DB …")
    conn = pyodbc.connect(DB_CONN)
    cursor = conn.cursor()

    # Check default plant exists
    cursor.execute("SELECT COUNT(*) FROM dbo.plants WHERE plant_code = ?", args.plant)
    if cursor.fetchone()[0] == 0:
        print(f"ERROR: plant_code '{args.plant}' not found in dbo.plants. Use --plant <code>.")
        sys.exit(1)

    inserted = 0
    skipped  = 0
    for item_code, attrs in items.items():
        cursor.execute(
            """
            IF NOT EXISTS (SELECT 1 FROM dbo.items WHERE item_code = ?)
            BEGIN
                INSERT INTO dbo.items (item_code, plant_code, pack_size_l, units_per_pallet)
                VALUES (?, ?, ?, ?)
            END
            """,
            item_code,
            item_code, args.plant, attrs["pack_size_l"], attrs["units_per_pallet"],
        )
        if cursor.rowcount > 0:
            inserted += 1
        else:
            skipped += 1

    conn.commit()
    conn.close()

    print(f"Done. Inserted: {inserted}  |  Already existed (skipped): {skipped}")


if __name__ == "__main__":
    main()
