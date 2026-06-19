"""
Generate a pre-filled Line Capacity Calendar Excel file for upload to RCCP One.

Covers 01/01/2026 – 31/12/2030.
One row per production line per day (14 lines × ~1826 days = ~25,564 rows).

Shift pattern (planned_hours = on-site hours minus 1h allocated to breaks):
  - Mon–Thu: is_working_day = 1, planned_hours = 8.0  (9.0h on site − 1h breaks)
  - Fri:     is_working_day = 1, planned_hours = 5.5  (6.5h on site − 1h breaks)
  - Sat/Sun: is_working_day = 0, planned_hours = 0

UK bank holidays: is_working_day = 0, planned_hours = 0, labelled in downtime_reason.

downtime_hours + downtime_reason capture lost time on working days (maintenance,
breakdown, stock check, planned shutdown) and SUBTRACT from planned_hours.

Run from the repo root:
    python scripts/generate_capacity_calendar.py
Output: uploads/capacity_calendar_2026_2030.xlsx
"""

import os
import sys
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Canonical bank-holiday list + shift pattern live in the backend so the Tmpl
# download and this script never drift.
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))
from app.services.uk_calendar import UK_BANK_HOLIDAYS, PLANNED_HOURS_BY_WEEKDAY  # noqa: E402

# ---------------------------------------------------------------------------
# Production lines (all 14 active lines)
# ---------------------------------------------------------------------------
LINES = [
    "A101", "A102", "A103",
    "A201", "A202",
    "A302", "A303", "A304", "A305", "A307", "A308",
    "A401",
    "A501", "A502",
]

# PLANNED_HOURS_BY_WEEKDAY and UK_BANK_HOLIDAYS are imported from
# app.services.uk_calendar (single source of truth).

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_DESC_FILL   = PatternFill("solid", fgColor="FFE066")
_DESC_FONT   = Font(color="7A5700", italic=True, size=9)
_BH_FILL     = PatternFill("solid", fgColor="FFF3CD")   # amber tint for bank holidays
_WE_FILL     = PatternFill("solid", fgColor="F3F4F6")   # light grey for weekends
_CENTER      = Alignment(horizontal="center", vertical="center")
_LEFT        = Alignment(horizontal="left",   vertical="center")

COLUMNS = [
    "line_code",
    "calendar_date",
    "is_working_day",
    "planned_hours",
    "downtime_hours",
    "downtime_reason",
]

DESCRIPTIONS = [
    "Production line code (e.g. A101, A202). Must match a line in the masterdata.",
    "Date in DD/MM/YYYY format (e.g. 01/03/2026).",
    "1 = working day, 0 = non-working day (weekend, bank holiday).",
    "Scheduled production hours for this line on this date (0–24). Required.",
    "Hours lost this day (subtracts from planned_hours). Enter 0 if none.",
    "Why the line is down: Breakdown / Maintenance / Stock check / Planned shutdown. Required when downtime_hours > 0.",
]

COL_WIDTHS = {
    "line_code":             12,
    "calendar_date":         16,
    "is_working_day":        16,
    "planned_hours":         16,
    "downtime_hours":        16,
    "downtime_reason":       22,
}


def generate() -> None:
    start = date(2026, 1, 1)
    end   = date(2030, 12, 31)

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Row 1: field descriptions (amber — validator ignores this row entirely)
    for col_idx, desc in enumerate(DESCRIPTIONS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=desc)
        cell.font      = _DESC_FONT
        cell.fill      = _DESC_FILL
        cell.alignment = _LEFT

    # Row 2: column keys
    for col_idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col, 16)

    ws.freeze_panes = "A3"

    row_num = 3
    current = start
    while current <= end:
        weekday       = current.weekday()          # 0=Mon … 6=Sun
        is_weekend    = weekday >= 5
        is_bank_hol   = current in UK_BANK_HOLIDAYS
        is_working    = not is_weekend and not is_bank_hol

        planned_hours = PLANNED_HOURS_BY_WEEKDAY[weekday] if is_working else 0.0

        reason = ""
        if is_bank_hol:
            reason = "Bank holiday"
        elif is_weekend:
            reason = "Weekend"

        for line in LINES:
            ws.cell(row=row_num, column=1, value=line)
            ws.cell(row=row_num, column=2, value=current.strftime("%d/%m/%Y"))
            ws.cell(row=row_num, column=3, value=1 if is_working else 0)
            ws.cell(row=row_num, column=4, value=planned_hours)
            ws.cell(row=row_num, column=5, value=0)          # downtime_hours
            ws.cell(row=row_num, column=6, value=reason)     # downtime_reason (label for non-working)

            if is_bank_hol:
                for c in range(1, 7):
                    ws.cell(row=row_num, column=c).fill = _BH_FILL
            elif is_weekend:
                for c in range(1, 7):
                    ws.cell(row=row_num, column=c).fill = _WE_FILL

            row_num += 1

        current += timedelta(days=1)

    # Info sheet
    ws2 = wb.create_sheet("Info")
    ws2.column_dimensions["A"].width = 70
    info = [
        "RCCP One — Line Capacity Calendar 2026–2030",
        "",
        f"Generated: {date.today().strftime('%d/%m/%Y')}",
        f"Lines: {len(LINES)}",
        f"Date range: 01/01/2026 – 31/12/2030",
        f"Total rows: {row_num - 3:,}",
        "",
        "Shift pattern applied (planned_hours = on-site hours minus 1h breaks):",
        "  Mon–Thu:  8.0 hours / day (9.0h on site − 1h breaks)",
        "  Friday:   5.5 hours / day (6.5h on site − 1h breaks)",
        "  Sat/Sun:  non-working     (is_working_day = 0, planned_hours = 0)",
        "",
        "UK bank holidays: is_working_day=0, planned_hours=0 (labelled in downtime_reason)",
        "",
        "Amber rows = UK bank holidays",
        "Grey rows  = weekends",
        "",
        "downtime_hours subtracts from planned_hours (available = planned − downtime).",
        "downtime_reason: Breakdown / Maintenance / Stock check / Planned shutdown.",
        "",
        "BEFORE UPLOADING — review and adjust:",
        "  - downtime_hours + downtime_reason for maintenance / shutdown windows",
        "  - Any site-specific shutdowns (e.g. factory closure weeks)",
        "  - Lines with different shift patterns (adjust planned_hours per row)",
    ]
    for i, line in enumerate(info, start=1):
        cell = ws2.cell(row=i, column=1, value=line)
        cell.font      = Font(bold=(i == 1), size=12 if i == 1 else 10)
        cell.alignment = _LEFT

    out_dir = os.path.join(os.path.dirname(__file__), "..", "uploads")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "capacity_calendar_2026_2030.xlsx")

    wb.save(out_path)
    print(f"Saved: {os.path.abspath(out_path)}")
    print(f"Rows:  {row_num - 3:,}  ({len(LINES)} lines × {(end - start).days + 1} days)")
    print(f"Shift: Mon–Thu 8.0h | Fri 5.5h | Sat–Sun 0h (on-site minus 1h breaks)")


if __name__ == "__main__":
    generate()
