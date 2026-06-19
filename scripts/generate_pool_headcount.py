"""
Generate the Pool Headcount plan with the ACTUAL pool numbers, held flat across
every month to the end of 2030 (headcount is stable; variations go on Exceptions).

Edit POOL_HEADCOUNT below if the standing numbers change, then re-run.

Output: uploads/headcount_plan_2026_2030.xlsx
  Sheet 1 "Pool Headcount" — pool_code, resource_type_code, plan_month, planned_headcount
  Sheet 2 "Exceptions"     — absences (header only; you maintain these)
"""

import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── ACTUAL standing headcount per pool per role (edit here if it changes) ──────
POOL_HEADCOUNT: dict[str, dict[str, float]] = {
    "POOL-FLEX": {            # Plants 1 / 3 / 4 — flexible crew
        "LINE_OPERATOR":        14,
        "LINE_LEADER":          4,
        "PALLETISING_OPERATOR": 1,
        "FORKLIFT_DRIVER":      2,    # site-shared: 2 drivers serve ALL plants (1/2/3/4)
        "MATERIAL_HANDLER":     2,    # site-shared: 2 material handlers serve ALL plants
        "ROBOT_OPERATOR":       1,
        "TECHNICIAN":           1,
    },
    "POOL-P2": {              # Plant 2 — dedicated line crew only (no palletising).
        "LINE_OPERATOR":        1,
        "LINE_LEADER":          1,
        # No palletising at Plant 2 (not required on A201/A202).
        # Forklift / material handler are SITE-shared (the 2 each held in POOL-FLEX
        # serve all 4 plants), so Plant 2 carries none of its own here.
    },
}

# Hold flat from the current plan-cycle month to end of 2030.
START_YEAR, START_MONTH = 2026, 7
END_YEAR,   END_MONTH   = 2030, 12

_HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_DESC_FILL   = PatternFill("solid", fgColor="FFE066")
_DESC_FONT   = Font(color="7A5700", italic=True, size=9)
_CENTER      = Alignment(horizontal="center", vertical="center")
_LEFT        = Alignment(horizontal="left",   vertical="center")


def _num(v):
    return int(v) if float(v).is_integer() else float(v)


def months():
    out, y, m = [], START_YEAR, START_MONTH
    while (y, m) <= (END_YEAR, END_MONTH):
        out.append(date(y, m, 1))
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return out


def _sheet(wb, name, is_first, columns, rows):
    ws = wb.active if is_first else wb.create_sheet(name)
    ws.title = name
    for ci, (_k, desc, _w) in enumerate(columns, 1):
        c = ws.cell(row=1, column=ci, value=desc); c.font, c.fill, c.alignment = _DESC_FONT, _DESC_FILL, _LEFT
    for ci, (k, _d, w) in enumerate(columns, 1):
        c = ws.cell(row=2, column=ci, value=k); c.font, c.fill, c.alignment = _HEADER_FONT, _HEADER_FILL, _CENTER
        ws.column_dimensions[get_column_letter(ci)].width = w
    for ri, row in enumerate(rows, 3):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    ws.freeze_panes = "A3"


def generate():
    ms = months()
    wb = Workbook()

    pool_cols = [
        ("pool_code",          "Labour pool (POOL-FLEX = Plants 1/3/4, POOL-P2 = Plant 2).",          14),
        ("resource_type_code", "Role (LINE_OPERATOR, LINE_LEADER, PALLETISING_OPERATOR, FORKLIFT_DRIVER, …).", 26),
        ("plan_month",         "1st of the month in DD/MM/YYYY.",                                      16),
        ("planned_headcount",  "People available in the pool for this role this month.",               20),
    ]
    # Month-major ordering: each month is a contiguous block of all pool×role rows
    # (pools in order, roles alphabetical) — easy to read and filter.
    rows = []
    for m in ms:
        for pool_code in sorted(POOL_HEADCOUNT):
            roles = POOL_HEADCOUNT[pool_code]
            for role in sorted(roles):
                rows.append([pool_code, role, m.strftime("%d/%m/%Y"), _num(roles[role])])
    _sheet(wb, "Pool Headcount", True, pool_cols, rows)

    exc_cols = [
        ("line_code",          "Production line (e.g. A101). For line-role absences; blank for plant-shared.", 12),
        ("plant_code",         "Plant (e.g. 'Plant 1'). For shared-role absences; blank for line.",            12),
        ("resource_type_code", "Role. Required for plant rows.",                                              24),
        ("start_date",         "First affected date DD/MM/YYYY.",                                             14),
        ("end_date",           "Last affected date DD/MM/YYYY (inclusive).",                                  14),
        ("delta_headcount",    "Change vs standing headcount. Negative for absences (e.g. -1).",              18),
        ("reason",             "Annual leave / sickness / training, etc.",                                   30),
    ]
    _sheet(wb, "Exceptions", False, exc_cols, [])

    wi = wb.create_sheet("Info")
    wi.column_dimensions["A"].width = 80
    info = [
        ("RCCP One — Pool Headcount (standing, flat to end 2030)", Font(bold=True, size=13)),
        ("", None),
        (f"Generated: {date.today().strftime('%d/%m/%Y')}", Font(size=10)),
        (f"Range: 01/{START_MONTH:02d}/{START_YEAR} – 12/{END_MONTH:02d}/{END_YEAR}  ({len(ms)} months)", Font(size=10)),
        (f"Pool Headcount rows: {len(rows):,}", Font(size=10)),
        ("", None),
        ("Numbers are held FLAT every month (headcount is stable). Manage day-to-day", Font(size=10)),
        ("variation — holiday, sickness, training, leavers — on the Exceptions sheet.", Font(size=10)),
        ("Re-run scripts/generate_pool_headcount.py if the standing numbers change.", Font(size=10)),
        ("", None),
        ("Standing headcount:", Font(bold=True, size=10)),
        *[(f"  {pc} · {rc}: {_num(hc)}", Font(size=10))
          for pc, roles in POOL_HEADCOUNT.items() for rc, hc in roles.items()],
    ]
    for i, (t, f) in enumerate(info, 1):
        c = wi.cell(row=i, column=1, value=t)
        if f:
            c.font = f
        c.alignment = _LEFT

    out_dir = os.path.join(os.path.dirname(__file__), "..", "uploads")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "headcount_plan_2026_2030.xlsx")
    wb.save(out_path)
    combos = sum(len(r) for r in POOL_HEADCOUNT.values())
    print(f"Saved: {os.path.abspath(out_path)}")
    print(f"  Pool Headcount: {len(rows):,} rows  ({combos} pool×role combos × {len(ms)} months)")


if __name__ == "__main__":
    generate()
