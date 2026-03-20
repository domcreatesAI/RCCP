"""
export_batch_data.py
--------------------
Exports all planning data for the most-recent PUBLISHED batch into a single
Excel workbook for manual verification.

Run from the repo root:
  cd c:/Claude/Moove/RCCP
  backend/venv/Scripts/python.exe scripts/export_batch_data.py

Output: documents_from_other_sources\batch_<id>_export.xlsx
"""

import os
import sys
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import pyodbc
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────

REPO_ROOT   = Path(__file__).resolve().parent.parent
ENV_FILE    = REPO_ROOT / "backend" / ".env"
OUT_DIR     = REPO_ROOT / "documents_from_other_sources"
OUT_DIR.mkdir(exist_ok=True)

load_dotenv(ENV_FILE)
DB_SERVER   = os.getenv("DB_SERVER", r"localhost\SQLEXPRESS")
DB_NAME     = os.getenv("DB_NAME",   "RCCP_One")
DB_USER     = os.getenv("DB_USER",   "rccp_app")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")

# ── DB connection ─────────────────────────────────────────────────────────────

def connect() -> pyodbc.Connection:
    conn_str = (
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={DB_SERVER};"
        f"DATABASE={DB_NAME};"
        f"UID={DB_USER};"
        f"PWD={DB_PASSWORD};"
    )
    return pyodbc.connect(conn_str)


def query(conn, sql: str, *params) -> pd.DataFrame:
    return pd.read_sql(sql, conn, params=list(params) if params else None)


# ── Helpers ───────────────────────────────────────────────────────────────────

def style_sheet(ws, freeze_row: int = 2):
    """Apply header styling and auto-width to a worksheet."""
    HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    CENTER      = Alignment(horizontal="center", vertical="center")
    LEFT        = Alignment(horizontal="left",   vertical="center")

    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER

    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0 for c in col_cells),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    if freeze_row:
        ws.freeze_panes = ws.cell(row=freeze_row, column=1)


def write_df(writer: pd.ExcelWriter, df: pd.DataFrame, sheet: str, note: str = ""):
    """Write a DataFrame to a sheet and apply styling."""
    if df.empty:
        df = pd.DataFrame({"(no rows)": []})

    # Insert a note row if provided
    if note:
        note_df = pd.DataFrame([{col: "" for col in df.columns}])
        note_df.iloc[0, 0] = f"ℹ {note}"
        df = pd.concat([note_df, df], ignore_index=True)

    df.to_excel(writer, sheet_name=sheet, index=False)
    ws = writer.sheets[sheet]
    style_sheet(ws)

    if note:
        ws.cell(row=2, column=1).font = Font(italic=True, color="7A5700", size=9)


# ── Main export ───────────────────────────────────────────────────────────────

def main():
    conn = connect()
    print("Connected to DB.")

    # 1. Find published batch
    batch_df = query(conn, """
        SELECT TOP 1
            batch_id, batch_name, plan_cycle_date, status,
            created_by, created_at, published_at, published_by, notes
        FROM dbo.import_batches
        WHERE status = 'PUBLISHED'
        ORDER BY published_at DESC
    """)

    if batch_df.empty:
        print("No PUBLISHED batch found. Trying most-recent batch instead.")
        batch_df = query(conn, """
            SELECT TOP 1
                batch_id, batch_name, plan_cycle_date, status,
                created_by, created_at, published_at, published_by, notes
            FROM dbo.import_batches
            ORDER BY created_at DESC
        """)

    if batch_df.empty:
        print("No batches found. Exiting.")
        sys.exit(1)

    batch_id   = int(batch_df.iloc[0]["batch_id"])
    batch_name = str(batch_df.iloc[0]["batch_name"])
    print(f"Exporting batch {batch_id}: {batch_name}")

    out_path = OUT_DIR / f"batch_{batch_id}_export.xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:

        # ── Sheet 1: Batch summary ─────────────────────────────────────────
        write_df(writer, batch_df, "0_Batch Info")

        # ── Sheet 2: Validation results ────────────────────────────────────
        val_df = query(conn, """
            SELECT
                ibf.file_type, ivr.validation_stage, ivr.stage_name,
                ivr.severity, ivr.message, ivr.created_at
            FROM dbo.import_validation_results ivr
            JOIN dbo.import_batch_files ibf ON ibf.batch_file_id = ivr.batch_file_id
            WHERE ibf.batch_id = ?
            ORDER BY ibf.file_type, ivr.validation_stage
        """, batch_id)
        write_df(writer, val_df, "1_Validation Results",
                 "Validation results for all files in this batch")

        # ── Sheet 3: Production Orders ─────────────────────────────────────
        po_df = query(conn, """
            SELECT
                po.sap_order_number, po.item_code, i.item_description,
                po.order_type, po.plant_code, po.production_line,
                po.basic_start_date,
                po.order_quantity, po.delivered_quantity, po.net_quantity,
                po.uom, po.system_status, po.mrp_controller, po.source_row_number
            FROM dbo.production_orders po
            LEFT JOIN dbo.items i ON i.item_code = po.item_code
            WHERE po.batch_id = ?
            ORDER BY po.basic_start_date, po.production_line, po.item_code
        """, batch_id)
        write_df(writer, po_df, "2_Production Orders",
                 f"{len(po_df)} rows — net_quantity = MAX(0, order_qty - delivered_qty)")

        # ── Sheet 4: Demand Plan ───────────────────────────────────────────
        dem_df = query(conn, """
            SELECT
                dp.item_code, i.item_description,
                dp.warehouse_code, dp.period_type, dp.demand_type,
                dp.period_start_date, dp.period_end_date, dp.demand_quantity
            FROM dbo.demand_plan dp
            LEFT JOIN dbo.items i ON i.item_code = dp.item_code
            WHERE dp.batch_id = ?
            ORDER BY dp.item_code, dp.period_start_date
        """, batch_id)
        write_df(writer, dem_df, "3_Demand Plan",
                 f"{len(dem_df)} rows — monthly demand in eaches")

        # ── Sheet 5: Line Capacity Calendar ───────────────────────────────
        lcc_df = query(conn, """
            SELECT
                lcc.line_code, lcc.calendar_date, lcc.is_working_day,
                lcc.standard_hours, lcc.planned_hours, lcc.maintenance_hours,
                lcc.public_holiday_hours, lcc.planned_downtime_hours,
                lcc.other_loss_hours, lcc.notes
            FROM dbo.line_capacity_calendar lcc
            WHERE lcc.batch_id = ?
            ORDER BY lcc.line_code, lcc.calendar_date
        """, batch_id)
        write_df(writer, lcc_df, "4_Capacity Calendar",
                 f"{len(lcc_df)} rows")

        # ── Sheet 6: Headcount Plan ────────────────────────────────────────
        hc_df = query(conn, """
            SELECT
                hp.line_code, hp.plan_date, hp.planned_headcount,
                hp.shift_code, hp.available_hours, hp.notes
            FROM dbo.headcount_plan hp
            WHERE hp.batch_id = ?
            ORDER BY hp.line_code, hp.plan_date
        """, batch_id)
        write_df(writer, hc_df, "5_Headcount Plan",
                 f"{len(hc_df)} rows — line-level operator headcount")

        # ── Sheet 7: Plant Headcount Plan ─────────────────────────────────
        php_df = query(conn, """
            SELECT
                php.plant_code, php.resource_type_code,
                php.plan_date, php.planned_headcount
            FROM dbo.plant_headcount_plan php
            WHERE php.batch_id = ?
            ORDER BY php.plant_code, php.resource_type_code, php.plan_date
        """, batch_id)
        write_df(writer, php_df, "6_Plant HC Plan",
                 f"{len(php_df)} rows — plant-level shared roles headcount")

        # ── Sheet 8: Master Stock ──────────────────────────────────────────
        ms_df = query(conn, """
            SELECT
                ms.item_code, i.item_description,
                ms.warehouse_code, ms.total_stock_ea,
                ms.free_stock_ea, ms.safety_stock_ea,
                ms.snapshot_date
            FROM dbo.master_stock ms
            LEFT JOIN dbo.items i ON i.item_code = ms.item_code
            WHERE ms.batch_id = ?
            ORDER BY ms.item_code, ms.warehouse_code
        """, batch_id)
        write_df(writer, ms_df, "7_Master Stock",
                 f"{len(ms_df)} rows")

        # ── Sheet 9: Portfolio Changes ─────────────────────────────────────
        pc_df = query(conn, """
            SELECT
                pc.change_type, pc.effective_date, pc.item_code,
                i.item_description, pc.description, pc.impact_notes,
                pc.initial_demand
            FROM dbo.portfolio_changes pc
            LEFT JOIN dbo.items i ON i.item_code = pc.item_code
            WHERE pc.batch_id = ?
            ORDER BY pc.effective_date
        """, batch_id)
        write_df(writer, pc_df, "8_Portfolio Changes",
                 f"{len(pc_df)} rows")

        # ── Sheet 10: RCCP Engine — Line monthly buckets ──────────────────
        # Run the engine and flatten to a DataFrame
        print("  Running RCCP engine...")
        sys.path.insert(0, str(REPO_ROOT / "backend"))
        from app.services.rccp_engine import compute_dashboard  # noqa: E402
        dashboard = compute_dashboard(conn, batch_id)

        rows = []
        for line in dashboard["lines"]:
            for m in line["monthly"]:
                rows.append({
                    "line_code":           line["line_code"],
                    "plant_code":          line["plant_code"],
                    "risk_status":         line["risk_status"],
                    "primary_driver":      line["primary_driver"],
                    "labour_status":       line["labour_status"],
                    "period":              m["period"],
                    "working_days":        m["working_days"],
                    "available_litres":    m["available_litres"],
                    "demand_litres":       m["demand_litres"],
                    "firm_litres":         m["firm_litres"],
                    "planned_litres":      m["planned_litres"],
                    "production_litres":   m["production_litres"],
                    "gap_litres":          m["gap_litres"],
                    "available_hours":     m["available_hours"],
                    "demand_hours":        m["demand_hours"],
                    "firm_hours":          m["firm_hours"],
                    "planned_hours":       m["planned_hours"],
                    "production_hours":    m["production_hours"],
                    "gap_hours":           m["gap_hours"],
                    "utilisation_pct":     m["utilisation_pct"],
                    "hc_required":         m["hc_required"],
                    "hc_planned_avg":      m["hc_planned_avg"],
                    "hc_shortfall":        m["hc_shortfall"],
                    "labour_status_month": m["labour_status"],
                })
        eng_df = pd.DataFrame(rows)
        write_df(writer, eng_df, "9_RCCP Engine Output",
                 "Computed monthly buckets per line — used directly by the dashboard")

        # ── Sheet 11: RCCP KPIs ────────────────────────────────────────────
        kpis = dashboard["kpis"]
        kpi_df = pd.DataFrame([{
            "metric":   k,
            "value":    v,
        } for k, v in kpis.items()])
        write_df(writer, kpi_df, "10_KPIs",
                 "Top-level KPI tiles shown on the dashboard")

        # ── Sheet 12: Unassigned orders ────────────────────────────────────
        if dashboard.get("unassigned_orders"):
            uo_df = pd.DataFrame(dashboard["unassigned_orders"])
            write_df(writer, uo_df, "11_Unassigned Orders",
                     "SKUs with no primary_line_code — excluded from utilisation")
        else:
            write_df(writer, pd.DataFrame({"result": ["No unassigned orders"]}),
                     "11_Unassigned Orders")

        # ── Sheet 13: Plant resource requirements ──────────────────────────
        pr_df = query(conn, """
            SELECT
                prr.plant_code, prr.resource_type_code,
                rt.resource_type_name, rt.scope,
                prr.headcount_required
            FROM dbo.plant_resource_requirements prr
            JOIN dbo.resource_types rt ON rt.resource_type_code = prr.resource_type_code
            ORDER BY prr.plant_code, prr.resource_type_code
        """)
        write_df(writer, pr_df, "12_Plant Resource Reqs",
                 "Static plant-level headcount requirements from masterdata")

        # ── Sheet 14: Line resource requirements ──────────────────────────
        lr_df = query(conn, """
            SELECT
                lrr.line_code, lrr.resource_type_code,
                rt.resource_type_name, rt.scope,
                lrr.headcount_required
            FROM dbo.line_resource_requirements lrr
            JOIN dbo.resource_types rt ON rt.resource_type_code = lrr.resource_type_code
            ORDER BY lrr.line_code, lrr.resource_type_code
        """)
        write_df(writer, lr_df, "13_Line Resource Reqs",
                 "Static per-line headcount requirements from masterdata")

    conn.close()
    print(f"\nDONE: {out_path}")
    print(f"  Sheets written:")
    print("    0  Batch Info")
    print("    1  Validation Results")
    print("    2  Production Orders")
    print("    3  Demand Plan")
    print("    4  Capacity Calendar")
    print("    5  Headcount Plan")
    print("    6  Plant HC Plan")
    print("    7  Master Stock")
    print("    8  Portfolio Changes")
    print("    9  RCCP Engine Output  (computed monthly buckets)")
    print("   10  KPIs")
    print("   11  Unassigned Orders")
    print("   12  Plant Resource Requirements")
    print("   13  Line Resource Requirements")


if __name__ == "__main__":
    main()
